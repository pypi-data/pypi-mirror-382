from orgparse import load, loads
import sys
import random
import argparse
import copy
from openpyxl import Workbook

cat_str = 'CATEGORY'
num_str = 'NUM_RESP'
num_cor = 'NUM_CORRECT'

class Question :
    def __init__(self, org_node, group, num) :
        # the question 
        self.q = org_node
        # the question belongs to the following categories
        if org_node.properties.get(cat_str) :
            self.categories = org_node.properties[cat_str].split()
        else :
            self.categories = []
        # the percentage of correct responses this question got
        # (a measure of how easy it is)
        if org_node.properties.get(num_cor) and org_node.properties.get(num_str) :
            self.rate = float(org_node.properties[num_cor])/float(org_node.properties[num_str])
        else :
            # if no responses yet, then I give an average score of 0.5
            self.rate = 0.5
        self.group = group
        self.num = num


def parse_arguments() :
    """ Parse the command line arguments """
    parser = argparse.ArgumentParser(description='Randomly generates list of questions for exames from an org-mode file',
                                     prog='qgen.py',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter,
                                     fromfile_prefix_chars='@')
    parser.add_argument('-d', '--db', 
                        help='the org file containing all questions',
                        default='db.org',
                        required=False)

    parser.add_argument('-t', '--title',
                        help='The title of each page',
                        default='Examen',
                        required=False)
    
    parser.add_argument('-i', '--ifile', 
                        help='Text file containing the instructions to be printed on each exam',
                        default='none',
                        required=False)
    
    parser.add_argument('outfile',
                        help='Output file')
    
    parser.add_argument('-n', '--ncopies',
                        type=int,
                        help='Number of copies to generate',
                        default=1,
                        required=False)
    
    parser.add_argument('-g', '--ng',
                        help='Number of questions per group',
                        type=int,
                        nargs='*',
                        default=[], #1, 1, 1],
                        required=False)
    
    parser.add_argument('-e', '--header',
                        help='Org-mode file header',
                        default='none',
                        required=False)
    
    return parser.parse_args()




def print_header(header, out) :
    for l in header :
        out.write(l)
    out.write('\\pagestyle{empty}\n\n')
    out.write('\\thispagestyle{empty}\n\n')


def intersection(s1, s2) :
    return list(set(s1) & set(s2))
    
    
def generate_questionnaire(q_groups, ng) :
    """
       This function generates one single questionnaire, by selecting a
       number of questions from each group, as specified in list ng
    """
    qql = []          # the list of generated questions to be returned
    tag_list = []     # list of tags already used in the previous questions  

    # for each group, it will generate nq
    for qs, nq in zip(q_groups, ng) :
        qs1 = copy.deepcopy(qs)
        # only selects the question that do not contain categories already selected before
        qs2 = [ x for x in qs1 if intersection(x.categories, tag_list) == [] ]
        if len(qs2) < nq :
            print("Error, too few categories left")
            sys.exit(-1)
            
        # selects nq questions from qs2
        ql = random.sample(qs2, k=nq)
        qql = qql + ql
        # adds the questions
        for q in ql :
            tag_list = tag_list + q.categories
        
    return qql


def generate_allcopies(question_groups, ncopies, ng) :
    nq = sum(ng)
    print("{} questions per questionnaire".format(nq)) 

    all_copies = []
    # the start of each questionnaire
    for exam in range(ncopies) :
        # randomly generates the questions for each questionnaire
        qlist = generate_questionnaire(question_groups, ng)
        all_copies.append(qlist)

    return all_copies


def output_exam(out, HEADER, title, ifile, all_copies) :
    print_header(HEADER, out)
    ncopies = len(all_copies)
    if ifile != 'none' :
        instructions = open(ifile, 'r')
        ilines = instructions.readlines()
    else:
        ilines = ""
        
    for exam in range(ncopies) :
        out.write('* ' + title + '\n')
        out.write('- N: ' + str(exam+1) + '\n')
        if ifile != 'none' :            
            out.write('** Instructions\n')
            for x in ilines:
                out.write(x)
        print_questions(all_copies[exam], out)
    


def print_questions(qlist, out) :
    for qx in qlist :
        out.write('** ')
        out.write(qx.q.heading)
        out.write('\n')
        out.write(qx.q.body)
        out.write('\n\n')
    # We generate an empty second-level heading before a pagebreak.  
    # This last part is to avoid strange behaviours (like printing the solution)
    # in the generated org-mode file. 
    out.write('** \n\n')
    # out.write('\\pagebreak\n\n\n\n')
    out.write('\\cleardoublepage\n\n')


def read_questions(filename) :
    # open the org file 
    print("Opening file", filename)
    # the top node 
    root = load(filename)

    # a list of lists of questions 
    question_groups = []
    i = 1
    # creates the questions
    for g in root.children :
        qql = []
        j = 1
        for node in g.children :
            q = Question(node, i, j)
            qql.append(q)
            j+=1
        print("Found", len(qql), "questions in group", i)
        i+=1
        question_groups.append(qql)    

    return question_groups
    

    
def create_spreadsheet(nq, all_copies, outfile) :
    """Generates the xls file for grading.
        Every spreadsheet has three sheets : 
        - first one for grading
        - second one contains the questions id, used for updating the database
        - third one contains the weigth for each question 

        In every sheet, the first line is the header of the table,
        second one contains the normalised score per question (max grade = 5 per question).
        From the third row, the grading starts for each question.  
        
        Two ways to compute the total grade: with or without weights. In the first case, 
        grades are computed so that the final grade is in the range 0-20 (French grading system). 
        In the second case, the grades are additionally weigthed by the question specific weigth. 

        Such grade is an indication of the difficulty of answering the
        question, so tough questions are weighted less. This is an
        experimental feature, do not use right now.
    """
    
    wb = Workbook()
    # first sheet contains notes
    ws1 = wb.active
    ws1.title = "Notes"
    # the second sheet contains the questions (group-number)
    ws2 = wb.create_sheet("Questions")
    # the thirs sheet contains the weights (experimental)
    ws3 = wb.create_sheet("Weights")
    
    header = ["N. questionnaire", "Identifiant étudiant", "Nom", "Prenom"]
    for x in range(nq) :
        header.append("q{}".format(x+1))
        
    ws2.append(header)
    ws3.append(header)
    header.append("Total")
    ws1.append(header)
    
    i = 1
    for c in all_copies :
        # first sheet 
        row = ["{}".format(i), "", "", ""]
        startcol = 'D'
        endcol = 'C'
        for q in c :
            row.append("")
            endcol = chr(ord(endcol)+1)
        row.append("=sum({}{}:{}{})".format(startcol, i+1, endcol,i+1))
        ws1.append(row)

        # second sheet
        row = ["{}".format(i), "", "", ""]
        for q in c :
            row.append("{}-{}".format(q.group, q.num))
            
        ws2.append(row)

        # third sheet
        row = ["{}".format(i), "", "", ""]
        for q in c :
            row.append("{}".format(q.rate))
        ws3.append(row)

        i+=1

    if ".org" in outfile :
        outfile = outfile.replace("org", "xls")
        print("Outfile is now", outfile)
    else :
        outfile += ".xls"
    print("Writing the excel file into", outfile)
    wb.save(outfile)
    
    

## BEGINNING OF THE SCRIPT ## 

def main() :    
    #
    # the standard org-mode header
    # this can be changed with the -e (--header) option
    # 
    HEADER = ["#+startup: overview\n",
              "#+options: num:nil ^:nil toc:nil\n",
              "#+LATEX_CLASS: article\n",
              "#+LATEX_CLASS_OPTIONS: [a4paper,11pt,twoside]\n",
              "#+LATEX_HEADER: \\usepackage[T1]{fontenc}\n",
              "#+LATEX_HEADER: \\usepackage[textwidth=18cm, textheight=22.5cm]{geometry}\n",
              "#+latex_header: \\usepackage{ifthen,changepage}\n",
              "#+exclude_tags: solution noexport\n"]

    # parse the arguments into object options 
    options = parse_arguments()

    # read all questions 
    question_groups = read_questions(options.db)
        
    # the output file 
    print("Output into ", options.outfile)
    out = open(options.outfile, 'w')

    # an alternative header file 
    if options.header != 'none' :
        fh = open(options.header)
        HEADER = fh.readlines();

    # an optional instruction file 

    # the groups are in the top-level headings
    ngroups = len(question_groups) 
    print("Found", ngroups, "groups of questions")

    if options.ng == [] :
        options.ng = [1] * ngroups
        
    # the number of groups should match the lenght of options.ng (a list containing
    # the number of question to be selected per each group). 
    if ngroups != len(options.ng) :
        print("Error : wrong number of groups")
        print("Consider specifying the --ng option")
        sys.exit(3)

    # outputs the header first 

    all_copies = generate_allcopies(question_groups, options.ncopies, options.ng)
    output_exam(out, HEADER, options.title, options.ifile, all_copies)
        
    print("Generated", options.ncopies, "exam copies into", options.outfile)

    create_spreadsheet(sum(options.ng), all_copies, options.outfile)
    
