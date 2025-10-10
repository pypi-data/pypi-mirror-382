from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill

# Based on Okabe-Ito palette
# Reasonable performance for most common forms of colourblindness
HEXCODES = {
    "green": "00009E73",
    "purple": "00CC79A7",
    "blue": "0056B4E9",
    "orange": "00E69F00",
    "yellow": "00F0E442",
    "red": "00FF0000",  # this one is only used for the heatmap, and isn't Okabe-Ito
}


HEATMAP_RULE = ColorScaleRule(
    start_type="num",
    start_value=0,
    start_color=HEXCODES["green"],
    mid_type="num",
    mid_value=1,
    mid_color=HEXCODES["yellow"],
    end_type="num",
    end_value=100,
    end_color=HEXCODES["red"],
)


ORANGE_FILL = PatternFill(
    start_color=HEXCODES["orange"],
    end_color=HEXCODES["orange"],
    fill_type="solid",
)


YELLOW_FILL = PatternFill(
    start_color=HEXCODES["yellow"],
    end_color=HEXCODES["yellow"],
    fill_type="solid",
)

# so that bump2version can auto-update this
VERSION = "1.2.8"

GUIDE_LINES = [
    ["Automated Metadata Validation Tool"],
    ["v" + VERSION],
    ["Welcome to the Automated Metadata Validation tabular output report!"],
    [
        'This report may be accompanied by a "commented copy" of the original form, with validation errors highlighted and annoted in place.'
    ],
    [""],
    ["If you have an ONS Digital github account, you can view the main readme here:"],
    ["https://github.com/ONSdigital/automated-metadata-validation"],
    [""],
    ["Metadata variables are either mandatory or optional."],
    ['Validation checks are considered to be "hard" or "soft".'],
    [
        'Most checks care only about the value of single cells at a time. Checks that involve multiple values are called "comparative".'
    ],
    [""],
    [
        "Hard checks are conditions that must be satisfied for the ingest pipeline to work successfully. Failing a hard check means that action must be taken before the metadata form can be submitted. For example, a hard check may inspect whether a mandatory field contains an expected and machine-readable value, without leading or trailing whitespace."
    ],
    [""],
    [
        "Soft checks are checks that require inspection, but not necessarily action, if they fail. Though they are still important for high quality metadata, they won't prevent a minimum successful ingest. For example, a soft check may inspect the format and style of a field to ensure that it is useful and legible to humans."
    ],
    [""],
    [
        "Comparative checks involve more than one cell value at a time. For example, a column of table names might require that each name be unique within that column. Or, for consistency, a table name appearing on one sheet might be required to also appear on a list of tables from a previous sheet."
    ],
]


SHEET_DEETS = {
    "single_validation_checks": {
        "Colour": "green",
        "Variables": "All",
        "Checks": "Hard and Soft",
        "Description": "Summarises checks that care only about single cell values at a time. Details the checks employed during this run, pass or fail.",
    },
    "rowwise_validations": {
        "Colour": "green",
        "Variables": "All",
        "Checks": "Comparative",
        "Description": "Summarises comparative checks that look along whole rows of metadata. Details the checks employed during this run, pass or fail.",
    },
    "comparative_validations": {
        "Colour": "green",
        "Variables": "All",
        "Checks": "Comparative",
        "Description": "Summarises comparative checks that look across multiple tabs. Details the checks employed during this run, pass or fail.",
    },
    "applied_fixes": {
        "Colour": "green",
        "Variables": "All",
        "Checks": "See description",
        "Description": "Summarises the attempted automatic fixes applied to the metadata, if save_corrected_copy was enabled. For example, stripping leading and trailing whitespace. This will be reflected in the commented copy - you can copy and paste tables with fixed values back into your working template.",
    },
    "Short % overview": {
        "Colour": "purple",
        "Variables": "Mandatory only",
        "Checks": "Hard and Soft",
        "Description": "Each row is a tab & variable name combination; columns list the % of records that are missing or failed a) any hard or b) only soft checks.",
    },
    "Long % overview": {
        "Colour": "purple",
        "Variables": "Mandatory only",
        "Checks": "Hard and Soft",
        "Description": "Each row is a tab & variable name combination; columns are fail %s for every check. This tab is still in development.",
    },
    "Fails by value": {
        "Colour": "orange",
        "Variables": "All",
        "Checks": "Hard and Soft",
        "Description": "Each row details a value appearing in a variable, all the cells that value appears in, and all the hard and soft checks that value fails.",
    },
    "Fails by cell - Mandatory": {
        "Colour": "orange",
        "Variables": "Mandatory only",
        "Checks": "Hard and Soft",
        "Description": "Each row details the names of all hard and soft checks failed by a single cell in a mandatory field.",
    },
    "Fails by check - Hard": {
        "Colour": "orange",
        "Variables": "All",
        "Checks": "Hard",
        "Description": "Each row details the cells of a single variable that have failed a particular hard check.",
    },
    "Missing values": {
        "Colour": "blue",
        "Variables": "All mandatory",
        "Checks": "Missingness",
        "Description": "each row details the cells with missing values for a single variable.",
    },
    "Comparative check fails": {
        "Colour": "blue",
        "Variables": "All",
        "Checks": "Comparative",
        "Description": "Each row details one instance of a failed comparative check.",
    },
    "Fails by cell - Non mandatory": {
        "Colour": "yellow",
        "Variables": "Non-mandatory only",
        "Checks": "Hard and Soft",
        "Description": "Each row details the names of all hard and soft checks failed by a single cell in a non-mandatory (optional or conditional) field.",
    },
    "Fails by check - Soft": {
        "Colour": "yellow",
        "Variables": "All",
        "Checks": "Soft",
        "Description": "Each row details the cells of a single variable that have failed a particular soft check.",
    },
}
