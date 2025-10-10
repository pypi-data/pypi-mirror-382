import inspect
import math

# TODO: logging
import os
import re
import warnings
from datetime import datetime
from functools import reduce
from types import ModuleType
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

import ons_metadata_validation.outputs.output_constants as oc
import ons_metadata_validation.validation._validation_checks as vc
from ons_metadata_validation.outputs.field_constants import TEMPLATE_ORDER
from ons_metadata_validation.processing.processor import MetadataProcessor
from ons_metadata_validation.processing.utils import get_excel_ref_components

# suppresses warning for inconsequential data validation extension
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


#######################
"""Main function"""
#######################


# TODO: remove make_report and save_report from processor and hand them directly in here
def outputs_main(
    processor: MetadataProcessor,
    save_folder: Optional[str] = None,
    save_commented_copy: bool = True,
) -> Workbook:
    """Creates an output report wb, possibly saves it, and possibly saves a commented copy, too.

    Args:

        processor - the interface object produced by the main process of this module. To produce
        outputs, the following attributes are used at various points:

            processor.fails_df - the main dataframe where each row is one check failure.
            processor.template_map - the structural plan of metadata variables by tab.
            processor.md_filepath - the location of the original filled template.
            processor.recorded_validation_checks - a dict of dfs naming the checks applied to each metadata field.
            processor.tab_lens - the lengths of each tab of the filled template.
            processor.make_report - whether or not to create the dfs for the tabular report.
            processor.save_report - whether or not to save the tabular report after creating it.
            processor.save_corrected_copy - if True, and a commented copy is saved, it will include autofixes.

        save_folder - the location for saving outputs. If None, will default to same folder as the original template.

        save_commented_copy - whether or not to save a copy of the filled templating with coloured
        highlights and comment boxes for validation check fails.

    Returns:

        out_dict - the dictionary of dataframes that form the tabular report, if make_report is true.
                   Otherwise, this dictionary is empty.

    Notes:

    save_commented_copy is currently always set to True, since the interface is still
    under development.
    """

    if not any([processor.make_report, processor.save_report, save_commented_copy]):
        raise ValueError(
            "Outputs_main has triggered, but no output types have been requested:"
        )  # TODO add keys and values here

    # just the 2 or 3
    template_version = float(re.sub(r"[A-Za-z _]", "", processor.target_version))
    template_order = _generate_template_order(
        processor.template_map, TEMPLATE_ORDER[template_version]
    )

    if save_folder is None:  # by default, save to same folder as input file
        save_folder = os.path.dirname(processor.md_filepath)

    # produced iff save_corrected_copy is True. Needs some formatting.
    # ...and needs formatting here, because it's potentially used in both output files
    if "applied_fixes" in processor.recorded_validation_checks.keys():
        processor.recorded_validation_checks["applied_fixes"] = _format_fix_df(
            processor, template_order
        )

    out_dict = {}

    # accounting for a possible but obviously unintentional combination
    if processor.save_report:
        processor.make_report = True

    if processor.make_report:
        report_dfs = create_report_dfs(processor, template_order)
        out_dict = report_dfs  # a dictionary of dataframes

        if processor.save_report:
            report_wb = create_output_report(report_dfs)
            save_path = _make_save_path(
                processor.md_filepath, save_folder, mode="main_report"
            )
            report_wb.save(save_path)
            print(f"Saved to {save_path}!")

    if save_commented_copy:
        # created and saved in one function, because it's not useful as an openpyxl object!
        comment_wb = create_copy_with_comments(processor, template_order, save_folder)

    return out_dict


######################################
"""Functions for sorting output dfs"""
######################################


def _generate_template_order(template_map: Dict, template_order: List) -> Dict:
    """
    Assign a unique ascending sort rank to each variable based on order of appearance in the template.

    Args:

        template_map (dict of dicts): a dictionary of information about the fields on the metadata template.
        For example, as of template v3, the first two records should look like
        {'DATASET_dataset_resource_name':
            {'tab': 'Dataset Resource',
            'name': 'Dataset Resource Name',
            (...),
            },
        'DATASET':
            {'tab': 'Dataset Resource',
            'name': 'Acronym',
            (...),
            },
        (...),
        }

        template_order (list): the (internal) names of metadata variables, in a meaningful order.

    Returns:

        out_dict (dict): keys look like "Dataset Resource | Acronym".
        Values are ascending integers, starting from zero.

    Notes:
        This relies on the order of dictionary elements being consistent,
        which is guaranteed as of python 3.7 and semi-reliable as of 3.6.
    """

    out_dict = {}
    order_marker = 0

    for varname in template_order:
        var_dict = template_map[varname]
        id = (" | ").join([var_dict["tab"], var_dict["name"]])
        rank = order_marker
        out_dict[id] = rank
        order_marker += 1

    return out_dict


# TODO: *might* be able to make this cleverer by specifying a 3rd column to sort by next
# i.e. by multiplying out the order dict
def _sort_by_template_order(df: pd.DataFrame, template_order: Dict) -> pd.DataFrame:
    """Sorts a df's rows based on tab and variable name to match the order of the template.

    Args:

        df - a dataframe to sort, i.e. an output table of grouped validation check failures.

        template_order - keys in the format 'tab | name'. Values are ascending integers.

    Returns:

        df - the original df, sorted, with the index updated to match the sort order.

    """
    df = df.copy()
    if "concat_id" not in df.columns:
        df["concat_id"] = df["tab"] + " | " + df["name"]

    assert set(df["concat_id"]).issubset(set(template_order.keys())), (
        "Your df contains unexpected tab & variable combinations!"
    )

    df = df.sort_values(by="concat_id", key=lambda x: x.map(template_order))
    df = df.drop("concat_id", axis=1)
    df = df.reset_index(drop=True)

    return df


def _generate_check_order(v_checks: ModuleType = vc) -> List:
    """Inspects the code for basic validation checks to find the order for sorting lists of checks.

    Args:
        v_checks (python module) - automated_metadata_validation.validation._validation_checks

    Returns:
        just_names (list of str): a list of the names of each non-commented, non-internal
        function in the vc script, in the meaningful order in which they were written.
    """
    vc_text = inspect.getsource(v_checks)
    # all functions that don't start with # and aren't internal (_)
    func_matches = re.findall(pattern=r"(?<!# )def [^_].*\(", string=vc_text)
    # strip the def and (
    # which I could equally do with a more complicated regex probably
    just_names = [funcmatch[4:-1] for funcmatch in func_matches]

    return just_names


def _sort_by_check_order(df: pd.DataFrame, check_list: List) -> pd.DataFrame:
    """Sorts a df's *columns* based on validation check names. Used for orderly heatmap summaries etc.

    Args:
        df - a dataframe where at least some columns are the names of validation checks.

        check_list - an ordered list of the names of validation checks.

    Returns:

        df - the original df, with all non-check-name columns in their original order,
        followed by check-name columns in the sort order.

    """

    not_check_cols = [colname for colname in df.columns if colname not in check_list]
    check_cols = [colname for colname in check_list if colname in df.columns]

    col_order = not_check_cols + check_cols
    df = df[col_order]

    return df


###################################################
"""Functions for wrangling excel cell references"""
###################################################


# this could achieve the same thing by looping over processing.utils.get_excel_ref_components
# but the extra steps it's doing would end up about the same length anyway
def _strip_col_letter(cell_list: List[str]) -> List[int]:
    """Takes a list of excel cells and removes their shared column reference letter.

    Args:
        cell_list (list[str]): A list of excel cell references, e.g. ['A6','A7','A8']. All cells must belong to the same column.

    Returns:
        num_list (list[int]): A list of the row references of the input cells as ints, e.g. [6,7,8].
    """
    assert isinstance(cell_list, list)
    assert all([isinstance(val, str) for val in cell_list])

    letter_list = []
    for val in cell_list:
        col_letter = "".join([char for char in str(val) if char.isalpha()])
        letter_list.append(col_letter)

    assert all([val == letter_list[0] for val in letter_list]), (
        "These cells aren't all from the same column!"
    )
    letter_len = len(letter_list[0])  # could be 2, e.g. AA
    num_list = [int(val[letter_len:]) for val in cell_list]

    return num_list


def _drop_consecutives(cell_list: List, strip_col_letter: bool = True) -> str:
    """Takes a list of individual excel cells, returns a string summarising cell ranges for human readers.

    Args:
        cell_list (list of str or int): the excel cells, each expressed in the format "F1", or just "1" or 1.

        strip_col_letter (bool): whether or not the cell_list includes column letters that must be removed.

    Returns:
        out_string (str): the cell ranges, expressed using hyphens and commas.

    Examples:
        cell_list = [F1,F2,F4,F6,F7,F8], strip_col_letter=True
        out_string = "1-2, 4, 6-8"
    """
    # may be the case with comparative fails
    # that don't neatly relate to particular cells
    if len(cell_list) == 0:
        return ""

    if strip_col_letter:
        num_list = _strip_col_letter(cell_list)
    else:
        num_list = [int(val) for val in cell_list]

    num_list = sorted(num_list)

    # replaced assert with warn
    if len(num_list) != len(set(num_list)):
        warnings.warn(
            "I've found duplicate cell refs - did something go wrong with your groupby?"
        )

    out_list = []
    for index, val in enumerate(num_list):
        # always keep first and last
        if index == 0 or index == (len(num_list) - 1):
            out_list.append(val)
        else:
            prev_val, next_val = num_list[index - 1], num_list[index + 1]
            # always keep isolates
            # and we know we'll never need to read them as ints for banding
            if (prev_val < val - 1) and (next_val > val + 1):
                out_list.append(str(val))
            # start of a band
            elif (prev_val != val - 1) and (next_val == val + 1):
                out_list.append(val)
            # end of a band
            elif (prev_val == val - 1) and (next_val != val + 1):
                prev_out = out_list.pop()
                out_list.append(str(prev_out) + "-" + str(val))

    # clean up, relying on the fact that isolates are already str
    if (
        (len(out_list) > 1)
        and isinstance(out_list[-2], int)
        and isinstance(out_list[-1], int)
    ):
        band_start = out_list.pop(-2)
        band_end = out_list.pop(-1)
        out_list.append(str(band_start) + "-" + str(band_end))

    out_list = [str(val) for val in out_list]
    out_string = (", ").join(out_list)

    return out_string


def _group_and_format_cell_refs(
    df: pd.DataFrame, group_cols: List, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df by the specified columns, finds the cell references belonging to each group, and expresses them as ranges.

    Args:
        df: a fails_df
        group_cols: the columns to group by
        template_order: used to sort the output's rows by the order of variables in the template, if the df contains "tab" and "name" columns.

    Returns:
        group_df - the original df, grouped and possibly sorted
    """
    group_df = df.groupby(group_cols, as_index=False).agg(
        {"cell_ref": lambda x: list(x)}
    )
    group_df["cell_ref"] = group_df["cell_ref"].apply(_drop_consecutives)

    # so that col ref is immediately followed by cell ref
    if "col_ref" in df.columns:
        cell_ref_col = group_df.pop("cell_ref")
        group_df.insert(
            list(group_df.columns).index("col_ref") + 1,
            "cell_ref",
            cell_ref_col,
        )

    if ("tab" in group_df.columns) and ("name" in group_df.columns):
        group_df = _sort_by_template_order(group_df, template_order)

    return group_df


# used for grouping by cell
def _group_and_concat(
    filter_df: pd.DataFrame,
    group_cols: List[str],
    template_order: Dict,
    concat_col="reason",
) -> pd.DataFrame:
    """Takes a filtered fails df, groups by specified cols, and concatenates a designated string column for all rows contributing to the group.

    Args:

        filter_df (pd.DataFrame): a (subset of) a fails df

        group_cols (List[str]): the columns to group by

        template_order (Dict): the order of variables on the excel template

        concat_col (str, optional): The string-type column to concatenate within groups (using a comma and a space).
                                    Defaults to "reason".

    Returns:

        group_df: the result of applying the operations to the original df

    """

    group_df = filter_df.groupby(group_cols, as_index=False).agg(
        {concat_col: lambda x: ", ".join(x)}
    )
    group_df = _sort_by_template_order(group_df, template_order)

    return group_df


#######################################
"""Grouping functions for output dfs"""
#######################################


# TODO: tests, referring to fixture
def _group_fails(
    full_df: pd.DataFrame,
    template_order: Dict,
    condition_col: str,
    is_in: List[str],
    group_cols: List[str],
    group_func: Callable,
) -> pd.DataFrame:
    """Filters a check fails df based on a specified condition, then groups using the specified function.

    Args:
        full_df (pd.DataFrame): an unfiltered fails df.
        template_order (Dict): the order of variables in the excel template.
        condition_col (str): the column to filter on.
        is_in (List[str]): the values to select with the filter.
        group_cols (List[str]): the columns to group by.
        group_func (Callable): the name of the group function to be applied, e.g. _group_and_format_cell_refs, or _group_and_concat.

    Returns:
        pd.DataFrame: a filtered and grouped df.

    Notes:
        The following series of functions are simply named uses of this function with set parameters.
    """

    if "mandatory_field" in full_df.columns:
        group_cols.append("mandatory_field")

    filter_df = full_df.loc[full_df[condition_col].isin(is_in)]
    group_df = group_func(filter_df, group_cols, template_order)

    return group_df


def group_comparatives(full_df: pd.DataFrame, template_order: Dict) -> pd.DataFrame:
    """Groups a fails_df so that each row details one instance of a failed comparative check.

    Args:
        full_df (pd.DataFrame): the full fails_df from the processor object.
        template_order (Dict): the order of variables in the excel template.

    Returns:
        pd.DataFrame
    """

    return _group_fails(
        full_df,
        template_order,
        condition_col="fail_type",
        is_in=["hard_comparative", "soft_comparative"],
        group_cols=[
            "tab",
            "name",
            "col_ref",
            "reason",
            "fail_type",
        ],
        group_func=_group_and_format_cell_refs,
    )


def group_missing_values_by_variable(
    full_df: pd.DataFrame, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df so that each row details the cells with missing values for a single variable.

    Args:
        full_df (pd.DataFrame): the full fails_df from the processor object.
        template_order (Dict): the order of variables in the excel template.

    Returns:
        pd.DataFrame
    """

    return _group_fails(
        full_df,
        template_order,
        condition_col="reason",
        is_in=["missing_value"],
        group_cols=["tab", "name", "col_ref"],
        group_func=_group_and_format_cell_refs,
    )


def group_by_check(full_df: pd.DataFrame, template_order: Dict) -> pd.DataFrame:
    """Groups a fails_df so that each row details the cells of a single variable that have failed a particular hard or soft check.

    Args:
        full_df (pd.DataFrame): the full fails_df from the processor object.
        template_order (Dict): the order of variables in the excel template.

    Returns:
        pd.DataFrame
    """

    return _group_fails(
        full_df,
        template_order,
        condition_col="fail_type",
        is_in=["hard", "soft"],
        group_cols=[
            "tab",
            "name",
            "col_ref",
            "reason",
        ],
        group_func=_group_and_format_cell_refs,
    )


# Not used in current version
# But could potentially be resurrected if a future generic user doesn't care about splitting hard and soft?
# def group_by_cell(full_df: pd.DataFrame, template_order: Dict) -> pd.DataFrame:
#     """Groups a fails_df so that each row details the names of all hard and soft checks failed by a single cell."""

#     return _group_fails(
#         full_df,
#         template_order,
#         condition_col="fail_type",
#         is_in=["hard", "soft"],
#         group_cols=[
#             "tab",
#             "name",
#             "cell_ref",
#             "value",
#         ],
#         group_func=_group_and_concat,  # concat_col defaults to 'reason'
#     )


# TODO: rename "group_fails_by_cell_and_type"?
# test: as before, hand a minimal df in, convert output to records dict, assert equivalence with expectation
def group_by_cell_hard_and_soft(
    full_df: pd.DataFrame,
    template_order: Dict,
) -> pd.DataFrame:
    """Groups a fails_df with one row per cell and two columns concatenating fail messages for hard_fails and soft_fails.

    Args:

        full_df (pd.DataFrame): a processor's fails_df, where each row is one check failure.

        template_order (Dict): a dictionary with keys in the format "tab : varname" and ascending integers as values.

    Returns:
        join_df (pd.DataFrame): a df where each row identifies one invalid template cell, with columns listing all hard and soft fails.
    """

    fail_dfs = {}

    for fail_type in ["hard", "soft"]:
        # can't use the even more generic function because of potential for comparatives with no cell ref
        filter_df = full_df[
            full_df["fail_type"].isin([fail_type, fail_type + "_comparative"])
            & (full_df["cell_ref"] != "")
        ]  # exclude unlocatable (hard) comp fails

        group_df = _group_and_concat(
            filter_df,
            ["tab", "name", "cell_ref", "value"],
            template_order,
            concat_col="reason",
        )
        group_df = group_df.rename({"reason": fail_type + "_fails"}, axis=1)

        fail_dfs[fail_type] = group_df

    # outer - it's possible to have only hard or only soft fails, of course
    join_df = pd.merge(
        fail_dfs["hard"],
        fail_dfs["soft"],
        on=["tab", "name", "value", "cell_ref"],
        how="outer",
    )
    join_df.fillna({"hard_fails": "", "soft_fails": ""}, inplace=True)
    join_df = _sort_by_template_order(join_df, template_order)

    return join_df


# a close cousin to group_by_cell_hard_and_soft...
# ...but not so similar that paramaterising would be both helpful and worthwhile
def group_by_value(full_df: pd.DataFrame, template_order: Dict) -> pd.DataFrame:
    """Groups a fails_df so that each row details a value appearing in a variable, all the cells that value appears in, and all the hard and soft checks that value fails.

    Args:

        full_df (pd.DataFrame): a processor's fails_df, where each row is one check failure.

        template_order (Dict): a dictionary with keys in the format "tab : varname" and ascending integers as values.

    Returns:
        join_df (pd.DataFrame): a df where each row identifies one invalid value, the checks it fails, and all cells it appears in for a given variable.

    """
    fail_dfs = {}
    for fail_type in ["hard", "soft"]:
        df = full_df[full_df["fail_type"] == fail_type]

        group_df = df.groupby(
            ["tab", "name", "col_ref", "value", "mandatory_field"],
            as_index=False,
        ).agg(
            {
                "reason": lambda x: ", ".join(list(dict.fromkeys(x).keys())),
                "cell_ref": lambda x: list(set(x)),
            }
        )

        group_df = group_df.rename({"reason": fail_type + "_fails"}, axis=1)

        group_df["cell_ref"] = group_df["cell_ref"].apply(_drop_consecutives)

        fail_dfs[fail_type] = group_df

    join_df = pd.merge(
        fail_dfs["hard"],
        fail_dfs["soft"],
        on=[
            "tab",
            "name",
            "value",
            "mandatory_field",
            "col_ref",
            "cell_ref",
        ],
        how="outer",
    )
    join_df.fillna({"hard_fails": "", "soft_fails": ""}, inplace=True)

    join_df = _sort_by_template_order(join_df, template_order)

    join_df = join_df[
        [
            "tab",
            "name",
            "col_ref",
            "cell_ref",
            "value",
            "mandatory_field",
            "hard_fails",
            "soft_fails",
        ]
    ]

    return join_df


# For "long % overview", which might be retired
def pivot_and_percent_by_fail_reason(
    full_df: pd.DataFrame,
    row_count_df: pd.DataFrame,
    template_order: Dict,
    check_list: List,
) -> pd.DataFrame:
    """Produces a df where rows are tab & variable name combos and columns are fail %s for every check."""

    non_comp_df = full_df.loc[full_df["fail_type"].isin(["hard", "soft"])]
    group_df = non_comp_df.groupby(["tab", "name", "reason"], as_index=False).agg(
        {"cell_ref": "count"}
    )

    pivot_df = group_df.pivot(
        index=["tab", "name"], columns="reason", values="cell_ref"
    ).reset_index()

    join_df = pd.merge(pivot_df, row_count_df, on="tab")
    join_df = join_df.loc[
        join_df["row_count"] != 0
    ]  # avoiding div/0 for Codes and Values
    join_df = join_df[["tab", "name", "row_count"] + list(pivot_df.columns[2:])]

    # TODO: will eventually want to denote checks not applied to particular variables
    # with N/As...
    # ...but for now, I'll mask with zeroes. Since we're doing fail %s, that'll still
    # be 0% fails!

    join_df = join_df.fillna(0)

    for colname in list(join_df.columns[3:]):
        join_df[colname] = np.round((join_df[colname] / join_df["row_count"]) * 100, 1)

    out_df = _sort_by_check_order(join_df, check_list)
    out_df = _sort_by_template_order(out_df, template_order)

    return out_df


# TODO:
# test: take the fixture(s) for v2 and v3 processor and their fail dfs
# see test_processor.py line ~37 for example
# produce the output, QA it, then save that as a fixture
# fill NA with None, convert to records dict, assert equivalence


# Used for "short % overview"
def make_short_summary_by_variable(
    full_df: pd.DataFrame,
    row_count_df: pd.DataFrame,
    template_order: dict,
) -> pd.DataFrame:
    """Produces a df where rows are tab & variable name combos and the columns list the % of records that failed validation checks.

    Args:
        full_df (pd.DataFrame): a fails_df where each row details a failed check.

        row_count_df (pd.DataFrame): a tiny df denoting the number of records received for each tabular tab.

        template_order (dict): the order of tabs and metadata fields, as seen on the template.

    Returns:
        out_df (pd.DataFrame): a df with % cols for missing values, "failed any hard checks", and "failed only soft checks".

    Notes:
        For brevity, only variables with at least one check failure are included in this summary df.
    """

    # TODO: might be able to link in "does at least one comparative check fail refer to this column"?

    cell_df = group_by_cell_hard_and_soft(
        full_df, template_order
    )  # currently sets aside comparative fails

    out_groups = {
        "missing values": (cell_df["hard_fails"] == "missing_value"),
        "failed any hard checks": (
            (cell_df["hard_fails"] != "missing_value") & (cell_df["hard_fails"] != "")
        ),
        "failed only soft checks": (cell_df["hard_fails"] == "")
        & (cell_df["soft_fails"] != ""),
    }

    df_list = []
    for group_name, condition in out_groups.items():
        # if it's missing, it will never have any other fail reasons, because they're moot
        cond_df = cell_df.loc[condition]
        group_df = cond_df.groupby(["tab", "name"], as_index=False).agg(
            {"cell_ref": "count"}
        )
        # want the renamed cols to be human readable, so can't use new_name=(_,_) syntax
        group_df = group_df.rename({"cell_ref": group_name}, axis=1)
        df_list.append(group_df)

    # May also want to add rows for variables that didn't fail at all,
    # so that it's clear that they were checked and didn't fail!
    # ...on the other hand, that could make this much longer, which sort of defeats the point.
    var_df = (
        full_df[["tab", "name"]]
        .drop_duplicates()
        .sort_values(["tab", "name"])
        .reset_index(drop=True)
    )

    base_df = pd.merge(var_df, row_count_df, on="tab", how="left")
    base_df = base_df.fillna(1)  # vars on cellwise tabs are just single fields
    base_df["row_count"] = base_df["row_count"].astype(int)
    base_df = base_df.rename({"row_count": "total records"}, axis=1)
    df_list.insert(0, base_df)

    join_df = reduce(
        lambda left, right: pd.merge(
            left, right, on=["tab", "name"], how="left", sort=True
        ),
        df_list,
    )
    join_df = join_df.fillna(0)
    # i.e. where there's missing values but not failed checks, or vice versa

    join_df["no issues detected"] = (
        join_df["total records"]
        - join_df["missing values"]
        - join_df["failed any hard checks"]
        - join_df["failed only soft checks"]
    )

    percent_df = join_df
    for colname in [
        "missing values",
        "failed any hard checks",
        "failed only soft checks",
        "no issues detected",
    ]:
        percent_df[colname] = np.round(
            (percent_df[colname] / percent_df["total records"]) * 100, 1
        )
        percent_df = percent_df.rename({colname: colname + " (%)"}, axis=1)

    out_df = _sort_by_template_order(percent_df, template_order)

    return out_df


def _format_fix_df(processor: MetadataProcessor, template_order: Dict) -> pd.DataFrame:
    """Groups the applied_fixes df and improves human readability.

    Args:
        processor (MetadataProcessor): the main object handed over from the processing stage.
        template_order (Dict): the order of variables on the excel template.

    Returns:
        join_df (pd.DataFrame): the "applied_fixes" df from the processor, in an improved format.
    """

    fix_df = processor.recorded_validation_checks["applied_fixes"]

    # since this formatting changes the processor object itself
    # we need to check if this function is unnecessary because it's already run
    if "attempted fixes: hard" in fix_df.columns:
        return fix_df

    fix_df["name"] = fix_df["name"].apply(lambda x: processor.template_map[x]["name"])

    out_dfs = {}
    for fail_type in ["hard", "soft"]:
        filter_df = fix_df[fix_df["fail_type"] == fail_type]
        filter_df = _group_and_concat(
            filter_df, ["tab", "name"], template_order, concat_col="attempted_fix"
        )
        filter_df["attempted_fix"] = filter_df["attempted_fix"].str.replace("_", " ")
        filter_df["attempted_fix"] = filter_df["attempted_fix"].str.replace(
            ", ", ", \n"
        )
        filter_df = filter_df.rename(
            {"attempted_fix": "attempted fixes: " + fail_type}, axis=1
        )
        out_dfs[fail_type] = filter_df

    join_df = pd.merge(
        out_dfs["hard"], out_dfs["soft"], on=["tab", "name"], how="outer"
    )
    join_df.fillna(
        {"attempted fixes: hard": "", "attempted fixes: soft": ""}, inplace=True
    )

    join_df = _sort_by_template_order(join_df, template_order)

    return join_df


########################################
"""Function used to save both outputs"""
########################################


# tests:
# main report
# comment copy
# neither of those things
# what did Ed learn about timestamps in tests?
# A: don't worry too much, just assert that it starts with. A year in the 2020s.
# do the assertions need to be os agnostic?
# A: replace double backslashes with forward slashes
# then can assert expected final string
def _make_save_path(
    md_filepath: str, save_folder: str, mode: str = "main_report"
) -> str:
    """Produces names and filepaths for output files, based on the name of the input file.

    Args:

        md_filepath (str): used to find the name of the original filled template.

        save_folder (str): the folder in which to save the output.

        mode (str): determines one part of the output file name.

    Raises:
        ValueError: if mode is not one of 'main_report' or 'comment_copy'.

    Returns:
        save_path (str): a filepath including a file name and a timestamp.
    """

    save_name = os.path.basename(md_filepath)
    save_name = os.path.splitext(save_name)[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    if mode == "main_report":
        save_name = f"{save_name}_VALIDATION_REPORT_{timestamp}.xlsx"
    elif mode == "comment_copy":
        save_name = f"COMMENT_COPY_DO_NOT_SUBMIT_{save_name}_{timestamp}.xlsx"
    else:
        raise ValueError("Please specify mode = 'main_report' or 'comment_copy'")

    save_path = os.path.join(save_folder, save_name)

    return save_path


#############################################
"""Functions for the tabular report output"""
#############################################


# test: probably not proportionate to test
def _fit_and_wrap(
    wb: Workbook, sheet_name: str, df: pd.DataFrame, filter_padding=True
) -> None:
    """Uses some rough conversions to approximate an 'autofit' for columns based on the number of characters in the cells, since openpyxl can't do this directly.

    Args:
        wb (Workbook): an openpyxl workbook, with columns whose width needs adjusting.
        sheet_name (str): the name of the sheet to edit.
        df (pd.DataFrame): the dataframe containing the data that's just been saved to this sheet.
        filter_padding (bool, optional): Whether or not to add some width to account for the drop-down arrow symbol. Defaults to True.

    Returns:
        None (the workbook object is edited in place).
    """

    for colname in df.columns:
        # columns where strings-with-commas also have newlines
        # TODO: consider moving this list to a constant, because I'm now using it in two and a half places
        if colname in [
            "hard_checks",
            "soft_checks",
            "hard_fails",
            "soft_fails",
            "reason",
            "names",
            "attempted fixes: hard",
            "attempted fixes: soft",
        ]:
            max_cell_len = max(
                df[colname].apply(lambda x: max([len(i) for i in x.split(",")]))
            )
        else:
            max_cell_len = max(df[colname].astype(str).str.len())

        max_len = min(200, max(max_cell_len, len(colname)))

        # the little filter arrow is about 3 units wide
        if (max_len <= len(colname) + 3) and filter_padding:
            max_len += 3

        # whether you offset by 1 or 2 depends on whether you're outputting the index to excel
        number_index = list(df.columns).index(colname) + 1
        letter_index = get_column_letter(number_index)

        wb[sheet_name].column_dimensions[letter_index].width = max_len

        for cell in wb[sheet_name][letter_index]:
            if (max_len == 200) or (
                isinstance(cell.value, str) and ("," in cell.value)
            ):
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")


def _split_by_mandatory(
    fails_df: pd.DataFrame, template_map: Dict
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Uses the template map to split processor.fails_df into fails for mandatory fields and fails for other fields.

    Args:

        fails_df - a dataframe where each row is an individual validation check failure.

        template_map - a nested dictionary detailing the shape and nature of the excel metadata template.

    Returns:

        mand_df - a df of all check failures belonging to mandatory variables.

        non_mand_df - a df of all check failures belonging to optional and conditional variables.

        mand_join_df - the original df, now with a column indicating mandatory status

    """
    mand_data_list = []
    for inner_dict in template_map.values():
        mand_data_list.append(
            [inner_dict["tab"], inner_dict["name"], inner_dict["mandatory"]]
        )

    mand_df = pd.DataFrame(
        data=mand_data_list, columns=["tab", "name", "mandatory_field"]
    )

    mand_join_df = pd.merge(fails_df, mand_df, on=["tab", "name"], how="left")

    mand_df = mand_join_df.loc[mand_join_df["mandatory_field"]]
    non_mand_df = mand_join_df.loc[~mand_join_df["mandatory_field"]]

    assert len(mand_df) + len(non_mand_df) == len(fails_df), (
        "Looks like something's gone wrong when joining the mandatory tags!"
    )

    return mand_df, non_mand_df, mand_join_df


# test: make the wb, then spot check assertions about it
def _write_guidance_tab(wb: Workbook, guide_df: pd.DataFrame) -> Workbook:
    """Refers to output_constants.py to create a workbook with a guidance sheet containing user documentation.

    Args:
        wb (openpyxl workbook) - the 'tabular report' workbook, awaiting its guidance page.
        guide_df (pd.DataFrame) - the dataframe containing information about the tabs of the workbook.

    Returns:
        wb (openpyxl workbook)

    Notes:
        This is necessary for packaging, since Pypi doesn't like including .xlsx files so we can't just keep a template.
        Refers to GUIDE_LINES from output_constants.py.
    """

    # relies on overwriting the default sheet, so that this is first
    wb["Sheet"].title = "Guidance"

    # These are simply known from having previously mocked up the sheet in excel
    wb["Guidance"].column_dimensions["A"].width = 30
    wb["Guidance"].column_dimensions["B"].width = 20
    wb["Guidance"].column_dimensions["C"].width = 20
    wb["Guidance"].column_dimensions["D"].width = 80

    """Written guidance"""

    for line in oc.GUIDE_LINES:
        # openpyxl wants each row to be expressed as a list, even if it's just one cell's text
        if isinstance(line, str):
            line = [line]
        wb["Guidance"].append(line)

        text_len = len(line[0])
        if text_len > 150:
            this_row = wb["Guidance"].max_row
            wb["Guidance"].row_dimensions[this_row].height = 16 * math.ceil(
                text_len / 150
            )
            wb["Guidance"].merge_cells(
                start_row=this_row,
                start_column=1,
                end_row=this_row,
                end_column=4,
            )
            wb["Guidance"]["A" + str(this_row)].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

    wb["Guidance"].cell(1, 1).font = Font(size=18, bold=True)

    wb["Guidance"].append([""])  # spacer
    table_start_row = wb["Guidance"].max_row + 1  # easiest to remember this now

    """Table explaining contents of sheets"""

    # turn that df back into rows (as list of lists), to segue neatly with existing code

    table_rows = [list(guide_df.columns)] + guide_df.values.tolist()

    for row in table_rows:
        wb["Guidance"].append(row)

    table_end_row = table_start_row + len(table_rows) - 1
    table_start_cell = "A" + str(table_start_row)
    table_end_cell = get_column_letter(len(table_rows[0])) + str(table_end_row)

    _format_as_table(
        wb["Guidance"], first_cell=table_start_cell, last_cell=table_end_cell
    )

    for row in range(
        table_start_row + 1, table_end_row + 1
    ):  # no headers, yes last row
        # do clever things based on sheet name
        sheet_key = (
            wb["Guidance"].cell(row, 1).value
        )  # not zero indexed - careful with interaction with range() here!
        colour = oc.HEXCODES[oc.SHEET_DEETS[sheet_key]["Colour"]]
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        wb["Guidance"].cell(row, 1).fill = fill

        # looks like I might need the file name for internal links to work, which isn't possible at this step
        # if sheet_key in wb.sheetnames:
        #     #internal sheet link format. Note that excel needs the ' !
        #     wb["Guidance"].cell(row,1).hyperlink = "'#"+sheet_key+"!"+"A1'"

        # wrap descriptions
        wb["Guidance"].row_dimensions[row].height = 16 * math.ceil(
            len(wb["Guidance"].cell(row, 4).value) / 80
        )
        for col in ["A", "B", "C", "D"]:
            wb["Guidance"][col + str(row)].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

    # Would be nice to have a little margin, but this causes problems with the table reference
    # wb['Guidance'].insert_rows(1)
    # wb['Guidance'].insert_cols(1)

    return wb


# TODO: consider moving the invocation of this earlier, to reduce duplication
def _underscores_to_spaces(df: pd.DataFrame) -> pd.DataFrame:
    """Helper function to replace underscores in fixed list of columns that need to be more human readable."""
    for colname in [
        "reason",
        "hard_fails",
        "soft_fails",
        "hard_checks",
        "soft_checks",
        "check",
        # intentionally excludes "attempted_fix", which is handled independently earlier
    ]:
        if colname in df.columns:
            df[colname] = df[colname].str.replace("_", " ")

    return df


def _colour_code_tabs(wb: Workbook) -> Workbook:
    """Colour codes the tabular report's sheets to indicate similar groups."""

    for sheet_name in wb.sheetnames:
        if sheet_name in oc.SHEET_DEETS.keys():
            wb[sheet_name].sheet_properties.tabColor = oc.HEXCODES[
                oc.SHEET_DEETS[sheet_name]["Colour"]
            ]

    return wb


# todo: decide if I want to pull this out for both short and long %
def _apply_heatmap(ws):
    pass


def _format_as_table(ws: Worksheet, first_cell=None, last_cell=None) -> Worksheet:
    """When a df has been written to a worksheet, format it nicely as a table with borders, shading, and filters.

    Args:
        ws (Worksheet): the worksheet to which the df has been written.
        first_cell (_type_, optional): the coordinates of the top left cell of the table (including headers). Defaults to A1.
        last_cell (_type_, optional): the coordinates of the bottom right cell of the table. Defaults to the last populated cell on the sheet.

    Returns:
        Worksheet: the worksheet, with table formatting applied.
    """
    # tables must have (unique) names, but they're picky
    base_name = ws.title.replace(" ", "_")
    safe_name = ""
    for i in base_name:
        if i.isalpha() or (i == "_"):
            safe_name += i

    if first_cell is None:
        first_cell = "A1"
    if last_cell is None:
        last_cell = get_column_letter(ws.max_column) + str(ws.max_row)
    table = Table(displayName=safe_name, ref=first_cell + ":" + last_cell)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    return ws


def _write_df_as_table(sheet_name: str, df: pd.DataFrame, wb: Workbook):
    """Writes a df to a new sheet and formats it nicely."""

    wb.create_sheet(sheet_name)

    # We now do want to keep the index, because it's often a meaningful sort order
    # However, we need to turn the index into a *column* in order to format as table
    df = df.reset_index(names="Index")

    if not df.empty:
        for row in dataframe_to_rows(df, index=False, header=True):
            wb[sheet_name].append(row)
        _fit_and_wrap(wb, sheet_name, df, filter_padding=True)
        _format_as_table(wb[sheet_name])
    else:
        wb[sheet_name]["A1"].value = f"{sheet_name}: No issues detected"


def _prep_background_dfs(processor):
    """Prepares dfs of suporting information: checks applied, fixes attempted, etc"""
    out_dfs = {}

    # note that the 'applied_fixes' df, if present, will have had some formatting in advance
    for name, df in processor.recorded_validation_checks.items():
        df = _underscores_to_spaces(df)
        out_dfs[name] = df

    guidance = []
    for sheet_name, deets in oc.SHEET_DEETS.items():
        if sheet_name == "applied_fixes" and "applied_fixes" not in out_dfs.keys():
            continue
        else:
            guidance.append(
                [
                    sheet_name,
                    deets.get("Variables"),
                    deets.get("Checks"),
                    deets.get("Description"),
                ]
            )

    guide_df = pd.DataFrame(
        columns=["Sheet", "Variables", "Checks", "Description"], data=guidance
    )
    # to mirror sheet order, we want the guidance df 'first' in the dict (to the extent that that's enforced)
    out_dfs = {**{"Guidance": guide_df}, **out_dfs}

    return out_dfs


# test: obviously, this needs a processor object
# if we have checks for the individual parts...
# ...just make sure it's producing a df with the expected keys
# and the values are all dfs.
def create_report_dfs(processor: MetadataProcessor, template_order: dict) -> Dict:
    """Draws the fails_df from a processor object, produces a dictionary of dataframes destined for separate tabs of a report."""

    fails_df = processor.fails_df
    # TODO: does get_excel_ref_components handle nulls?
    fails_df["col_ref"] = (
        fails_df["cell_ref"]
        .apply(lambda x: get_excel_ref_components(x)[0] if x else None)
        .str[0]
    )  # returns a tuple, e.g. ["A",1]

    """supporting dfs"""

    background_dfs = _prep_background_dfs(processor)

    """data dfs"""

    # temporary - some issues with duplication earlier in processing
    fails_df = fails_df.drop_duplicates()

    mand_df, non_mand_df, mand_join_df = _split_by_mandatory(
        fails_df, processor.template_map
    )

    # may ask for row counts as df in next iteration
    row_count_df = pd.DataFrame(
        data=list(processor.tab_lens.items()), columns=["tab", "row_count"]
    )

    check_order = _generate_check_order(vc)
    data_dfs = {}

    # "Purple group"
    data_dfs["Short % overview"] = make_short_summary_by_variable(
        mand_df, row_count_df, template_order
    )
    data_dfs["Long % overview"] = pivot_and_percent_by_fail_reason(
        mand_df, row_count_df, template_order, check_order
    )

    # "Orange group"
    data_dfs["Fails by value"] = group_by_value(mand_join_df, template_order)
    data_dfs["Fails by cell - Mandatory"] = group_by_cell_hard_and_soft(
        mand_df, template_order
    )

    data_dfs["Fails by check - Hard"] = group_by_check(
        mand_join_df[mand_join_df["fail_type"] == "hard"], template_order
    )

    # "Blue group"
    data_dfs["Missing values"] = group_missing_values_by_variable(
        mand_join_df, template_order
    )
    data_dfs["Comparative check fails"] = group_comparatives(
        mand_join_df, template_order
    )

    # "Yellow group"
    data_dfs["Fails by cell - Non mandatory"] = group_by_cell_hard_and_soft(
        non_mand_df, template_order
    )
    data_dfs["Fails by check - Soft"] = group_by_check(
        mand_join_df[mand_join_df["fail_type"] == "soft"], template_order
    )

    # finally replace those underscores for human readability
    for df in data_dfs.values():
        df = _underscores_to_spaces(df)

    out_dfs = {**background_dfs, **data_dfs}

    return out_dfs


# test: create the wb, assert that it has certain sheets names in a certain order
def create_output_report(df_dict: Dict) -> openpyxl.Workbook:
    """Takes a dictionary of dataframes and writes them to the sheets of an openpyxl workbook."""

    wb = Workbook()

    # special case - needs merging with its blurb
    df_dict = df_dict.copy()
    guide_df = df_dict.pop("Guidance")

    # line breaks within cell - formatting step that's only relevant when writing to excel
    for df in df_dict.values():
        for colname in [
            "hard_checks",
            "soft_checks",
            "hard_fails",
            "soft_fails",
            "reason",
            "names",
        ]:
            if colname in df.columns:
                # Avoids aberrant behaviour when running in succession
                df[colname] = df[colname].str.replace(r", (\n)*", ", \n", regex=True)

    for sheet_name, df in df_dict.items():
        _write_df_as_table(sheet_name, df, wb)

    # heatmap for %s in short overview
    if "Short % overview" in wb.sheetnames:
        max_row = wb["Short % overview"].max_row
        wb["Short % overview"].conditional_formatting.add(
            "E2:G" + str(max_row), oc.HEATMAP_RULE
        )

    wb = _colour_code_tabs(wb)

    wb = _write_guidance_tab(wb, guide_df)

    return wb


######################################
"""Functions for the commented copy"""
######################################


# tests:
# hard only, soft only, hard and soft
def _prep_hard_and_soft_comment_text(hard_fails: str, soft_fails: str) -> str:
    """Takes concatenated strings of hard and soft fails, and makes them human readable for the commented output.

    Args:
        hard_fails (str) - a string of one or more hard fail reasosns, such as 'missing_value'.

        soft_fails (str) - a string of one or more soft fail reasons, such as 'must_not_end_with_whitespace'.

    Returns:
        com_text (str) - in the format "HARD FAILS (#): ... SOFT FAILS (#): ...".

    """
    com_text = ""

    if hard_fails != "":
        hard_count = hard_fails.count(",") + 1  # fenceposts
        com_text = (
            f"HARD FAILS ({hard_count}): " + hard_fails.replace("_", " ") + " "
        )  # trail a little whitespace in case there's soft fails too

    if soft_fails != "":
        soft_count = soft_fails.count(",") + 1
        com_text += f"SOFT FAILS ({soft_count}): " + soft_fails.replace("_", " ")

    return com_text


def _split_and_report_unlocatable_fails(
    wb: Workbook, fails_df: pd.DataFrame
) -> Workbook:
    """Separates out any fails_df rows with no cell ref and reports on them in a separate sheet."""

    no_cell_ref_df = fails_df[
        fails_df["cell_ref"].isna() | (fails_df["cell_ref"] == "")
    ].copy()
    wb.create_sheet("OTHER CHECK FAILS")
    wb["OTHER CHECK FAILS"].cell(
        1, 1
    ).value = "This sheet is for noting comparative checks that can't easily be associated with one single cell"
    wb["OTHER CHECK FAILS"].cell(
        2, 1
    ).value = "And therefore can't be reported on through highlighting and cell notes."

    if no_cell_ref_df.empty:
        wb["OTHER CHECK FAILS"].cell(
            3, 1
        ).value = "...however, no fails of this type were detected this run!"
    else:
        wb["OTHER CHECK FAILS"].append([""])
        for row in dataframe_to_rows(no_cell_ref_df, index=False, header=True):
            wb["OTHER CHECK FAILS"].append(row)

        _fit_and_wrap(wb, "OTHER CHECK FAILS", no_cell_ref_df)

    return wb


# tests: active=important notice,
# A10="See 'OTHER COMP CHECK FAILS' sheet for details."
# resource A1 = "Validation comments doc - NOT FOR INGESTION"
def _write_commented_copy_notices(metadata_wb: Workbook) -> Workbook:
    """
    Places warnings and intentionally sabotages the commented copy to discourage ingestion. Writes boilerplate guidance page.

    Args:
        metadata_wb - an openpyxl workbook of the original filled template

    Returns:
        metadata_wb - the template, with added text and a new sheet.

    Notes:
        Unfortunately, attempts at protecting sheets don't work because of the existing protection.
    """

    # this will intentionally cause ingest to fail if this file is submitted
    metadata_wb["Change history"][
        "D1"
    ].value = "Validation comments doc - NOT FOR INGESTION"
    metadata_wb["Dataset Resource"][
        "A1"
    ].value = "Validation comments doc - NOT FOR INGESTION"
    metadata_wb.create_sheet("Important notice")
    metadata_wb.active = metadata_wb["Important notice"]

    notice_text = [
        "DO NOT edit and submit this workbook.",
        "It is a commented copy for reference only.",
        "If you submit this file, it will not be accepted.",
        "Please reopen and edit the original, using this file as a guide.",
        "",
        "Orange cells contain 'hard' check fails that will **prevent a successful ingest.**",
        "Yellow cells contain only 'soft' fails.",
        "Mouse over highlighted cells to see details of failed checks.",
        "",
        "Note that some checks don't belong to single cells.",
        "See 'OTHER CHECK FAILS' sheet for details.",
        "",
        "If you set 'save_corrected_copy', automatic fixes will be reflected in this file.",
        "For example, removing leading and trailing whitespace.",
        "Refer to 'APPLIED FIXES' for a summary.",
        "In this case, you may copy and paste (values only) the fixed contents of this form back into your working copy.",
    ]

    for text in notice_text:
        row_list = [text]  # append expects one list per row
        metadata_wb["Important notice"].append(row_list)

    return metadata_wb


# test: write the output, read it in, processor.load_xl...
# make sure it's writing to a relative location, so that things don't go weird on github
# ...if that works, we know that it's written right
# convert_to_version.py test has example of temporary files in tests
def create_copy_with_comments(
    processor: MetadataProcessor, template_order: dict, save_folder: str
) -> openpyxl.Workbook:
    """Uses the df of hard and soft fails to add comments and highlighting to a copy of the original template.

    Args:

        processor - the interface containing preprocessed objects used to produce outputs.
        This function uses processor.fails_df and processor.md_filepath.

        template_order (Dict): a dictionary with keys in the format "tab : varname" and ascending integers as values.

        save_folder - the folder in which to save outputs.

    Returns:

        metadata_wb - the filled template, with comments added.

    """

    fails_df = processor.fails_df.drop_duplicates()
    cell_hard_soft_df = group_by_cell_hard_and_soft(fails_df, template_order)

    metadata_wb = openpyxl.load_workbook(processor.md_filepath)

    metadata_wb = _split_and_report_unlocatable_fails(metadata_wb, fails_df)

    # I don't think excel 'notes' actually display this
    author = "Automated Metadata Validation Tool"

    for index, df_row in cell_hard_soft_df.iterrows():
        sheet = metadata_wb[df_row["tab"]]
        cell = sheet[df_row["cell_ref"]]

        com_text = _prep_hard_and_soft_comment_text(
            df_row["hard_fails"], df_row["soft_fails"]
        )

        comment = Comment(com_text, author)
        comment.width = 300
        comment.height = 20 * (
            1 + math.ceil(len(com_text) / 40)
        )  # based on chars per 300-pixel line
        cell.comment = comment

        if "HARD FAILS" in com_text:
            cell.fill = oc.ORANGE_FILL
        else:
            cell.fill = oc.YELLOW_FILL

    # only relevant if save_commented_copy was True
    if "applied_fixes" in processor.recorded_validation_checks.keys():
        _write_df_as_table(
            "APPLIED FIXES",
            processor.recorded_validation_checks["applied_fixes"],
            metadata_wb,
        )

    metadata_wb = _write_commented_copy_notices(metadata_wb)

    save_path = _make_save_path(processor.md_filepath, save_folder, mode="comment_copy")
    metadata_wb.save(save_path)
    print(f"Saved to {save_path}!")

    return metadata_wb
