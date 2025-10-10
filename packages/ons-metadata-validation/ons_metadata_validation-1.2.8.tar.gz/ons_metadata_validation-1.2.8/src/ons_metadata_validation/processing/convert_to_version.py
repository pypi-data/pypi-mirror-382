from copy import deepcopy
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import load_workbook

from ons_metadata_validation.processing.processor import MetadataProcessor
from ons_metadata_validation.processing.utils import (
    apply_version_changes,
    write_dfs_to_wb,
)
from ons_metadata_validation.reference.constants import (
    MetaMapper,
    SheetMapper,
    SupplyType,
)
from ons_metadata_validation.reference.delta_table import DELTA_TABLE
from ons_metadata_validation.reference.template import V2_TEMPLATE


def convert_template_to_version(
    orig_path: str,
    empty_target_path: str,
    save_path: str,
    target_version: float = 3.0,
    default_values: Optional[Dict] = None,
    apply_autofixes: bool = False,
) -> bool:
    """Convert a metadata template to another version.

    Args:
        orig_path (str): Path to the original template. Version is inferred.
        empty_target_path (str): Path to the empty dst template.
        save_path (str): Where to save the dst template.
        target_version (float, optional): The version to transform to. Defaults to 3.0.
        default_values (Dict, optional): Default values to populate the template with. Defaults to None.
        apply_autofixes (bool, optional): Apply autofixes before converting to V3 template. Defaults to False.

    Returns:
        bool: True if successful. False otherwise.

    Notes:
        Please also run the main function to validate the results of your conversion!

    Example:
    >>> convert_template_to_version(
    >>>     "path/to/original_v1/template.xlsx",
    >>>     "path/to/empty_v3/template.xlsx",
    >>>     "path/where/to/save/v3/template.xlsx",
    >>>     target_version=3,
    >>>     default_values={
    >>>         'DATASET_access_level': "Access level 3",
    >>>         "DATASET_safe_settings": "ESRC SafePods, Assured Organisational Connectivity (office-based), Assured Organisational Connectivity (homeworking)",
    >>>         "DATASET_subject_to_low_level_access_control": "No",
    >>>         "VARS_variable_availability": "Standard",
    >>>         "VARS_row_level_restrictions": "No"
    >>>     }
    >>> )
    """
    inp_args = locals()
    try:
        mp = MetadataProcessor(orig_path, "full", False, False, apply_autofixes)
        mp.load_xl()
        version_template = apply_version_changes(
            V2_TEMPLATE, DELTA_TABLE, target_version=target_version
        )
        wb = load_workbook(empty_target_path)

        default_values = default_values or {}

        for key, default_value in default_values.items():
            node = version_template.get(key, {})

            if node.get("enum") and default_value not in node.get("enum", []):
                print(
                    f"Invalid default value: `{default_value}` for `{node.get('name')}`. Use one of: {node.get('enum')}"
                )
                continue

            mp.xl[node["tab"]][key] = default_value

        write_dfs_to_wb(version_template, mp.xl, wb)
        for sheet in wb.worksheets:
            sheet.protection.sheet = False
        wb.save(save_path)
        return True
    except Exception as e:
        print(f"{e} for input args {inp_args}")
        return False


def combine_raw_and_mastered_templates(
    raw_path: Union[str, List[str]],
    mastered_path: str,
    blank_orig_path: str,
    save_path: str,
) -> bool:
    """Combines data from the raw version of the metadata template with the mastered version.

    Args:
        raw_path (Union[str, List[str]]):
            if given a str, this is the path to the single raw metadata template.
            if given a list it will sequentially apply the changes from the templates to
            the first one in order.
        mastered_path (str): the path to the mastered metadata template
        blank_orig_path (str): the path to the blank V2 metadata template
        save_path (str): the path to save the combined V2 template

    Returns:
        bool: True if successful else false

    Using dictionary unpacking we can ensure that information in the raw version that isn't
    updated in the mastered version is retained.

    Example:
    >>> raw = {"A": "1", "B": "2", "C": ""}
    >>> mastered = {"A": "4", "B": "", "C": "6"}
    >>> unpacked = {**raw, **mastered}
    >>> # {"A": "4", "B": "2", "C": "6"}

    We can see that the values for "B" from raw is retained. This is done on each row in the
    dataframe to combine the data in the raw template and any modifications in the mastered copy.

    """
    inp_args = locals()

    collapse_config = [
        (
            SheetMapper.SERIES.value,
            MetaMapper.SERIES_google_cloud_platform_bigquery_table_name.value,
            [MetaMapper.SERIES_google_cloud_platform_bigquery_table_name.value],
        ),
        (
            SheetMapper.FILE.value,
            MetaMapper.FILE_google_cloud_platform_bigquery_table_name.value,
            [MetaMapper.FILE_file_path_and_name.value],
        ),
        (
            SheetMapper.VARS.value,
            MetaMapper.VARS_google_cloud_platform_bigquery_table_name.value,
            [
                MetaMapper.VARS_google_cloud_platform_bigquery_table_name.value,
                MetaMapper.VARS_variable_name.value,
            ],
        ),
    ]

    try:
        if isinstance(raw_path, str):
            raw_mp = _get_mp_obj(raw_path)
        if isinstance(raw_path, list):
            raw_mp = _combine_raw_md_templates(raw_path, collapse_config)

        mastered_mp = _get_mp_obj(mastered_path)
        wb = load_workbook(blank_orig_path)

        final_xl = {}

        for tab, gcp_table_name_col, collapse_cols in collapse_config:
            raw_xl = raw_mp.xl[tab].copy()
            mastered_xl = mastered_mp.xl[tab].copy()
            mastered_xl = mastered_xl[
                ~mastered_xl[gcp_table_name_col].str.endswith("_curated")
            ]

            raw_json = _collapse_tab(raw_xl, collapse_cols)
            mastered_json = _collapse_tab(mastered_xl, collapse_cols)

            final_xl[tab] = pd.DataFrame(
                [{**v, **mastered_json.get(k, {})} for k, v in raw_json.items()]
            )

        for tab in [SheetMapper.RESOURCE.value, SheetMapper.BACK_OFFICE.value]:
            final_xl[tab] = pd.DataFrame(
                {
                    **_convert_df_to_flat_dict(raw_mp.xl[tab]),
                    **_convert_df_to_flat_dict(mastered_mp.xl[tab]),
                },
                index=[0],
            )

        write_dfs_to_wb(mastered_mp.template_map, final_xl, wb)
        wb.save(save_path)
        return True

    except Exception as e:
        print(f"{e} for input args {inp_args}")
        return False


def _combine_raw_md_templates(
    raw_paths: List[str], collapse_config: List[Tuple[str, str, List[str]]]
) -> MetadataProcessor:
    """Combine multiple raw templates into a single raw template.

    Args:
        raw_paths (List[str]): the raw metadata paths to combine sequentially
        collapse_config (List[Tuple[str, str, List[str]]]): the config passed down from
            the combine_raw_and_mastered_templates function

    Raises:
        ValueError: if not all templates have the same dataset name

    Returns:
        MetadataProcessor: The first metadata processor object with the combined raw template
    """
    raw_mps = [_get_mp_obj(raw_path) for raw_path in raw_paths]
    dataset_names = {
        v.md_filepath: v.xl[SheetMapper.BACK_OFFICE.value][
            MetaMapper.DATASET_google_cloud_platform_big_query_dataset_name.value
        ].values[0]
        for v in raw_mps
    }

    if not all(d == dataset_names[raw_paths[0]] for d in dataset_names.values()):
        raise ValueError(f"Not all templates for the same dataset: {dataset_names}")

    prev_raw_xl = deepcopy(raw_mps[0].xl)
    change_data = deepcopy(prev_raw_xl)

    for curr_mp in raw_mps[1:]:
        curr_raw_xl = deepcopy(curr_mp.xl)
        records = curr_raw_xl[SheetMapper.SERIES.value].to_dict("records")

        for record in records:
            table_name = record[
                MetaMapper.SERIES_google_cloud_platform_bigquery_table_name.value
            ]
            supply_type = record[MetaMapper.SERIES_supply_type.value].strip().upper()
            existing_tables = prev_raw_xl[SheetMapper.SERIES.value][
                MetaMapper.SERIES_google_cloud_platform_bigquery_table_name.value
            ].values

            if table_name in existing_tables and supply_type == "FULL":
                for sheet, key, _ in collapse_config:
                    # only case where we overwrite
                    prev_raw_xl[sheet] = prev_raw_xl[sheet][
                        prev_raw_xl[sheet][key] != table_name
                    ]

                    if table_name in prev_raw_xl[sheet][key].unique():
                        ValueError(
                            f"Failed to remove table {table_name} from sheet {sheet}"
                        )

                    change_data[sheet] = pd.concat(
                        [
                            curr_raw_xl[sheet][curr_raw_xl[sheet][key] == table_name],
                            prev_raw_xl[sheet],
                        ]
                    )

            elif (table_name in existing_tables and supply_type == "APPEND") or (
                table_name not in existing_tables and supply_type == "FULL"
            ):
                for sheet, key, unique_cols in collapse_config:
                    change_data[sheet] = pd.concat(
                        [
                            curr_raw_xl[sheet][curr_raw_xl[sheet][key] == table_name],
                            prev_raw_xl[sheet],
                        ]
                    ).drop_duplicates(subset=unique_cols)

            else:
                print(
                    f"{record} set to '{supply_type}' but not present in {existing_tables}"
                )

            prev_raw_xl = deepcopy(change_data)

    raw_mp = raw_mps[0]
    raw_mp.xl = change_data

    # by stacking changes over time the pipeline should see everything as FULL for this reingest
    raw_mp.xl[SheetMapper.SERIES.value][MetaMapper.SERIES_supply_type.value] = (
        SupplyType.FULL.value
    )

    return raw_mp


def _get_mp_obj(path: str) -> MetadataProcessor:
    mp = MetadataProcessor(path, "full", False, False, False)
    mp.load_xl()
    return mp


def _convert_df_to_flat_dict(df: pd.DataFrame) -> Dict[str, Any]:
    return _remove_empty_values(df.T.to_dict()[0])


def _collapse_tab(df: pd.DataFrame, key_cols: list) -> Dict[str, Dict[str, Any]]:
    return {
        _get_key(node, key_cols): _remove_empty_values(node)
        for node in df.to_dict("records")
    }


def _remove_empty_values(record: dict) -> Dict[str, Any]:
    return {k: v for k, v in record.items() if v}


def _get_key(node: dict, key_cols: list) -> str:
    return "|table_var_delim|".join([node[k] for k in key_cols])
