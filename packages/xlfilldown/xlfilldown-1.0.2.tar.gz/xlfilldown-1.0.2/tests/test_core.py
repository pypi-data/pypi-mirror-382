import io
from pathlib import Path
import sqlite3
from datetime import datetime
from decimal import Decimal
import subprocess
import sys
import shutil

import openpyxl
import pytest

from xlfilldown.core import qident, normalize_headers, canon_list, sha256_hex
from xlfilldown.api import ingest_excel_to_sqlite, ingest_excel_to_excel


# -----------------------------
# Helpers
# -----------------------------

def _make_workbook(tmp_path: Path):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header at row 1
    ws.append(["tier0", "tier1"])
    # Data rows per the example:
    ws.append(["value1", "tiervalue1"])
    ws.append([None, "tiervalue2"])
    ws.append([None, None])          # completely empty row (no fill-down)
    ws.append([None, "tiervalue3"])
    ws.append(["value2", "tiervalue4"])
    ws.append([None, "tiervalue5"])
    wb.save(p)
    return p


def _write_wb_simple(tmp_path: Path, title="S"):
    p = tmp_path / "in_cli.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["A", "B"])
    ws.append(["x1", "y1"])
    ws.append([None, "y2"])
    ws.append(["x3", None])
    wb.save(p)
    return p


def _mk_in(tmp_path: Path):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    ws.append(["x"])
    wb.save(p)
    wb.close()
    return p


# -----------------------------
# Core unit tests
# -----------------------------

def test_qident():
    assert qident('a') == '"a"'
    assert qident('a"b') == '"a""b"'


def test_normalize_headers():
    headers = normalize_headers([None, ' A ', 'nan', 'ok'])
    assert headers == ['', 'A', '', 'ok']


def test_canon_list_and_hash():
    s = canon_list([None, ' A ', 5])
    assert s == '["","A","5"]'
    h = sha256_hex(s)
    assert len(h) == 64


def test_ingest_to_sqlite_with_empty_row_handling(tmp_path):
    infile = _make_workbook(tmp_path)
    db_path = tmp_path / "out.db"

    summary = ingest_excel_to_sqlite(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        fill_cols=["tier0"],                 # API uses fill_cols
        db=db_path,
        table="t",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=True,
        excel_row_numbers=True,
        if_exists="fail",
        batch_size=50,
        fill_mode="hierarchical",            # explicit for clarity
    )

    assert summary["table"] == "t"
    assert summary["row_hash"] is True
    assert summary["excel_row_numbers"] is True

    with sqlite3.connect(str(db_path)) as conn:
        cur = conn.cursor()
        rows = cur.execute('SELECT excel_row, tier0, tier1 FROM t ORDER BY excel_row').fetchall()
        assert len(rows) == 6
        by_row = {r[0]: (r[1], r[2]) for r in rows}
        assert by_row["2"] == ("value1", "tiervalue1")
        assert by_row["3"] == ("value1", "tiervalue2")
        assert by_row["4"] == (None, None)          # spacer preserved
        assert by_row["5"] == ("value1", "tiervalue3")
        assert by_row["6"] == ("value2", "tiervalue4")
        assert by_row["7"] == ("value2", "tiervalue5")


def test_ingest_to_excel_with_append_and_headers(tmp_path):
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "out.xlsx"

    summary1 = ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        fill_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        fill_mode="hierarchical",
    )
    assert summary1["sheet"] == "Processed"
    assert summary1["rows_written"] == 6

    summary2 = ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        fill_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="append",
        fill_mode="hierarchical",
    )
    assert summary2["rows_written"] == 6

    wb = openpyxl.load_workbook(outfile)
    ws = wb["Processed"]
    assert ws.max_row == 1 + 12
    assert ws.cell(row=2, column=1).value == "value1"
    assert ws.cell(row=3, column=1).value == "value1"
    assert ws.cell(row=4, column=1).value is None and ws.cell(row=4, column=2).value is None


def test_row_hash_float_decimal_stability():
    cases = [
        [1, 2.50, 1000000],
        [1.0, 2.5, 1_000_000.0],
        [Decimal("1"), Decimal("2.5000"), Decimal("1000000.000")],
    ]
    texts = [canon_list(c) for c in cases]
    hashes = [sha256_hex(t) for t in texts]
    assert len(set(texts)) == 1
    assert len(set(hashes)) == 1


def test_whitespace_only_cells_are_blank_for_padding_and_filters(tmp_path):
    p = tmp_path / "in_ws.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B"])
    ws.append(["x", "   "])   # B whitespace -> blank
    ws.append([None, "y"])    # A should fill to "x"
    wb.save(p)

    db_path = tmp_path / "o.db"
    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["A"],
        db=db_path,
        table="t",
        drop_blank_rows=False,
        require_non_null=["A"],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=10,
        fill_mode="hierarchical",
    )
    assert summary["rows_ingested"] == 2

    with sqlite3.connect(str(db_path)) as conn:
        rows = conn.execute("SELECT A, B FROM t ORDER BY rowid").fetchall()
    assert rows == [("x", None), ("x", "y")]


def test_drop_blank_rows_with_whitespace_only_fill_cols(tmp_path):
    p = tmp_path / "in_drop.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["grp", "val"])
    ws.append([None, "  "])
    ws.append(["", "   "])
    ws.append(["g1", "v1"])
    wb.save(p)

    db_path = tmp_path / "o2.db"
    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["grp"],
        db=db_path,
        table="t",
        drop_blank_rows=True,
        require_non_null=[],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=10,
        fill_mode="hierarchical",
    )
    assert summary["rows_ingested"] == 1
    with sqlite3.connect(str(db_path)) as conn:
        rows = conn.execute("SELECT grp, val FROM t").fetchall()
    assert rows == [("g1", "v1")]


def test_row_hash_matches_expected_with_float_padding(tmp_path):
    p = tmp_path / "in_hash.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["n", "txt"])
    ws.append([1.0, "a"])
    ws.append([None, "b"])  # will fill n down to 1.0
    wb.save(p)

    db_path = tmp_path / "hash.db"
    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["n"],
        db=db_path,
        table="t",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=True,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=10,
        fill_mode="hierarchical",
    )
    assert summary["rows_ingested"] == 2

    with sqlite3.connect(str(db_path)) as conn:
        rows = conn.execute("SELECT row_hash, n, txt FROM t ORDER BY rowid").fetchall()

    expected1 = sha256_hex(canon_list([1, "a"]))  # 1.0 canonical as '1'
    expected2 = sha256_hex(canon_list([1, "b"]))

    assert rows[0][0] == expected1
    assert rows[1][0] == expected2
    assert rows[0][2] == "a" and rows[1][2] == "b"


def test_xlsx_headers_include_row_hash_and_excel_row_in_order(tmp_path):
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "out_hash.xlsx"

    summary = ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        fill_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=True,
        excel_row_numbers=True,
        if_exists="replace",
        fill_mode="hierarchical",
    )
    assert summary["sheet"] == "Processed"

    wb = openpyxl.load_workbook(outfile)
    ws = wb["Processed"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["row_hash", "excel_row", "tier0", "tier1"]


def _make_tiered_workbook_for_pad_modes(tmp_path: Path):
    p = tmp_path / "padmodes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Tier 1", "Tier 2", "Tier 3", "Tier 4"])  # header row = 1
    ws.append(["apple",  None,   None,   None])          # 2: group header
    ws.append([None,     "red",  "sour", "value1"])      # 3: detail
    ws.append(["potato", None,   None,   None])          # 4: new group header
    ws.append([None,     None,   None,   "value2"])      # 5: Tier4 BEFORE T2/T3
    ws.append([None,     "fried","yellow","value3"])     # 6: detail
    wb.save(p)
    return p


def test_fill_mode_hierarchical_resets_lower_tiers_with_tier4(tmp_path):
    p = _make_tiered_workbook_for_pad_modes(tmp_path)
    db = tmp_path / "hier.db"

    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["Tier 1", "Tier 2", "Tier 3"],
        db=db,
        table="t",
        drop_blank_rows=True,
        require_non_null=["Tier 4"],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=50,
        fill_mode="hierarchical",
    )
    assert summary["rows_ingested"] == 3

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Tier 1","Tier 2","Tier 3","Tier 4" FROM t ORDER BY rowid').fetchall()

    assert rows == [
        ("apple",  "red",   "sour",   "value1"),
        ("potato", None,    None,     "value2"),   # lower tiers reset
        ("potato", "fried", "yellow", "value3"),
    ]


def test_fill_mode_independent_carries_lower_tiers_across_groups(tmp_path):
    p = _make_tiered_workbook_for_pad_modes(tmp_path)
    db = tmp_path / "indep.db"

    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["Tier 1", "Tier 2", "Tier 3"],
        db=db,
        table="t",
        drop_blank_rows=True,
        require_non_null=["Tier 4"],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=50,
        fill_mode="independent",     # new API flag
    )
    assert summary["rows_ingested"] == 3

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Tier 1","Tier 2","Tier 3","Tier 4" FROM t ORDER BY rowid').fetchall()

    assert rows == [
        ("apple",  "red",   "sour",   "value1"),
        ("potato", "red",   "sour",   "value2"),   # carried across Tier 1 change
        ("potato", "fried", "yellow", "value3"),
    ]


def test_duplicate_headers_raise(tmp_path):
    """Duplicate header names after normalization should error (core behavior)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "A"])  # duplicate headers
    ws.append(["x", "y"])
    p = tmp_path / "dup.xlsx"
    wb.save(p)
    wb.close()

    with pytest.raises(ValueError, match="Duplicate header names"):
        ingest_excel_to_sqlite(
            file=p, sheet="S", header_row=1, fill_cols=["A"], db=tmp_path / "dup.db"
        )


def test_missing_require_non_null_header_errors(tmp_path):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B"])
    ws.append(["x", "y"])
    wb.save(p)
    wb.close()

    with pytest.raises(ValueError, match="require_non_null header\\(s\\) not found"):
        ingest_excel_to_sqlite(
            file=p,
            sheet="S",
            header_row=1,
            fill_cols=["A"],
            db=tmp_path / "o.db",
            require_non_null=["C"],  # not present
        )


def test_header_row_not_one(tmp_path):
    p = tmp_path / "in_hdr3.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append([None, None])         # row 1 spacer
    ws.append([None, None])         # row 2 spacer
    ws.append(["H1", "H2"])         # row 3 header
    ws.append(["v1", None])         # row 4
    ws.append([None, "v2"])         # row 5
    wb.save(p)
    wb.close()

    db = tmp_path / "o_hdr3.db"
    summary = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=3, fill_cols=["H1"], db=db,
        row_hash=False, excel_row_numbers=True
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT excel_row, "H1", "H2" FROM "S" ORDER BY rowid').fetchall()
    assert rows == [("4", "v1", None), ("5", "v1", "v2")]


def test_sqlite_excel_row_is_text_type(tmp_path):
    p = tmp_path / "in_text.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    ws.append(["x"])
    wb.save(p)
    wb.close()

    db = tmp_path / "t.db"
    ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, fill_cols=["A"], db=db,
        excel_row_numbers=True
    )
    with sqlite3.connect(str(db)) as conn:
        info = conn.execute('PRAGMA table_info("S");').fetchall()
    names_and_types = [(r[1], r[2]) for r in info]
    assert names_and_types[0] == ("excel_row", "TEXT")
    assert names_and_types[1] == ("A", "TEXT")


def test_sqlite_append_exact_schema_match_succeeds(tmp_path):
    infile = _make_workbook(tmp_path)
    db = tmp_path / "samedb.db"

    s1 = ingest_excel_to_sqlite(
        file=infile, sheet="Sheet1", header_row=1, fill_cols=["tier0"],
        db=db, table="t", row_hash=True, excel_row_numbers=True,
        if_exists="replace", fill_mode="hierarchical"
    )
    assert s1["rows_ingested"] == 6

    s2 = ingest_excel_to_sqlite(
        file=infile, sheet="Sheet1", header_row=1, fill_cols=["tier0"],
        db=db, table="t", row_hash=True, excel_row_numbers=True,
        if_exists="append", fill_mode="hierarchical"
    )
    assert s2["rows_ingested"] == 6
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute("SELECT COUNT(*) FROM t;").fetchone()[0]
    assert n == 12


def test_sqlite_batch_size_one(tmp_path):
    infile = _make_workbook(tmp_path)
    db = tmp_path / "b1.db"
    s = ingest_excel_to_sqlite(
        file=infile, sheet="Sheet1", header_row=1, fill_cols=["tier0"],
        db=db, table="t", batch_size=1, if_exists="replace", fill_mode="hierarchical"
    )
    assert s["rows_ingested"] == 6
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute("SELECT COUNT(*) FROM t;").fetchone()[0]
    assert n == 6


def test_excel_if_exists_replace_recreates(tmp_path):
    infile = _make_workbook(tmp_path)
    out = tmp_path / "r.xlsx"

    ingest_excel_to_excel(
        file=infile, sheet="Sheet1", header_row=1, fill_cols=["tier0"],
        outfile=out, outsheet="Out", if_exists="fail", fill_mode="hierarchical"
    )
    ingest_excel_to_excel(
        file=infile, sheet="Sheet1", header_row=1, fill_cols=["tier0"],
        outfile=out, outsheet="Out", if_exists="replace", fill_mode="hierarchical"
    )

    wb = openpyxl.load_workbook(out)
    ws = wb["Out"]
    assert ws.max_row == 7  # header + 6 rows


def test_fill_cols_duplicates_are_deduped_preserving_first(tmp_path):
    """Duplicate names in fill_cols are interpreted once in order."""
    p = tmp_path / "in_dupcols.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])
    ws.append(["g1", None, "v1"])
    ws.append([None, "b1", "v2"])
    wb.save(p)
    wb.close()

    db = tmp_path / "o.db"
    summary = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1,
        fill_cols=["A", "B", "A", "B"],   # duplicates
        db=db, table="t", if_exists="replace", fill_mode="hierarchical"
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT A, B, C FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "v1"), ("g1", "b1", "v2")]


def test_require_non_null_after_padding_keeps_row(tmp_path):
    p = tmp_path / "in_req.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "Val"])
    ws.append(["g1", None])
    ws.append([None, "v1"])
    wb.save(p)
    wb.close()

    db = tmp_path / "o.db"
    s = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, fill_cols=["Grp"],
        db=db, table="t", require_non_null=["Grp"], if_exists="replace",
        fill_mode="hierarchical"
    )
    assert s["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None), ("g1", "v1")]


def test_sheet_not_found_raises(tmp_path):
    infile = _write_wb_simple(tmp_path, title="Good")
    with pytest.raises(ValueError, match="Sheet not found"):
        ingest_excel_to_sqlite(
            file=infile, sheet="BAD", header_row=1, fill_cols=["A"], db=tmp_path / "o.db"
        )


def test_all_blank_header_row_rejected(tmp_path):
    p = tmp_path / "in_blankhdr.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["  ", None, "nan", "   "])  # all blank/nan → normalize to empty
    ws.append(["x", "y", "z", "w"])
    wb.save(p)
    wb.close()

    with pytest.raises(ValueError, match="No non-empty headers"):
        ingest_excel_to_sqlite(
            file=p, sheet="S", header_row=1, fill_cols=["x"], db=tmp_path / "o.db"
        )


def test_new_workbook_has_no_extra_default_sheet(tmp_path):
    src = _mk_in(tmp_path)
    out = tmp_path / "fresh.xlsx"

    ingest_excel_to_excel(
        file=src, sheet="S", header_row=1, fill_cols=["A"],
        outfile=out, outsheet="Processed", if_exists="fail", fill_mode="hierarchical"
    )

    wb = openpyxl.load_workbook(out)
    try:
        assert wb.sheetnames == ["Processed"]
        ws = wb["Processed"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ["A"]
    finally:
        wb.close()


def test_excel_multiple_sheets_in_same_workbook(tmp_path):
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "multi.xlsx"

    ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        fill_cols=["tier0"],
        outfile=outfile,
        outsheet="S1",
        if_exists="fail",
        fill_mode="hierarchical",
    )

    ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        fill_cols=["tier0"],
        outfile=outfile,
        outsheet="S2",
        if_exists="fail",
        fill_mode="hierarchical",
    )

    wb = openpyxl.load_workbook(outfile)
    try:
        assert set(wb.sheetnames) == {"S1", "S2"}
        for name in ("S1", "S2"):
            ws = wb[name]
            assert ws.max_row == 7
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert header == ["tier0", "tier1"]
    finally:
        wb.close()


def test_date_canonicalization_and_hash(tmp_path):
    p = tmp_path / "dates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp","When","Txt"])
    d1 = datetime(2024, 5, 17, 13, 45, 0)
    ws.append(["g1", d1, "a"])
    ws.append([None, None, "b"])  # When should fill
    wb.save(p)
    wb.close()

    db = tmp_path / "dates.db"
    summary = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, fill_cols=["Grp","When"],
        db=db, table="t", row_hash=True, if_exists="replace",
        fill_mode="hierarchical"
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT row_hash, "Grp","When","Txt" FROM t ORDER BY rowid').fetchall()

    assert rows[0][2] == d1.isoformat()
    assert rows[1][2] == d1.isoformat()
    h1, h2 = rows[0][0], rows[1][0]
    assert h1 != h2


def test_fill_cols_letters_out_of_range_mentions_context(tmp_path):
    p = tmp_path / "oor.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Only"])
    ws.append(["x"])
    wb.save(p)
    wb.close()

    db = tmp_path / "oor.db"
    # Use CLI to validate error messaging path for letters mapping
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "ZZ",
        "--db", str(db),
    ], capture_output=True, text=True)
    assert r.returncode != 0
    msg = (r.stderr + r.stdout)
    assert "out of range" in msg and "header row 1" in msg and "only headered columns are ingested" in msg


def test_fill_mode_hierarchical_excel_writer(tmp_path):
    p = tmp_path / "padmodes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Tier 1","Tier 2","Tier 3","Tier 4"])
    ws.append(["apple",  None,   None,   None])
    ws.append([None,     "red",  "sour", "value1"])
    ws.append(["potato", None,   None,   None])
    ws.append([None,     None,   None,   "value2"])
    ws.append([None,     "fried","yellow","value3"])
    wb.save(p)
    wb.close()

    out = tmp_path / "o.xlsx"
    s = ingest_excel_to_excel(
        file=p, sheet="S", header_row=1,
        fill_cols=["Tier 1","Tier 2","Tier 3"],
        outfile=out, outsheet="O",
        drop_blank_rows=True, require_non_null=["Tier 4"],
        fill_mode="hierarchical", if_exists="replace"
    )
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["O"]
    data = [[c.value for c in r] for r in ws2.iter_rows(min_row=2, max_row=1+s["rows_written"])]
    assert data == [
        ["apple","red","sour","value1"],
        ["potato", None, None, "value2"],
        ["potato","fried","yellow","value3"],
    ]
    wb2.close()


def test_drop_blank_rows_whitespace_with_fill_flags(tmp_path):
    p = tmp_path / "wsdrop2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["grp","val"])
    ws.append(["   ", "x"])    # grp blank -> droppable when drop_blank_rows=True
    ws.append(["g1",  "y"])
    wb.save(p)
    wb.close()

    db = tmp_path / "wsdrop2.db"
    s = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, fill_cols=["grp"], db=db,
        drop_blank_rows=True, if_exists="replace", fill_mode="hierarchical"
    )
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT grp, val FROM "S" ORDER BY rowid').fetchall()
    assert rows == [("g1","y")]


# -----------------------------
# CLI sanity tests (unchanged behavior)
# -----------------------------

def test_cli_db_happy_path_prints_summary(tmp_path):
    infile = _write_wb_simple(tmp_path, title="Sheet1")
    db = tmp_path / "out.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "Sheet1",
        "--header-row", "1",
        "--fill-cols", '["A"]',
        "--db", str(db),
        "--table", "t",
        "--row-hash",
        "--excel-row-numbers",
        "--if-exists", "replace",
        "--batch-size", "2",
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    assert "Loaded 'Sheet1' →" in r.stdout
    assert "(cols=2 (+ excel_row, row_hash);" in r.stdout
    assert db.exists()
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute("SELECT COUNT(*) FROM t;").fetchone()[0]
    assert n == 3


def test_cli_xlsx_happy_path_prints_summary(tmp_path):
    infile = _write_wb_simple(tmp_path, title="S1")
    outfile = tmp_path / "out.xlsx"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "xlsx",
        "--infile", str(infile),
        "--insheet", "S1",
        "--header-row", "1",
        "--fill-cols", '["A"]',
        "--outfile", str(outfile),
        "--outsheet", "Processed",
        "--row-hash",
        "--excel-row-numbers",
        "--if-exists", "fail",
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    assert "Wrote 'S1' →" in r.stdout
    assert "(cols=2 (+ excel_row, row_hash);" in r.stdout
    wb = openpyxl.load_workbook(outfile)
    assert "Processed" in wb.sheetnames
    ws = wb["Processed"]
    assert ws.max_row == 4
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["row_hash", "excel_row", "A", "B"]


def test_cli_pad_cols_bad_json_exits(tmp_path):
    infile = _write_wb_simple(tmp_path, title="S")
    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols", 'not-a-json',
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "must be a valid JSON list of header names" in (r.stderr + r.stdout)


def test_cli_pad_cols_empty_list_exits(tmp_path):
    infile = _write_wb_simple(tmp_path, title="S")
    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols", "[]",
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "--fill-cols cannot be empty" in (r.stderr + r.stdout)


def test_cli_letters_happy_path_db(tmp_path):
    p = tmp_path / "in2.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Grp","Val","Note"])
    ws.append(["g1", None, "n1"])
    ws.append([None,"v2","n2"])
    wb.save(p); wb.close()

    db = tmp_path / "o2.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "A",     # => "Grp"
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val","Note" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "n1"), ("g1","v2","n2")]


def test_cli_letters_out_of_range_errors(tmp_path):
    p = tmp_path / "letters_oor.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])
    ws.append([1, 2, 3])
    wb.save(p); wb.close()

    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols-letters", "ZZ",
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    msg = (r.stderr + r.stdout)
    assert "out of range" in msg and "header row" in msg


def test_cli_letters_mutually_exclusive_errors(tmp_path):
    p = tmp_path / "letters_me.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "Val"])
    ws.append(["g1", "v1"])
    wb.save(p); wb.close()

    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols", '["Grp"]',
        "--fill-cols-letters", "A",
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "only one of --fill-cols or --fill-cols-letters" in (r.stderr + r.stdout)


def test_cli_header_row_zero_errors(tmp_path):
    infile = _mk_in(tmp_path)
    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "S",
        "--header-row", "0",
        "--fill-cols", '["A"]',
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "--header-row must be >= 1" in (r.stderr + r.stdout)


def test_cli_letters_header_row_exceeds_max_errors(tmp_path):
    p = tmp_path / "hdr_oor_letters.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    wb.save(p); wb.close()

    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "999",
        "--fill-cols-letters", "A",
        "--db", str(tmp_path / "o.db"),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "exceeds sheet 'S' max row" in (r.stderr + r.stdout)


def test_cli_no_subcommand_shows_help(tmp_path):
    r = subprocess.run([sys.executable, "-m", "xlfilldown.cli"],
                       capture_output=True, text=True)
    assert r.returncode == 0
    assert "db" in r.stdout and "xlsx" in r.stdout


def test_cli_version_flag(tmp_path):
    r = subprocess.run([sys.executable, "-m", "xlfilldown.cli", "--version"],
                       capture_output=True, text=True)
    assert r.returncode == 0
    assert r.stdout.strip()

# --- New tests for API-only letters support (no CLI involved) ---

def _mk_letters_wb(tmp_path: Path, title="S"):
    p = tmp_path / "letters_api.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["Grp", "Val", "Note"])   # A, B, C
    ws.append(["g1", None, "n1"])
    ws.append([None, "v2", "n2"])
    wb.save(p); wb.close()
    return p


def test_api_fill_cols_letters_sqlite_happy_path(tmp_path):
    """fill_cols_letters maps letters → headers for the API (SQLite path)."""
    infile = _mk_letters_wb(tmp_path)
    db = tmp_path / "letters_api.db"

    summary = ingest_excel_to_sqlite(
        file=infile,
        sheet="S",
        header_row=1,
        fill_cols_letters=["A"],     # => "Grp"
        db=db,
        table="t",
        if_exists="replace",
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val","Note" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "n1"), ("g1", "v2", "n2")]  # "Grp" filled down


def test_api_fill_cols_letters_excel_happy_path(tmp_path):
    """fill_cols_letters also works for the Excel writer path."""
    infile = _mk_letters_wb(tmp_path)
    out = tmp_path / "letters_api_out.xlsx"

    s = ingest_excel_to_excel(
        file=infile,
        sheet="S",
        header_row=1,
        fill_cols_letters=["A"],     # => "Grp"
        outfile=out,
        outsheet="Processed",
        if_exists="replace",
    )
    assert s["rows_written"] == 2
    wb = openpyxl.load_workbook(out)
    ws = wb["Processed"]
    data = [[c.value for c in r] for r in ws.iter_rows(min_row=2, max_row=3)]
    assert data == [["g1", None, "n1"], ["g1", "v2", "n2"]]
    wb.close()


def test_api_require_non_null_letters_merges_with_names(tmp_path):
    """
    require_non_null_letters should merge (order-preserving, de-duped) with require_non_null (names).
    Behavior-wise, row must satisfy the final merged set after padding.
    """
    p = tmp_path / "req_letters_merge.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "A", "B"])  # A,B,C letters → Grp,A,B
    ws.append([None, "x", None])  # drop: Grp missing (A present)
    ws.append(["g1", None, "y"])  # keep: Grp present & B present
    wb.save(p); wb.close()

    db = tmp_path / "req_letters_merge.db"
    ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["Grp"],
        require_non_null=["B"],              # name
        require_non_null_letters=["A"],      # letter "A" -> "Grp"
        db=db,
        table="t",
        if_exists="replace",
    )
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","A","B" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "y")]      # first row dropped; second kept


def test_api_fill_cols_and_fill_cols_letters_mutually_exclusive(tmp_path):
    infile = _mk_letters_wb(tmp_path)
    with pytest.raises(ValueError, match="only one of fill_cols or fill_cols_letters"):
        ingest_excel_to_sqlite(
            file=infile,
            sheet="S",
            header_row=1,
            fill_cols=["Grp"],
            fill_cols_letters=["A"],
            db=tmp_path / "me.db",
        )


def test_api_fill_cols_letters_empty_header_errors(tmp_path):
    """Letters pointing at an empty/whitespace header cell should raise a clear error."""
    p = tmp_path / "empty_hdr.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "   ", "C"])  # column B header normalizes to empty
    ws.append(["x", "y", "z"])
    wb.save(p); wb.close()

    with pytest.raises(ValueError, match="refers to an empty header cell"):
        ingest_excel_to_sqlite(
            file=p,
            sheet="S",
            header_row=1,
            fill_cols_letters=["B"],      # points to empty header after normalization
            db=tmp_path / "eh.db",
        )


def test_api_fill_cols_letters_out_of_range_errors(tmp_path):
    """Letters beyond header width should raise with context (row/sheet/width hint)."""
    p = tmp_path / "oor_hdr.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Only"])
    ws.append(["x"])
    wb.save(p); wb.close()

    with pytest.raises(ValueError) as ei:
        ingest_excel_to_sqlite(
            file=p,
            sheet="S",
            header_row=1,
            fill_cols_letters=["ZZ"],   # out of range
            db=tmp_path / "oor.db",
        )
    msg = str(ei.value)
    assert "out of range" in msg and "header row 1" in msg and "only headered columns are ingested" in msg


def test_api_require_non_null_letters_dedup_and_order(tmp_path):
    """
    If the same header is specified in both names and letters, final set should de-dup,
    but behavior should still enforce all unique headers.
    """
    p = tmp_path / "req_dedup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["H1", "H2", "V"])
    ws.append([None, None, "x"])    # drop: H1 and H2 missing
    ws.append(["a", None, "y"])     # drop: H2 missing
    ws.append(["a", "b", "z"])      # keep
    wb.save(p); wb.close()

    db = tmp_path / "req_dedup.db"
    ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        fill_cols=["H1", "H2"],
        require_non_null=["H2", "H1", "H2"],     # names with duplicates
        require_non_null_letters=["A"],          # "H1" again via letter
        db=db,
        table="t",
        if_exists="replace",
    )
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "H1","H2","V" FROM t ORDER BY rowid').fetchall()
    assert rows == [("a", "b", "z")]  # only the row satisfying both H1 & H2 remains


# -----------------------------
# Raw ingest mode tests
# -----------------------------

def _mk_raw_wb(tmp_path: Path, title="S"):
    p = tmp_path / "raw_in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["A", "B"])            # header row
    ws.append(["x", None])           # row 2
    ws.append([None, "y"])           # row 3
    ws.append([None, None])          # row 4 (completely empty)
    wb.save(p); wb.close()
    return p


def test_raw_sqlite_no_fill_with_audit_columns(tmp_path):
    infile = _mk_raw_wb(tmp_path)
    db = tmp_path / "raw.db"

    summary = ingest_excel_to_sqlite(
        file=infile, sheet="S", header_row=1,
        ingest_mode="raw",                  # <-- NEW
        db=db, table="t",
        row_hash=True, excel_row_numbers=True,
        if_exists="replace",
    )
    assert summary["rows_ingested"] == 3  # rows 2,3,4 (we do not drop blank by default)
    assert summary["row_hash"] is True and summary["excel_row_numbers"] is True

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT row_hash, excel_row, "A","B" FROM t ORDER BY excel_row').fetchall()

    # No fill-down should occur
    # Row 2: ["x", None]
    expected2 = sha256_hex(canon_list(["x", None]))
    # Row 3: [None, "y"]
    expected3 = sha256_hex(canon_list([None, "y"]))
    # Row 4: [None, None] (completely empty row)
    expected4 = sha256_hex(canon_list([None, None]))

    assert rows == [
        (expected2, "2", "x", None),
        (expected3, "3", None, "y"),
        (expected4, "4", None, None),
    ]


def test_raw_excel_no_fill_with_audit_columns(tmp_path):
    infile = _mk_raw_wb(tmp_path)
    outfile = tmp_path / "raw.xlsx"

    summary = ingest_excel_to_excel(
        file=infile, sheet="S", header_row=1,
        ingest_mode="raw",                  # <-- NEW
        outfile=outfile, outsheet="RawOut",
        row_hash=True, excel_row_numbers=True,
        if_exists="replace",
    )
    assert summary["sheet"] == "RawOut"
    assert summary["rows_written"] == 3  # rows 2,3,4

    wb = openpyxl.load_workbook(outfile)
    try:
        ws = wb["RawOut"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ["row_hash", "excel_row", "A", "B"]

        data = [[c.value for c in r] for r in ws.iter_rows(min_row=2, max_row=4)]
        # Recompute expected hashes
        h2 = sha256_hex(canon_list(["x", None]))
        h3 = sha256_hex(canon_list([None, "y"]))
        h4 = sha256_hex(canon_list([None, None]))
        assert data == [
            [h2, "2", "x", None],
            [h3, "3", None, "y"],
            [h4, "4", None, None],
        ]
    finally:
        wb.close()


def test_raw_mode_drop_blank_and_require_non_null(tmp_path):
    # Build a sheet where require_non_null and drop_blank_rows matter
    p = tmp_path / "raw_filters.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "Val"])
    ws.append([None, None])   # row 2: completely empty -> dropped by drop_blank_rows=True
    ws.append(["g1", None])   # row 3: kept unless require_non_null=["Val"]
    ws.append([None, "v1"])   # row 4: kept unless require_non_null=["Grp"]
    ws.append(["g2", "v2"])   # row 5: always kept
    wb.save(p); wb.close()

    db = tmp_path / "raw_filters.db"
    # In raw mode, padding is off. Filters still apply AFTER "no padding".
    s = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1,
        ingest_mode="raw",
        db=db, table="t",
        drop_blank_rows=True,                 # drop the completely empty row 2
        require_non_null=["Grp","Val"],       # keep only rows where both present (rows 5)
        if_exists="replace",
    )
    assert s["rows_ingested"] == 1

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val" FROM t').fetchall()
    assert rows == [("g2", "v2")]


def test_raw_mode_ignores_fill_cols_and_fill_mode(tmp_path):
    infile = _mk_raw_wb(tmp_path)
    db = tmp_path / "raw_ignore.db"

    # Even if we pass fill options, raw mode must not fill anything
    s = ingest_excel_to_sqlite(
        file=infile, sheet="S", header_row=1,
        ingest_mode="raw",
        fill_cols=["A", "B"],                 # should be ignored
        fill_mode="hierarchical",             # should be ignored
        db=db, table="t",
        if_exists="replace",
    )
    assert s["rows_ingested"] == 3

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "A","B" FROM t ORDER BY rowid').fetchall()
    # No fill-down: matches source rows 2..4 exactly
    assert rows == [("x", None), (None, "y"), (None, None)]


def test_cli_raw_db_no_fillcols_required(tmp_path):
    infile = _mk_raw_wb(tmp_path, title="Sheet1")
    db = tmp_path / "cli_raw.db"

    # No --fill-cols / --fill-cols-letters needed in raw mode
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "Sheet1",
        "--header-row", "1",
        "--db", str(db),
        "--table", "t",
        "--ingest-mode", "raw",
        "--if-exists", "replace",
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "A","B" FROM t ORDER BY rowid').fetchall()
    assert rows == [("x", None), (None, "y"), (None, None)]


def test_cli_raw_xlsx_no_fillcols_required(tmp_path):
    infile = _mk_raw_wb(tmp_path, title="S1")
    out = tmp_path / "cli_raw.xlsx"

    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "xlsx",
        "--infile", str(infile),
        "--insheet", "S1",
        "--header-row", "1",
        "--outfile", str(out),
        "--outsheet", "RawOut",
        "--ingest-mode", "raw",
        "--if-exists", "replace",
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr

    wb = openpyxl.load_workbook(out)
    try:
        ws = wb["RawOut"]
        data = [[c.value for c in r] for r in ws.iter_rows(min_row=2, max_row=4)]
        assert data == [["x", None], [None, "y"], [None, None]]
    finally:
        wb.close()















