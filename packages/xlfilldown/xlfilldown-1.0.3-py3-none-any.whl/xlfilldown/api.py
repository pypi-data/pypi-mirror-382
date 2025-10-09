# xlfilldown/api.py
from __future__ import annotations

from typing import Iterable, Optional, Sequence, Union, Any
from os import PathLike

import openpyxl

# Re-export public helpers from core
from .core import (
    qident,
    normalize_headers,
    canon_list,
    sha256_hex,
)

# Import the existing core ingest functions (which still take pad_* params)
from .core import ingest_excel_to_sqlite as _core_ingest_excel_to_sqlite
from .core import ingest_excel_to_excel as _core_ingest_excel_to_excel


__all__ = [
    "ingest_excel_to_sqlite",
    "ingest_excel_to_excel",
    "qident",
    "normalize_headers",
    "canon_list",
    "sha256_hex",
]


def _resolve_fill_mode(
    *,
    fill_mode: Optional[str] = None,
    pad_hierarchical: Optional[bool] = None,
) -> bool:
    """
    Returns the boolean value expected by core.pad_hierarchical.

    Priority:
      1) explicit pad_hierarchical (for backward-compat)
      2) fill_mode string ("hierarchical" | "independent") — default is "hierarchical"
    """
    if pad_hierarchical is not None:
        return bool(pad_hierarchical)

    if fill_mode is None:
        return True  # default to hierarchical

    fm = str(fill_mode).strip().lower()
    if fm in ("hierarchical", "hier", "h"):
        return True
    if fm in ("independent", "ind", "i", "legacy"):
        return False
    raise ValueError("fill_mode must be 'hierarchical' or 'independent'.")


def _excel_letters_to_index(letter: str) -> int:
    """Convert 'A' -> 1, 'B' -> 2, ..., 'Z'->26, 'AA'->27, etc."""
    s = letter.strip().upper()
    if not s.isalpha():
        raise ValueError(f"Invalid Excel column letter: {letter!r}")
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _read_normalized_headers(
    *,
    file: Union[str, PathLike[str]],
    sheet: str,
    header_row: int,
) -> Sequence[str]:
    """Open workbook in read-only mode and return normalized headers for the header_row."""
    try:
        wb = openpyxl.load_workbook(filename=file, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to open workbook: {e}") from e

    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet!r}")
        ws = wb[sheet]
        max_row = ws.max_row or 0
        if header_row < 1:
            raise ValueError("Header row must be >= 1.")
        if header_row > max_row:
            raise ValueError(
                f"Header row {header_row} exceeds sheet '{sheet}' max row ({max_row})."
            )

        # Read the header row across all used columns
        max_col = ws.max_column or 0
        raw = []
        for c in range(1, max_col + 1):
            raw.append(ws.cell(row=header_row, column=c).value)
        return normalize_headers(raw)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _letters_to_headers(
    *,
    letters: Sequence[str],
    headers: Sequence[str],
    sheet: str,
    header_row: int,
) -> list[str]:
    """Map Excel column letters to header names with validation."""
    result: list[str] = []
    width = len(headers)
    for L in letters:
        idx = _excel_letters_to_index(L)
        if idx < 1 or idx > width:
            raise ValueError(
                f"Column {L!r} is out of range for header row {header_row} on sheet '{sheet}' "
                f"(width={width}). Hint: only headered columns are ingested."
            )
        name = headers[idx - 1]
        if not name:
            raise ValueError(
                f"Column {L!r} refers to an empty header cell on sheet '{sheet}' (row {header_row})."
            )
        result.append(name)
    return result


def _merge_unique(primary: Sequence[str], secondary: Sequence[str]) -> list[str]:
    """Merge two lists, keeping order and de-duplicating (primary wins on ordering)."""
    out: list[str] = []
    seen: set[str] = set()
    for src in (primary, secondary):
        for x in src:
            if x not in seen:
                out.append(x)
                seen.add(x)
    return out


def _resolve_fill_cols(
    *,
    file: Union[str, PathLike[str]],
    sheet: str,
    header_row: int,
    fill_cols: Optional[Sequence[str]] = None,
    fill_cols_letters: Optional[Sequence[str]] = None,
    pad_cols: Optional[Sequence[str]] = None,
    ingest_mode: Optional[str] = None,
) -> Sequence[str]:
    """
    Returns the list of column names to fill.

    Modes:
      - ingest_mode == "raw"  -> skip fill entirely (return [])
      - default ("fill")      -> resolve as before

    Resolution priority in "fill" mode:
      - NEW API: fill_cols OR fill_cols_letters (mutually exclusive)
      - OLD API: pad_cols (names only)
      - Otherwise: []
    """
    if ingest_mode and str(ingest_mode).strip().lower() == "raw":
        return []

    if fill_cols is not None and fill_cols_letters is not None:
        raise ValueError("Use only one of fill_cols or fill_cols_letters.")

    if fill_cols_letters is not None:
        headers = _read_normalized_headers(file=file, sheet=sheet, header_row=header_row)
        return _letters_to_headers(
            letters=fill_cols_letters, headers=headers, sheet=sheet, header_row=header_row
        )

    if fill_cols is not None:
        return list(fill_cols)

    if pad_cols is not None:
        return list(pad_cols)

    # Let core handle empty/missing errors in its usual way.
    return []



def _resolve_require_non_null(
    *,
    file: Union[str, PathLike[str]],
    sheet: str,
    header_row: int,
    require_non_null: Optional[Sequence[str]] = None,
    require_non_null_letters: Optional[Sequence[str]] = None,
) -> Sequence[str]:
    """
    Merge require_non_null names + letters→names (order-preserving, de-duped).
    """
    names = list(require_non_null or [])
    if require_non_null_letters:
        headers = _read_normalized_headers(file=file, sheet=sheet, header_row=header_row)
        letter_names = _letters_to_headers(
            letters=require_non_null_letters, headers=headers, sheet=sheet, header_row=header_row
        )
        names = _merge_unique(names, letter_names)
    return names


def ingest_excel_to_sqlite(
    *,
    file: Union[str, PathLike[str]],
    sheet: str,
    header_row: int,
    # NEW API
    fill_cols: Optional[Sequence[str]] = None,
    fill_cols_letters: Optional[Sequence[str]] = None,
    fill_mode: Optional[str] = None,
    require_non_null_letters: Optional[Sequence[str]] = None,
    ingest_mode: str = "fill",   # <-- NEW
    # OLD API (aliases)
    pad_cols: Optional[Sequence[str]] = None,
    pad_hierarchical: Optional[bool] = None,
    # passthrough options
    db: Union[str, PathLike[str]],
    table: Optional[str] = None,
    drop_blank_rows: bool = False,
    require_non_null: Optional[Sequence[str]] = None,
    row_hash: bool = False,
    excel_row_numbers: bool = False,
    if_exists: str = "fail",
    batch_size: int = 1000,
    **kwargs: Any,
) -> dict:
    """
    API wrapper that mirrors the CLI terms (fill_cols/_letters, fill_mode) while remaining
    backward-compatible with pad_cols / pad_hierarchical.

    ingest_mode:
      - "fill" (default): existing behavior (may require fill columns upstream)
      - "raw": skip all fill/down-carry logic (pad_cols becomes [])
    """
    cols = _resolve_fill_cols(
        file=file,
        sheet=sheet,
        header_row=header_row,
        fill_cols=fill_cols,
        fill_cols_letters=fill_cols_letters,
        pad_cols=pad_cols,
        ingest_mode=ingest_mode,   # NEW
    )
    hierarchical = _resolve_fill_mode(fill_mode=fill_mode, pad_hierarchical=pad_hierarchical)
    req_non_null = _resolve_require_non_null(
        file=file,
        sheet=sheet,
        header_row=header_row,
        require_non_null=require_non_null,
        require_non_null_letters=require_non_null_letters,
    )

    # Forward to core using its expected parameter names
    return _core_ingest_excel_to_sqlite(
        file=file,
        sheet=sheet,
        header_row=header_row,
        pad_cols=cols,
        db=db,
        table=table,
        drop_blank_rows=drop_blank_rows,
        require_non_null=req_non_null,
        row_hash=row_hash,
        excel_row_numbers=excel_row_numbers,
        if_exists=if_exists,
        batch_size=batch_size,
        pad_hierarchical=hierarchical,
        **kwargs,
    )



def ingest_excel_to_excel(
    *,
    file: Union[str, PathLike[str]],
    sheet: str,
    header_row: int,
    # NEW API
    fill_cols: Optional[Sequence[str]] = None,
    fill_cols_letters: Optional[Sequence[str]] = None,
    fill_mode: Optional[str] = None,
    require_non_null_letters: Optional[Sequence[str]] = None,
    ingest_mode: str = "fill",   # <-- NEW
    # OLD API (aliases)
    pad_cols: Optional[Sequence[str]] = None,
    pad_hierarchical: Optional[bool] = None,
    # passthrough options
    outfile: Union[str, PathLike[str]],
    outsheet: str,
    drop_blank_rows: bool = False,
    require_non_null: Optional[Sequence[str]] = None,
    row_hash: bool = False,
    excel_row_numbers: bool = False,
    if_exists: str = "fail",
    **kwargs: Any,
) -> dict:
    """
    API wrapper that mirrors the CLI terms (fill_cols/_letters, fill_mode) while remaining
    backward-compatible with pad_cols / pad_hierarchical.

    ingest_mode:
      - "fill" (default): existing behavior
      - "raw": skip all fill/down-carry logic (pad_cols becomes [])
    """
    cols = _resolve_fill_cols(
        file=file,
        sheet=sheet,
        header_row=header_row,
        fill_cols=fill_cols,
        fill_cols_letters=fill_cols_letters,
        pad_cols=pad_cols,
        ingest_mode=ingest_mode,   # NEW
    )
    hierarchical = _resolve_fill_mode(fill_mode=fill_mode, pad_hierarchical=pad_hierarchical)
    req_non_null = _resolve_require_non_null(
        file=file,
        sheet=sheet,
        header_row=header_row,
        require_non_null=require_non_null,
        require_non_null_letters=require_non_null_letters,
    )

    return _core_ingest_excel_to_excel(
        file=file,
        sheet=sheet,
        header_row=header_row,
        pad_cols=cols,
        outfile=outfile,
        outsheet=outsheet,
        drop_blank_rows=drop_blank_rows,
        require_non_null=req_non_null,
        row_hash=row_hash,
        excel_row_numbers=excel_row_numbers,
        if_exists=if_exists,
        pad_hierarchical=hierarchical,
        **kwargs,
    )



