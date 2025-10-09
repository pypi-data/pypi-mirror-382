from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import List, Optional

from . import __version__
from .api import ingest_excel_to_sqlite, ingest_excel_to_excel, normalize_headers

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def _resolve_headers_from_letters(
    *,
    infile: str,
    insheet: str,
    header_row: int,
    letters: List[str],
) -> List[str]:
    """
    Map Excel column letters (A, B, AE, ...) to normalized header names found on `header_row`.
    - Validates the sheet exists.
    - Validates the header_row is within bounds (>=1 and <= sheet max row).
    - Validates each letter is a valid Excel column label and within the header row's width.
    - Rejects empty headers (None / whitespace / 'nan' → normalize to '') with a clear error.
    Returns a list of resolved header names in the order of the provided letters.
    """
    if header_row < 1:
        raise SystemExit(f"Header row must be >= 1 (got {header_row}).")

    wb = load_workbook(filename=str(infile), read_only=True, data_only=True)
    try:
        if insheet not in wb.sheetnames:
            raise SystemExit(f"Sheet not found: {insheet} (available: {wb.sheetnames})")
        ws = wb[insheet]

        max_row = getattr(ws, "max_row", 0) or 0
        if header_row > max_row:
            raise SystemExit(
                f"Header row {header_row} exceeds sheet '{insheet}' max row {max_row}."
            )

        # Read raw header row and normalize exactly like the engine does
        header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers_norm = normalize_headers(header_cells)
        n_cols = len(headers_norm)

        resolved: List[str] = []
        for raw in letters:
            letter = (raw or "").strip()
            # Validate the label itself
            try:
                idx = column_index_from_string(letter) - 1  # A -> 0
            except Exception:
                raise SystemExit(f"Invalid column letter: {raw!r}")

            # Bounds check against header row width
            if idx < 0 or idx >= n_cols:
                raise SystemExit(
                    f"Column letter {letter} is out of range on header row {header_row} "
                    f"(header row has {n_cols} column(s); only headered columns are ingested)."
                )

            # Enforce non-empty header at that position after normalization
            header_name = headers_norm[idx]
            if header_name == "":
                raise SystemExit(
                    f"Column letter {letter} refers to an empty header cell on row {header_row}. "
                    f"Use a column with a non-empty header, or pass --fill-cols with explicit header names."
                )

            resolved.append(header_name)

        return resolved
    finally:
        try:
            wb.close()
        except Exception:
            pass




def _parse_json_list(arg_val: Optional[str], flag_name: str) -> List[str]:
    if not arg_val:
        return []
    try:
        parsed = json.loads(arg_val)
        if not isinstance(parsed, list):
            raise ValueError
        return [str(h) for h in parsed]
    except Exception:
        raise SystemExit(
            f"{flag_name} must be a valid JSON list of header names, e.g. '[\"columnname1\",\"columnname2\",\"columnname,4\"]'"
        )


def _shared_args(sub: argparse.ArgumentParser) -> None:
    sub.add_argument("--infile", required=True, help="Input .xlsx file")
    sub.add_argument("--insheet", required=True, help="Sheet name to read")
    sub.add_argument(
        "--header-row",
        type=int,
        required=True,
        help="Header row number as seen in Excel (1-based)",
    )
    sub.add_argument(
        "--fill-cols",
        required=False,
        help='JSON array of header names to forward-fill. Example: \'["columnname1","columnname2","columnname,4"]\'',
    )
    sub.add_argument(
        "--fill-cols-letters",
        nargs="+",
        help="Excel column letters to forward-fill (e.g. A B C AE). "
             "Resolved to header names using --header-row.",
    )
    sub.add_argument(
        "--fill-mode",
        choices=["hierarchical", "independent"],
        default="hierarchical",
        help=(
            "Fill-down mode for --fill-cols. "
            "'hierarchical' (default) resets lower tier columns when a higher tier column changes. "
            "'independent' carries each column separately (pandas-style ffill)."
        ),
    )
    sub.add_argument(
        "--drop-blank-rows",
        action="store_true",
        help="Drop rows where ALL fill columns are empty AFTER padding (true spacer rows).",
    )
    sub.add_argument(
        "--require-non-null",
        help='JSON array of header names; drop row if ANY are null/blank AFTER padding. Example: \'["columnname1","columnname2"]\'',
    )
    sub.add_argument(
        "--require-non-null-letters",
        nargs="+",
        help="Excel column letters for required-non-null (e.g. A D). "
             "Resolved to header names using --header-row, then merged with --require-non-null.",
    )
    sub.add_argument(
        "--row-hash",
        action="store_true",
        help="Include a NON-UNIQUE indexed row_hash column (DB) / write a row_hash column (Excel).",
    )
    sub.add_argument(
        "--excel-row-numbers",
        action="store_true",
        help="Include Excel row numbers in the output (column 'excel_row').",
    )
    sub.add_argument(
        "--if-exists",
        choices=["fail", "replace", "append"],
        default="fail",
        help="Behavior if the destination already exists (DB table or Excel sheet).",
    )
    sub.add_argument(
        "--ingest-mode",
        choices=["fill", "raw"],
        default="fill",
        help=(
            "Ingestion strategy. 'fill' (default) applies forward-fill per --fill-cols/letters "
            "and --fill-mode. 'raw' skips fill-down entirely and writes rows as-is while still "
            "supporting --row-hash, --excel-row-numbers, --drop-blank-rows, and --require-non-null."
        ),
    )




def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Fill-down ETL: stream an Excel sheet to SQLite or Excel, "
            "with forward-fill for named columns, optional Excel row numbers, and a stable row hash."
        )
    )
    ap.add_argument("-V", "--version", action="version", version=f"%(prog)s {__version__}")

    subparsers = ap.add_subparsers(dest="command")

    # DB subcommand
    ap_db = subparsers.add_parser(
        "db",
        help="Write output to SQLite database",
        description="Stream an Excel sheet into SQLite with fill-down, optional row hash and excel_row.",
    )
    _shared_args(ap_db)
    ap_db.add_argument("--db", required=True, help="SQLite database path (created if missing)")
    ap_db.add_argument("--table", help="SQLite table name (default: derived from sheet name)")
    ap_db.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Number of rows per executemany() batch",
    )

    # Excel subcommand
    ap_xlsx = subparsers.add_parser(
        "xlsx",
        help="Write output to Excel file",
        description="Stream an Excel sheet into a new or existing Excel workbook/sheet with fill-down.",
    )
    _shared_args(ap_xlsx)
    ap_xlsx.add_argument("--outfile", required=True, help="Output .xlsx file")
    ap_xlsx.add_argument("--outsheet", help="Output sheet name (default: derived from input sheet name)")

    args = ap.parse_args()

    # If no subcommand was given, show help and exit 0
    if args.command is None:
        ap.print_help()
        return

    # Guard: header row must be ≥ 1
    if getattr(args, "header_row", None) is not None and args.header_row < 1:
        raise SystemExit("--header-row must be >= 1.")

    # Resolve fill columns depending on ingest mode
    if args.ingest_mode == "raw":
        # Raw mode: ignore any --fill-cols / --fill-cols-letters; no fill-down
        pad_cols = []
    else:
        # Fill mode: require either names (JSON) or letters
        if getattr(args, "fill_cols_letters", None) and args.fill_cols:
            raise SystemExit("Use only one of --fill-cols or --fill-cols-letters, not both.")

        if getattr(args, "fill_cols_letters", None):
            pad_cols = _resolve_headers_from_letters(
                infile=args.infile,
                insheet=args.insheet,
                header_row=args.header_row,
                letters=args.fill_cols_letters,
            )
        elif args.fill_cols is not None:
            pad_cols = _parse_json_list(args.fill_cols, "--fill-cols")
            if not pad_cols:
                raise SystemExit("--fill-cols cannot be empty; provide at least one header name.")
        else:
            # Neither option provided in fill mode
            raise SystemExit(
                "You must provide either --fill-cols (JSON header names) or --fill-cols-letters (Excel column letters) "
                "when --ingest-mode=fill. Example: --fill-cols '[\"Tier 1\",\"Tier 2\"]'  or  --fill-cols-letters A C AE\n"
                "Tip: use --ingest-mode raw to skip fill-down entirely."
            )


    # Resolve required_non_null by names + letters (additive)
    required_non_null = _parse_json_list(args.require_non_null, "--require-non-null") if args.require_non_null else []
    if getattr(args, "require_non_null_letters", None):
        required_non_null += _resolve_headers_from_letters(
            infile=args.infile,
            insheet=args.insheet,
            header_row=args.header_row,
            letters=args.require_non_null_letters,
        )
    # De-dup while preserving order
    seen_r = set()
    required_non_null = [h for h in required_non_null if not (h in seen_r or seen_r.add(h))]

    pad_hierarchical = (args.fill_mode == "hierarchical")

    if args.command == "db":
        summary = ingest_excel_to_sqlite(
            file=Path(args.infile),
            sheet=args.insheet,
            header_row=args.header_row,
            pad_cols=pad_cols,
            db=Path(args.db),
            table=args.table,
            drop_blank_rows=args.drop_blank_rows,
            require_non_null=required_non_null,
            row_hash=args.row_hash,
            excel_row_numbers=args.excel_row_numbers,
            if_exists=args.if_exists,
            batch_size=args.batch_size,
            pad_hierarchical=pad_hierarchical,
            ingest_mode=args.ingest_mode,
        )

        cols = len(summary["columns"])
        extras = []
        if summary.get("excel_row_numbers"):
            extras.append("excel_row")
        if summary.get("row_hash"):
            extras.append("row_hash")
        extras_str = f" (+ {', '.join(extras)})" if extras else ""
        print(
            f"✅ Loaded '{args.insheet}' → {args.db}:{summary['table']} "
            f"(cols={cols}{extras_str}; rows={summary['rows_ingested']}; if_exists={args.if_exists})"
        )
        return

    if args.command == "xlsx":
        summary = ingest_excel_to_excel(
            file=Path(args.infile),
            sheet=args.insheet,
            header_row=args.header_row,
            pad_cols=pad_cols,
            outfile=Path(args.outfile),
            outsheet=args.outsheet,
            drop_blank_rows=args.drop_blank_rows,
            require_non_null=required_non_null,
            row_hash=args.row_hash,
            excel_row_numbers=args.excel_row_numbers,
            if_exists=args.if_exists,
            pad_hierarchical=pad_hierarchical,
            ingest_mode=args.ingest_mode,
        )

        cols = len(summary["columns"])
        extras = []
        if summary.get("excel_row_numbers"):
            extras.append("excel_row")
        if summary.get("row_hash"):
            extras.append("row_hash")
        extras_str = f" (+ {', '.join(extras)})" if extras else ""
        print(
            f"✅ Wrote '{args.insheet}' → {args.outfile}:{summary['sheet']} "
            f"(cols={cols}{extras_str}; rows={summary['rows_written']}; if_exists={args.if_exists})"
        )
        return





if __name__ == "__main__":
    main()


