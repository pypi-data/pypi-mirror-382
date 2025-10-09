from __future__ import annotations

import json
import sqlite3
import hashlib
from pathlib import Path
from typing import List, Optional, Sequence, Iterable, Dict, Any, Tuple
from datetime import datetime, date
from decimal import Decimal, InvalidOperation



from openpyxl import load_workbook, Workbook


def qident(name: str) -> str:
    """Quote a SQLite identifier (double-quotes; escape inner quotes)."""
    return '"' + name.replace('"', '""') + '"'


def normalize_headers(cells: Sequence[Optional[str]]) -> List[str]:
    """Trim whitespace; coerce None/blank/'nan' to empty strings."""
    out: List[str] = []
    for c in cells:
        if c is None:
            out.append("")
        else:
            s = str(c).strip()
            out.append("" if not s or s.lower() == "nan" else s)
    return out

def _decimal_to_canonical_str(d: Decimal) -> str:
    """
    Convert a Decimal to a stable, non-scientific string with no trailing zeros.
    Examples:
      1        -> "1"
      1.0      -> "1"
      2.50     -> "2.5"
      1000000  -> "1000000"
      -0.0     -> "0"
    """
    # For NaN/inf (unlikely from Excel), fall back to str()
    if not d.is_finite():
        return str(d)

    # Remove exponent while keeping exact value
    d = d.normalize()
    s = format(d, "f")  # plain decimal string, no scientific notation

    # Strip trailing zeros and any dangling dot
    if "." in s:
        s = s.rstrip("0").rstrip(".")

    # Normalize signed zero to "0"
    if s in ("-0", "+0", ""):
        s = "0"
    return s


def _is_blank_cell(v):
    """True if v should be considered blank/empty for padding & filtering logic."""
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False

def _strip_for_storage(v):
    """
    Normalize values for storage as TEXT:
      - None stays None (→ NULL in SQLite / empty cell in Excel).
      - strings: .strip(); empty string → None.
      - non-strings: converted to canonical string (matches hashing canonicalization).
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return None if s == "" else s
    # Use the same canonicalization as hashing to keep DB/Excel text stable
    return _canon_scalar(v)





def _canon_scalar(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return _decimal_to_canonical_str(v)
    if isinstance(v, float):
        # Convert via Decimal for stable text (avoids binary float quirks & sci notation)
        try:
            d = Decimal(str(v))
            return _decimal_to_canonical_str(d)
        except (InvalidOperation, ValueError):
            # Fallback if something odd slips through
            return str(v)
    # Ints, bools, etc. (bool is subclass of int): stable string already
    return str(v)


def canon_list(values: Sequence[Any]) -> str:
    """Canonicalize a list of values to a stable JSON string for hashing."""
    return json.dumps([_canon_scalar(v) for v in values], ensure_ascii=False, separators=(",", ":"))



def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


# ---------------------------
# Internal helpers
# ---------------------------

def _validate_headers(headers: List[str]) -> None:
    if not headers:
        raise ValueError("No non-empty headers found on the specified header row.")
    if "" in headers:
        raise ValueError("Empty header names are not allowed after normalization.")
    seen: Dict[str, int] = {}
    dups: List[str] = []
    for h in headers:
        seen[h] = seen.get(h, 0) + 1
        if seen[h] == 2:
            dups.append(h)
    if dups:
        raise ValueError(f"Duplicate header names found: {sorted(dups)}")


def _sheet_headers(ws, header_row_idx: int) -> Tuple[List[str], List[int]]:
    """Return normalized headers and kept positions using values_only for speed.

    Adds clear validation for header_row bounds (>=1 and <= max_row).
    """
    if header_row_idx < 1:
        raise ValueError(f"Header row must be >= 1 (got {header_row_idx}).")
    # openpyxl's max_row can be 1 even for empty sheets with a default row
    max_row = getattr(ws, "max_row", 0) or 0
    if header_row_idx > max_row:
        raise ValueError(
            f"Header row {header_row_idx} exceeds sheet max row {max_row}."
        )

    header_cells = next(
        ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
    )
    headers_norm = normalize_headers(header_cells)
    kept_positions = [i for i, h in enumerate(headers_norm) if h != ""]
    headers = [headers_norm[i] for i in kept_positions]
    _validate_headers(headers)
    return headers, kept_positions



def _process_rows(
    ws,
    *,
    header_row_idx: int,
    headers: List[str],
    kept_positions: List[int],
    pad_cols: Sequence[str],
    drop_blank_rows: bool,
    require_non_null: Sequence[str] | None,
    row_hash: bool,
    excel_row_numbers: bool,
    pad_hierarchical: bool,   # NEW: True = hierarchical (default), False = independent/legacy
) -> Iterable[List[Optional[str]]]:
    """
    Stream rows after the header and yield processed rows as lists matching:
      [row_hash?] + [excel_row?] + headers...

    Behavior:
      - Hierarchical fill-down (default): when a higher-level pad_col gets a new value,
        all lower-level pad_col carries are reset for that row.
      - Independent fill-down (legacy): each pad_col carries independently (no resets).
      - Completely empty rows (all headers blank) are treated as spacer rows:
          * dropped if --drop-blank-rows
          * otherwise emitted as empty (no fill-down applied), carry persists.
      - Dropping by --require-non-null happens AFTER padding.
      - Dropping by --drop-blank-rows across pad_cols happens AFTER padding.
      - Strings are stripped for storage; whitespace-only becomes None.
      - row_hash (if enabled) is computed over the stripped values (data columns only).
    """
    require_non_null = list(require_non_null or [])

    # Validate requested names exist
    for flagname, cols in (("pad_cols", pad_cols), ("require_non_null", require_non_null)):
        missing = [c for c in cols if c not in headers]
        if missing:
            raise ValueError(f"{flagname} header(s) not found: {missing}")

    # De-dup pad cols, keep order
    seen = set()
    pad_cols = [c for c in pad_cols if not (c in seen or seen.add(c))]

    # Carry only for padded columns
    carry = {h: None for h in pad_cols}
    current_excel_row = header_row_idx + 1  # first data row

    for row_vals in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        excel_row_num = current_excel_row
        current_excel_row += 1

        # Extract only the kept positions (match headers)
        values: List[Optional[str]] = [
            (row_vals[i] if i < len(row_vals) else None) for i in kept_positions
        ]
        row_map = {h: values[idx] for idx, h in enumerate(headers)}

        # Determine if this is a completely empty row across all headers
        is_completely_empty = all(_is_blank_cell(row_map.get(h)) for h in headers)

        # Spacer rows: drop early if requested; no padding and carry persists
        if is_completely_empty:
            if drop_blank_rows:
                continue
            stripped_vals = [_strip_for_storage(row_map.get(h)) for h in headers]
            row_out_vals = stripped_vals
            if excel_row_numbers:
                row_out_vals = [_strip_for_storage(str(excel_row_num))] + row_out_vals
            if row_hash:
                row_hash_value = sha256_hex(canon_list(stripped_vals))
                row_out_vals = [row_hash_value] + row_out_vals
            yield row_out_vals
            continue

        # -----------------------------
        # Fill-down (hierarchical or independent)
        # -----------------------------
        if pad_hierarchical:
            # Hierarchical: when a higher-level value appears, reset lower levels
            for i, h in enumerate(pad_cols):
                x = row_map.get(h)
                if _is_blank_cell(x):
                    row_map[h] = carry[h]
                else:
                    carry[h] = x
                    for lower_h in pad_cols[i + 1:]:
                        carry[lower_h] = None
        else:
            # Independent (legacy): each padded column carries independently
            for h in pad_cols:
                x = row_map.get(h)
                if _is_blank_cell(x):
                    row_map[h] = carry[h]
                else:
                    carry[h] = x

        # Enforce required_non_null *after* padding
        if require_non_null:
            if any(_is_blank_cell(row_map.get(h)) for h in require_non_null):
                continue

        # Drop rows that are still blank across pad_cols *after* padding
        if drop_blank_rows and pad_cols:
            if all(_is_blank_cell(row_map.get(h)) for h in pad_cols):
                continue

        # Build stripped output row in header order
        stripped_vals = [_strip_for_storage(row_map[h]) for h in headers]
        row_out_vals = stripped_vals

        # Optional excel_row
        if excel_row_numbers:
            row_out_vals = [_strip_for_storage(str(excel_row_num))] + row_out_vals

        # Optional row_hash (over data headers only)
        if row_hash:
            row_hash_value = sha256_hex(canon_list(stripped_vals))
            row_out_vals = [row_hash_value] + row_out_vals

        yield row_out_vals




# ---------------------------
# Public ingestion functions
# ---------------------------

def ingest_excel_to_sqlite(
    *,
    file: str | Path,
    sheet: str,
    header_row: int,
    pad_cols: Sequence[str],
    db: str | Path,
    table: Optional[str] = None,
    drop_blank_rows: bool = False,
    require_non_null: Sequence[str] | None = None,
    row_hash: bool = False,
    excel_row_numbers: bool = False,
    if_exists: str = "fail",  # "fail" | "replace" | "append"
    batch_size: int = 1000,
    pad_hierarchical: bool = True,   # NEW: default hierarchical; False -> legacy independent
) -> Dict[str, Any]:
    """
    Stream an Excel sheet into SQLite with fill-down on selected columns by name,
    preserving original Excel row numbers (1-based), and creating the table on the fly.
    """
    if int(header_row) < 1:
        raise ValueError(f"Header row must be >= 1 (got {header_row}).")

    xlsx_path = Path(file)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Input not found: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet} (available: {wb.sheetnames})")
        ws = wb[sheet]

        header_row_idx = int(header_row)  # no clamping; let _sheet_headers validate bounds
        headers, kept_positions = _sheet_headers(ws, header_row_idx)

        table_name = table or sheet

        with sqlite3.connect(str(db)) as conn:
            cur = conn.cursor()

            if if_exists == "replace":
                cur.execute(f"DROP TABLE IF EXISTS {qident(table_name)};")
            elif if_exists == "fail":
                r = cur.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table_name,)
                ).fetchone()
                if r:
                    raise ValueError(
                        f"Table already exists: {table_name} (use if_exists='replace' or 'append')"
                    )

            # Create table: [row_hash TEXT if enabled] + excel_row TEXT + all headers TEXT
            cols_sql: List[str] = []
            if row_hash:
                cols_sql.append(f'{qident("row_hash")} TEXT')
            if excel_row_numbers:
                cols_sql.append(f'{qident("excel_row")} TEXT')
            cols_sql.extend(f"{qident(h)} TEXT" for h in headers)
            cur.execute(
                f"CREATE TABLE IF NOT EXISTS {qident(table_name)} ({', '.join(cols_sql)});"
            )

            # Helpful indexes
            if excel_row_numbers:
                cur.execute(
                    f"CREATE INDEX IF NOT EXISTS {qident(table_name + '_row_idx')} "
                    f"ON {qident(table_name)} ({qident('excel_row')});"
                )
            if row_hash:
                cur.execute(
                    f"CREATE INDEX IF NOT EXISTS {qident(table_name + '_rowhash_idx')} "
                    f"ON {qident(table_name)} ({qident('row_hash')});"
                )
            conn.commit()

            # INSERT statement
            insert_cols = headers.copy()
            if excel_row_numbers:
                insert_cols = ["excel_row"] + insert_cols
            if row_hash:
                insert_cols = ["row_hash"] + insert_cols
            placeholders = ",".join(["?"] * len(insert_cols))
            insert_sql = (
                f"INSERT INTO {qident(table_name)} ("
                + ", ".join(qident(c) for c in insert_cols)
                + f") VALUES ({placeholders});"
            )

            # If appending, validate schema matches exactly (column order too)
            if if_exists == "append":
                r = cur.execute(f"PRAGMA table_info({qident(table_name)});").fetchall()
                if r:
                    existing_cols = [row[1] for row in r]  # name column
                    expected = []
                    if row_hash:
                        expected.append("row_hash")
                    if excel_row_numbers:
                        expected.append("excel_row")
                    expected += headers
                    if existing_cols != expected:
                        raise ValueError(
                            f"Existing schema {existing_cols} does not match expected {expected} for append."
                        )

            # Stream/process rows
            n_rows = 0
            batch: List[List[Optional[str]]] = []
            batch_size = max(1, int(batch_size))

            for out_row in _process_rows(
                ws,
                header_row_idx=header_row_idx,
                headers=headers,
                kept_positions=kept_positions,
                pad_cols=pad_cols,
                drop_blank_rows=drop_blank_rows,
                require_non_null=require_non_null,
                row_hash=row_hash,
                excel_row_numbers=excel_row_numbers,
                pad_hierarchical=pad_hierarchical,  # pass through
            ):
                batch.append(out_row)
                n_rows += 1
                if len(batch) >= batch_size:
                    cur.executemany(insert_sql, batch)
                    batch.clear()
            if batch:
                cur.executemany(insert_sql, batch)

            conn.commit()

        return {
            "table": table_name,
            "columns": headers,
            "rows_ingested": n_rows,
            "row_hash": bool(row_hash),
            "excel_row_numbers": bool(excel_row_numbers),
        }
    finally:
        # Explicitly close the source workbook (important on Windows/long runs)
        try:
            wb.close()
        except Exception:
            pass




def ingest_excel_to_excel(
    *,
    file: str | Path,
    sheet: str,
    header_row: int,
    pad_cols: Sequence[str],
    outfile: str | Path,
    outsheet: Optional[str] = None,
    drop_blank_rows: bool = False,
    require_non_null: Sequence[str] | None = None,
    row_hash: bool = False,
    excel_row_numbers: bool = False,
    if_exists: str = "fail",  # "fail" | "replace" | "append"
    pad_hierarchical: bool = True,     # NEW: default hierarchical; False -> legacy independent
) -> Dict[str, Any]:
    """
    Stream an Excel sheet into another Excel sheet, padding (fill-down) selected columns by name,
    preserving original Excel row numbers (1-based), and writing headers in the destination.
    """
    if int(header_row) < 1:
        raise ValueError(f"Header row must be >= 1 (got {header_row}).")

    src_path = Path(file)
    if not src_path.exists():
        raise FileNotFoundError(f"Input not found: {src_path}")

    wb_src = load_workbook(filename=str(src_path), read_only=True, data_only=True)
    wb_dst = None  # ensure defined for finally
    try:
        if sheet not in wb_src.sheetnames:
            raise ValueError(f"Sheet not found: {sheet} (available: {wb_src.sheetnames})")
        ws_src = wb_src[sheet]

        header_row_idx = int(header_row)  # no clamping; let _sheet_headers validate bounds
        headers, kept_positions = _sheet_headers(ws_src, header_row_idx)

        dest_path = Path(outfile)
        sheet_name = outsheet or sheet

        # Decide if we can stream-write (write_only): only when creating a NEW workbook
        can_stream_write = (not dest_path.exists())

        # Open or create destination workbook
        if dest_path.exists():
            wb_dst = load_workbook(filename=str(dest_path))
        else:
            wb_dst = Workbook(write_only=True)
            # Remove the auto-created active sheet unless it already matches our target
            try:
                default_ws = getattr(wb_dst, "active", None)
                if default_ws is not None and default_ws.title != sheet_name:
                    wb_dst.remove(default_ws)
            except Exception:
                # If anything odd happens, it's safe to proceed; we'll create our own sheet next.
                pass


        # Handle if_exists at sheet level
        if sheet_name in wb_dst.sheetnames and not getattr(wb_dst, "write_only", False):
            if if_exists == "fail":
                raise ValueError(f"Sheet already exists: {sheet_name} in {dest_path}")
            elif if_exists == "replace":
                ws_old = wb_dst[sheet_name]
                wb_dst.remove(ws_old)
                ws_dst = wb_dst.create_sheet(title=sheet_name)
                write_headers = True
                start_row = 2
            elif if_exists == "append":
                ws_dst = wb_dst[sheet_name]
                # Validate header row matches expected
                expected_cols = []
                if row_hash:
                    expected_cols.append("row_hash")
                if excel_row_numbers:
                    expected_cols.append("excel_row")
                expected_cols += headers

                if ws_dst.max_row < 1:
                    raise ValueError(f"Destination sheet '{sheet_name}' is empty; cannot append without header.")
                dest_header_row = [c.value for c in next(ws_dst.iter_rows(min_row=1, max_row=1))]
                dest_header_norm = normalize_headers(dest_header_row)
                dest_kept_positions = [i for i, h in enumerate(dest_header_norm) if h != ""]
                dest_headers = [dest_header_norm[i] for i in dest_kept_positions]
                if dest_headers != expected_cols:
                    raise ValueError(
                        f"Destination header {dest_headers} does not match expected {expected_cols} for append."
                    )
                write_headers = False
                start_row = ws_dst.max_row + 1
            else:
                raise ValueError(f"Unknown if_exists mode: {if_exists}")
        else:
            ws_dst = wb_dst.create_sheet(title=sheet_name)
            write_headers = True
            start_row = 2

        # Build the output header order (destination columns)
        out_header = []
        if row_hash:
            out_header.append("row_hash")
        if excel_row_numbers:
            out_header.append("excel_row")
        out_header.extend(headers)

        # Write header if needed
        if write_headers:
            if getattr(wb_dst, "write_only", False):
                ws_dst.append(out_header)
            else:
                for col_idx, name in enumerate(out_header, start=1):
                    ws_dst.cell(row=1, column=col_idx, value=name)

        # Write rows
        n_rows = 0
        if getattr(wb_dst, "write_only", False):
            for out_row in _process_rows(
                ws_src,
                header_row_idx=header_row_idx,
                headers=headers,
                kept_positions=kept_positions,
                pad_cols=pad_cols,
                drop_blank_rows=drop_blank_rows,
                require_non_null=require_non_null,
                row_hash=row_hash,
                excel_row_numbers=excel_row_numbers,
                pad_hierarchical=pad_hierarchical,  # pass through
            ):
                ws_dst.append(out_row)
                n_rows += 1
        else:
            row_index = start_row
            for out_row in _process_rows(
                ws_src,
                header_row_idx=header_row_idx,
                headers=headers,
                kept_positions=kept_positions,
                pad_cols=pad_cols,
                drop_blank_rows=drop_blank_rows,
                require_non_null=require_non_null,
                row_hash=row_hash,
                excel_row_numbers=excel_row_numbers,
                pad_hierarchical=pad_hierarchical,  # pass through
            ):
                for col_idx, val in enumerate(out_row, start=1):
                    ws_dst.cell(row=row_index, column=col_idx, value=val)
                row_index += 1
                n_rows += 1

        wb_dst.save(str(dest_path))

        return {
            "workbook": str(dest_path),
            "sheet": sheet_name,
            "columns": headers,
            "rows_written": n_rows,
            "row_hash": bool(row_hash),
            "excel_row_numbers": bool(excel_row_numbers),
        }
    finally:
        # Close both workbooks explicitly (protect against file-handle leaks)
        try:
            wb_src.close()
        except Exception:
            pass
        try:
            if wb_dst is not None:
                wb_dst.close()
        except Exception:
            pass



