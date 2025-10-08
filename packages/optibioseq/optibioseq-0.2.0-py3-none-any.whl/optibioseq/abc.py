from bs4 import BeautifulSoup
import re 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl import load_workbook

def input_seq ():
    secuencia = open(wdir+"/input/input-seq.txt", "r")
    Secuencia = open(wdir+"/input/input-seq.txt", "r")
    red_seq = Secuencia.read()
    Secuencia.close

    with open(wdir+"/input/input-seq.txt") as file:
        lineas = [line.rstrip() for line in file]
    lista = []
    for i in lineas:
        if i.startswith(">"):
            lista.append("aquies")
        else:
            lista.append(i)
            
    if lista[0] == "aquies":
        lista.remove("aquies")
    seq = "".join(lista)
    seq = seq.upper()
    input_seq.seq = seq
    
    no_amino_acidos = ['J','O','U','X']
    for i in no_amino_acidos:
        isit = seq.find(i)
        if isit != -1:
            print ("\nEsta no es una secuencia válida de amino acidoácidos\n")
            print ("\nCaracter no válido en la posición ", isit,".\n") 
            break


def abcpredresults(seq):
    abc_url = wdir+"/abcpred/abcpred_res.html"
    abc_page = open(abc_url)
    
    soup = BeautifulSoup(abc_page.read(), "html.parser")
    
    #Paso 1: Dejar únicamente el ovrlap display (la parte que sí sirve.
    #cambiar los <br> por retornos de linea.
    phase1 = stringTable_02 = re.sub(r'<.+?>','',str(soup).replace('<br/>','\n'))
    #Paso 2: delimitar hasta cual línea se debe cortar
    beginning = phase1.index ("DISPLAY")
    phase2 = phase1[beginning + 7: ]
    phase2 = re.sub(r'([A-Z])','Z',str(phase2))
    #Paso 3: Eliminar la secuencia, dejando solo los resultados
    phase3 = phase2.split("\n")
    del phase3[0]
    phase3.reverse()
    results_list = ' '.join(phase3).split()
    abcpredresults.results_list = results_list
    
def results2xls(seq):
    seq = input_seq.seq
    the_matrix = abcpredresults.results_list
    wb = load_workbook(wdir+'/ICE-BIOSEQ-out.xlsx') 
    ws = wb.active
    
    ws["A13"] = "ABCPred"
    ws["A13"].font = Font(bold=True, color = "FFFFFF")
    ws["A13"].fill = PatternFill(fgColor= "7030A0", fill_type = "solid")
    
    #Esribir la secuencia
    for x, y in enumerate(seq):
        ws.cell(row= 13, column= x+2).value = y
        
    #Colorear el primer resultados
    hex_base = int(6303950)
    for n in the_matrix:
        for i, h in enumerate (seq):
            dummy = n
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                if hex_base < int(0x10000):
                    color = hex(7352480)
                else:
                    color = hex(hex_base)
                painting = color[2:]
                ws.cell(row=13, column=2+i).fill = PatternFill(fgColor= painting, fill_type = "solid")
                ws.cell(row=13, column=2+i).font = Font(color = "FFFFFF")
        hex_base = hex_base - int(1572888)

    wb.save(wdir+"/ICE-BIOSEQ-out.xlsx")

def start(wodir):
    global wdir
    wdir = wodir
    input_seq()
    abcpredresults(input_seq.seq)
    results2xls(input_seq.seq)