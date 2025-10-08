from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl import load_workbook



def imput_seq():
    Secuencia = open(wdir+"/input/input-seq.txt", "r")
    red_seq = Secuencia.read()
    Secuencia.close

    with open(wdir+"/input/input-seq.txt") as file:
        lineas = [line.rstrip() for line in file]
    lista = []
    for i in lineas:
        if i.startswith(">"):
            lista.append("aquies")
        else:
            lista.append(i)
            
    if lista[0] == "aquies":
        lista.remove("aquies")
    seq = "".join(lista)
    seq = seq.upper()
    imput_seq.seq = seq
    
    no_amino_acidos = ['J','O','U','X']
    for i in no_amino_acidos:
        isit = seq.find(i)
        if isit != -1:
            print ("\nEsta no es una secuencia válida de amino acidoácidos\n")
            print ("\nCaracter no válido en la posición ", isit,".\n") 
            break

def iedb1_results(seq):
    iedb_url = wdir+"/iedb1/iedb-mhci.html"
    iedb_page = open(iedb_url)
        
    soup = BeautifulSoup(iedb_page.read(), "html.parser")
    #Dejar únicamente la tabla de los resultados.
    #La salida de BS4 debe ser convertida a string antes de trabajarla como tal
    sopa = str(soup)
    #Cambir tabuladores por espacios
    sopa2 = re.sub(r'\t',' ',str(sopa))
    #Hacer una lista con cada "renglón"
    sopa_table = sopa2.split("\n")
    #Lista donde cada renglón se divide como lista
    sopa_table2 = []
    for i in sopa_table:
        sopa_table2.append(i.split())
    #Crear una lista únicamente con los resultados.
    sopa_table3 = []
    for i in sopa_table2:
        if len(i) == 8:
            sopa_table3.append(i)
    del sopa_table3[0]
    del sopa_table3[0]
    #Hacer una lista con los datos relevantes para el análisis antigénicos
    chosen_ones = []
    for i in sopa_table3:
        if float(i[6]) >= 0.85:
            chosen_ones.append(i)
        else:
            break
    iedb1_results.results_list = chosen_ones
     
def results2xls(seq):
    seq = imput_seq.seq
    the_matrix = iedb1_results.results_list
    wb = load_workbook(wdir+'/ICE-BIOSEQ-out.xlsx') 
    ws = wb.active
    
    ws["A17"] = "IEDB MHCI"
    ws["A17"].font = Font(bold=True)
    ws["A17"].fill = PatternFill(fgColor= "5B9BD5", fill_type = "solid")
    
    #Esribir la secuencia
    for x, y in enumerate(seq):
        ws.cell(row= 17, column= x+2).value = y
    
    for i in the_matrix:
        start = int(i[2])
        end = int(i[3])
        while start <= end:
            ws.cell(row=17, column=1+start).fill = PatternFill(fgColor= "5B9BD5", fill_type = "solid")
            start = start+1
        
    wb.save(wdir+"/ICE-BIOSEQ-out.xlsx")
    
def start(wodir):
    global wdir
    wdir = wodir
    imput_seq()
    iedb1_results(imput_seq.seq)
    results2xls(imput_seq.seq)