from bs4 import BeautifulSoup
import re
import pandas as pd
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl import load_workbook

def input_seq ():
    Secuencia = open(wdir+"/input/input-seq.txt", "r")
    red_seq = Secuencia.read()
    Secuencia.close

    with open(wdir+"/input/input-seq.txt") as file:
        lineas = [line.rstrip() for line in file]
    lista = []
    for i in lineas:
        if i.startswith(">"):
            lista.append("aquies")
        else:
            lista.append(i)
            
    if lista[0] == "aquies":
        lista.remove("aquies")
    seq = "".join(lista)
    seq = seq.upper()
    input_seq.seq = seq
    
    no_amino_acidos = ['J','O','U','X']
    for i in no_amino_acidos:
        isit = seq.find(i)
        if isit != -1:
            print ("\nEsta no es una secuencia válida de amino acidoácidos\n")
            print ("\nCaracter no válido en la posición ", isit,".\n") 
            break

def BEPI_results(seq):
    zf = ZipFile(wdir+"/bepipred/bepi.zip", "r")
    zf.extractall(wdir+"/bepipred")
    zf.close()
    data = pd.read_csv(wdir+"/bepipred/raw_output.csv")
    scores = list(data["BepiPred-3.0 linear epitope score"])
    BEPI_results.results_list = scores
    
def results2xls(seq):
    seq = input_seq.seq
    the_matrix = BEPI_results.results_list
    wb = load_workbook(wdir+'/ICE-BIOSEQ-out.xlsx') 
    ws = wb.active
    
    ws["A15"] = "BEPIPred"
    ws["A15"].font = Font(bold=True)
    ws["A15"].fill = PatternFill(fgColor= "00FF00", fill_type = "solid")
    
    #Esribir la secuencia
    for x, y in enumerate(seq):
        ws.cell(row= 15, column= x+2).value = y
    
    #Colorear los resultados
    for i, h in enumerate (seq):
        dummy = the_matrix[i]
        poc = float(0.1512)
        if dummy >= poc:
            ws.cell(row=15, column=2+i).fill = PatternFill(fgColor= "00FF00", fill_type = "solid")
    
    wb.save(wdir+"/ICE-BIOSEQ-out.xlsx")

def start(wodir):
    global wdir
    wdir = wodir
    input_seq()
    BEPI_results(input_seq.seq)
    results2xls(input_seq.seq)