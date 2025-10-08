from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl import load_workbook

Thresholds = [2,1.9,2,1.9,2.4,2.3,1.8,1.9]
def imput_seq():
    Secuencia = open(wdir+"/input/input-seq.txt", "r")
    red_seq = Secuencia.read()
    Secuencia.close

    with open(wdir+"/input/input-seq.txt") as file:
        lineas = [line.rstrip() for line in file]
    lista = []
    for i in lineas:
        if i.startswith(">"):
            lista.append("aquies")
        else:
            lista.append(i)
            
    if lista[0] == "aquies":
        lista.remove("aquies")
    seq = "".join(lista)
    seq = seq.upper()
    imput_seq.seq = seq
    
    no_amino_acidos = ['J','O','U','X']
    for i in no_amino_acidos:
        isit = seq.find(i)
        if isit != -1:
            print ("\nEsta no es una secuencia válida de amino acidoácidos\n")
            print ("\nCaracter no válido en la posición ", isit,".\n") 
            break

def bcepredresults(seq,Thresholds):
    bce_file = wdir+"/bcepred/bcepred.html"
    page = open(bce_file)
    
    soup = BeautifulSoup(page.read(), "html.parser")
    
    
    #Paso 1: Dejar únicamente el ovrlap display (la parte que sí sirve.
    #cambiar los <br> por retornos de linea.
    phase0 = re.sub(r'<u>([A-Z])</u>','Z',str(soup))
    phase1 = re.sub(r'<.+?>','',str(phase0).replace('<tr>','\n'))
    phase2 = re.sub(r'<.+?>','',str(phase1).replace('<br/>','\n'))
    
    
    #Paso 2: delimitar hasta cual línea se debe cortar
    beginning = phase2.index ("The predicted")
    phase3 = phase2[beginning + 71: ]
    phase4 = re.sub(r'<u>([A-Z])</u>','Z',str(phase3))
    phase5 = re.sub(r'<.+?>','',str(phase4).replace('<tr>','\n'))
    phase6 = re.sub(r'1','\n',str(phase5))
    
    antigen_results = re.sub(r'([0-9])','',str(phase6))
    
    results_list = antigen_results.split("\n")
    
    while("" in results_list):
        results_list.remove("") 
    
    #Tomando en cuenta que este programa da un resultado mas complejo (múltiples resultados, resultados 
    #Asegurar que todos los resultados estén en orden
    
        
    while len(results_list) <= 16:
        results_list.append ("N/A")
    
    
    if results_list[2].find("Hydrophilicity") == -1:
        results_list.insert(2, "Hydrophilicity")
        if results_list [3] != "N/A":
            results_list.insert(3, "N/A")
    
    if results_list[4].find("Flexibility") == -1:
        results_list.insert(4, "Flexibility")
        if results_list [5] != "N/A":
            results_list.insert(5, "N/A")
    
    if results_list[6].find("Accessibility") == -1:
        results_list.insert(6, "Accessibility")
        if results_list [7] != "N/A":
            results_list.insert(7, "N/A")
    
    if results_list[8].find("Turns") == -1:
        results_list.insert(8, "Turns")
        if results_list [9] != "N/A":
            results_list.insert(9, "N/A")
    
    if results_list[10].find("Exposed") == -1:
        results_list.insert(10, "Exposed Surface")
        if results_list [11] != "N/A":
            results_list.insert(11, "N/A")
    
    if results_list[12].find("Polarity") == -1:
        results_list.insert(12, "Polarity")
        if results_list [13] != "N/A":
            results_list.insert(13, "N/A")
    
    if results_list[14].find("Antigenic") == -1:
        results_list.insert(14, "Antigenic Propensity")
        if results_list [15] != "N/A":
            results_list.insert(15, "N/A")
    
    while len(results_list) > 16:
        del results_list[16]
    bcepredresults.results_list = results_list
  
    
    
def results2xls(seq):
    seq = imput_seq.seq
    the_matrix = bcepredresults.results_list
    wb = load_workbook(wdir+'/ICE-BIOSEQ-out.xlsx') 
    ws = wb.active
    ws['A2'] = "Secuencia"
    ws['A2'].font = Font(bold=True)
    
    ws["A4"] = "BCEPred"
    ws['A4'].font = Font(bold=True)
    
    tasks = ["Hydrophilicity", "Flexibility", "Accesibility", "Turns", "Exposed surface", "Polarity", "Antigenic propencity"]
    colors = ["00b0f0", "92d050", "ff0000", "ffff00", "f4b084", "ffd966", "c9c9c9"]
    
    coor = 5
    for t in (tasks):
        ws.cell(row=coor, column=1).value = t
        coor = coor + 1
    
    coor = 5
    for c in colors:
        ws.cell(row=coor, column=1).fill = PatternFill(fgColor= c, fill_type = "solid")
        coor = coor+1
    
    #Escribir la secuencia base
    for x, y in enumerate(seq):
        ws.cell(row= 2, column= x+2).value = y
        
    #Escribir la secuencia en cada una de las tasks    
    for i, h in enumerate(tasks):
        le_row = i + 5
        for x, y in enumerate(seq):
            ws.cell(row= le_row, column= x+2).value = y
    
    #colorear los resultados de hydrophilicity     
    if the_matrix[3] != "N/A":
        for i, h in enumerate (seq):
            dummy = (the_matrix[3])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=5, column=2+i).fill = PatternFill(fgColor= colors[0], fill_type = "solid")
    
    #colorear los resultados de Flexibility   
    if the_matrix[5] != "N/A":
        for i, h in enumerate (seq):
            dummy = (the_matrix[5])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=6, column=2+i).fill = PatternFill(fgColor= colors[1], fill_type = "solid")
            
    #colorear los resultados de Accesibility
    if the_matrix[7] != "N/A":
        for i, h in enumerate (seq):
            dummy = (the_matrix[7])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=7, column=2+i).fill = PatternFill(fgColor= colors[2], fill_type = "solid")

    #colorear los resultados de Turns
    if the_matrix[9] != "N/A":
        for i, h in enumerate (seq):
            dummy = (the_matrix[9])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=8, column=2+i).fill = PatternFill(fgColor= colors[3], fill_type = "solid")
            
    #colorear los resultados de Exposed surface
    if the_matrix[11] != "N/A":       
        for i, h in enumerate (seq):
            dummy = (the_matrix[11])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=9, column=2+i).fill = PatternFill(fgColor= colors[4], fill_type = "solid")
    
    #colorear los resultados de Polarity
    if the_matrix[13] != "N/A":        
        for i, h in enumerate (seq):
            dummy = (the_matrix[13])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=10, column=2+i).fill = PatternFill(fgColor= colors[5], fill_type = "solid")    
    
    #colorear los resultados de Antigenic propencity
    if the_matrix[15] != "N/A":        
        for i, h in enumerate (seq):
            dummy = (the_matrix[15])
            dummer = dummy[i]
            poc = "Z"
            if poc == dummer:
                ws.cell(row=11, column=2+i).fill = PatternFill(fgColor= colors[6], fill_type = "solid")  
    wb.save(wdir+'/ICE-BIOSEQ-out.xlsx')

def start(wodir):
    global wdir
    wdir = wodir
    imput_seq()
    bcepredresults(imput_seq.seq, Thresholds)
    results2xls(imput_seq.seq)