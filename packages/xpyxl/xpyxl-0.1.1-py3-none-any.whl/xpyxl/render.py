from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from typing import Any, assert_never

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .nodes import (
    CellNode,
    ColumnNode,
    HorizontalStackNode,
    RenderableItem,
    RowNode,
    SheetComponent,
    SheetNode,
    SpacerNode,
    TableNode,
    VerticalStackNode,
)
from .styles import (
    BorderStyleName,
    Style,
    bold,
    combine_styles,
    normalize_hex,
    to_argb,
)

__all__ = ["render_sheet"]


DEFAULT_FONT_NAME = "Calibri"
DEFAULT_FONT_SIZE = 11.0
DEFAULT_MONO_FONT = "Consolas"
DEFAULT_TEXT_COLOR = normalize_hex("#111827")
DEFAULT_BORDER_COLOR = normalize_hex("#D0D5DD")
DEFAULT_BORDER_STYLE: BorderStyleName = "thin"
DEFAULT_ROW_HEIGHT = 16.0
DEFAULT_TABLE_HEADER_BG = normalize_hex("#F2F4F7")
DEFAULT_TABLE_HEADER_TEXT = normalize_hex("#1E293B")
DEFAULT_TABLE_STRIPE_COLOR = normalize_hex("#F8FAFC")
DEFAULT_TABLE_COMPACT_HEIGHT = 18.0


@dataclass
class EffectiveStyle:
    font_name: str
    font_size: float
    bold: bool
    italic: bool
    text_color: str
    fill_color: str | None
    horizontal_align: str | None
    vertical_align: str | None
    indent: int | None
    wrap_text: bool
    number_format: str | None
    border: BorderStyleName | None
    border_color: str | None


@dataclass(frozen=True)
class _Size:
    width: int
    height: int


@dataclass(frozen=True)
class _Placement:
    row: int
    col: int
    item: RenderableItem


def _resolve(styles: Sequence[Style]) -> EffectiveStyle:
    base_style = Style(
        font_name=DEFAULT_FONT_NAME,
        font_size=DEFAULT_FONT_SIZE,
        text_color=DEFAULT_TEXT_COLOR,
    )
    merged = combine_styles(styles, base=base_style)

    font_name = merged.font_name or DEFAULT_FONT_NAME
    if merged.mono:
        font_name = DEFAULT_MONO_FONT
    font_size = merged.font_size if merged.font_size is not None else DEFAULT_FONT_SIZE
    if merged.font_size_delta is not None:
        font_size += merged.font_size_delta

    bold_flag = merged.bold if merged.bold is not None else False
    italic_flag = merged.italic if merged.italic is not None else False

    text_color = normalize_hex(merged.text_color or DEFAULT_TEXT_COLOR)
    fill_color = normalize_hex(merged.fill_color) if merged.fill_color else None
    border_color = normalize_hex(merged.border_color) if merged.border_color else None

    return EffectiveStyle(
        font_name=font_name,
        font_size=font_size,
        bold=bold_flag,
        italic=italic_flag,
        text_color=text_color,
        fill_color=fill_color,
        horizontal_align=merged.horizontal_align,
        vertical_align=merged.vertical_align,
        indent=merged.indent,
        wrap_text=merged.wrap_text if merged.wrap_text is not None else False,
        number_format=merged.number_format,
        border=merged.border,
        border_color=border_color,
    )


def _default_row_height() -> float:
    return DEFAULT_ROW_HEIGHT


def _update_dimensions(
    *,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
    column_index: int,
    row_index: int,
    value: Any,
    style: EffectiveStyle,
    prefer_height: float | None = None,
) -> None:
    text = "" if value is None else str(value)
    width_hint = max(len(text), 1.0)
    existing_width = col_widths.get(column_index, 0.0)
    col_widths[column_index] = max(existing_width, width_hint)

    base_height = prefer_height if prefer_height is not None else _default_row_height()
    row_heights[row_index] = max(row_heights.get(row_index, 0.0), base_height)


def _apply_style(cell, effective: EffectiveStyle, border_fallback_color: str) -> None:
    cell.font = Font(
        name=effective.font_name,
        size=effective.font_size,
        bold=effective.bold,
        italic=effective.italic,
        color=to_argb(effective.text_color),
    )

    if effective.fill_color:
        color = to_argb(effective.fill_color)
        cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)

    align_kwargs: dict[str, Any] = {}
    if effective.horizontal_align:
        align_kwargs["horizontal"] = effective.horizontal_align
    if effective.vertical_align:
        align_kwargs["vertical"] = effective.vertical_align
    if effective.indent is not None:
        align_kwargs["indent"] = effective.indent
    if effective.wrap_text:
        align_kwargs["wrap_text"] = True
    if align_kwargs:
        align_kwargs.setdefault("vertical", "bottom")
        cell.alignment = Alignment(**align_kwargs)
    elif cell.alignment is None and effective.wrap_text:
        cell.alignment = Alignment(wrap_text=True)

    if effective.number_format:
        cell.number_format = effective.number_format

    if effective.border:
        border_color = effective.border_color or border_fallback_color
        side = Side(style=effective.border, color=to_argb(border_color))
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def _render_row(
    ws,
    node: RowNode,
    start_row: int,
    start_col: int,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> None:
    row_index = start_row
    for column_offset, cell_node in enumerate(node.cells, start=1):
        styles = (*node.styles, *cell_node.styles)
        effective = _resolve(styles)
        column_index = start_col + column_offset - 1
        target_cell = ws.cell(row=row_index, column=column_index, value=cell_node.value)
        _apply_style(target_cell, effective, DEFAULT_BORDER_COLOR)
        _update_dimensions(
            col_widths=col_widths,
            row_heights=row_heights,
            column_index=column_index,
            row_index=row_index,
            value=cell_node.value,
            style=effective,
        )


def _render_column(
    ws,
    node: ColumnNode,
    start_row: int,
    start_col: int,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> None:
    row_index = start_row
    for cell_node in node.cells:
        styles = (*node.styles, *cell_node.styles)
        effective = _resolve(styles)
        target_cell = ws.cell(row=row_index, column=start_col, value=cell_node.value)
        _apply_style(target_cell, effective, DEFAULT_BORDER_COLOR)
        _update_dimensions(
            col_widths=col_widths,
            row_heights=row_heights,
            column_index=start_col,
            row_index=row_index,
            value=cell_node.value,
            style=effective,
        )
        row_index += 1


def _render_cell(
    ws,
    node: CellNode,
    row_index: int,
    column_index: int,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> None:
    effective = _resolve(node.styles)
    target_cell = ws.cell(row=row_index, column=column_index, value=node.value)
    _apply_style(target_cell, effective, DEFAULT_BORDER_COLOR)
    _update_dimensions(
        col_widths=col_widths,
        row_heights=row_heights,
        column_index=column_index,
        row_index=row_index,
        value=node.value,
        style=effective,
    )


def _render_table(
    ws,
    node: TableNode,
    start_row: int,
    start_col: int,
    col_widths: dict[int, float],
    row_heights: dict[int, float],
) -> None:
    table_style = combine_styles(node.styles)
    banded = table_style.table_banded if table_style.table_banded is not None else True
    bordered = (
        table_style.table_bordered if table_style.table_bordered is not None else True
    )
    compact = (
        table_style.table_compact if table_style.table_compact is not None else False
    )
    border_color = (
        table_style.border_color
        if table_style.border_color is not None
        else DEFAULT_BORDER_COLOR
    )
    border_style = (
        table_style.border if table_style.border is not None else DEFAULT_BORDER_STYLE
    )

    table_border_style = (
        Style(border=border_style, border_color=border_color) if bordered else None
    )
    stripe_style = Style(fill_color=DEFAULT_TABLE_STRIPE_COLOR) if banded else None
    compact_height = DEFAULT_TABLE_COMPACT_HEIGHT if compact else None

    current_row = start_row

    def render(
        row_node: RowNode,
        *,
        extra: Sequence[Style] = (),
        prefer_height: float | None = None,
    ) -> None:
        for column_offset, cell_node in enumerate(row_node.cells, start=1):
            style_chain = (*node.styles, *row_node.styles, *extra, *cell_node.styles)
            if table_border_style:
                style_chain = (*style_chain, table_border_style)
            effective = _resolve(style_chain)
            column_index = start_col + column_offset - 1
            target_cell = ws.cell(
                row=current_row, column=column_index, value=cell_node.value
            )
            _apply_style(target_cell, effective, border_color)
            _update_dimensions(
                col_widths=col_widths,
                row_heights=row_heights,
                column_index=column_index,
                row_index=current_row,
                value=cell_node.value,
                style=effective,
                prefer_height=prefer_height,
            )

    if node.header:
        header_extras: list[Style] = [bold]
        header_extras.append(Style(fill_color=DEFAULT_TABLE_HEADER_BG))
        header_extras.append(Style(text_color=DEFAULT_TABLE_HEADER_TEXT))
        render(node.header, extra=header_extras, prefer_height=compact_height)
        current_row += 1

    for idx, row_node in enumerate(node.rows):
        extras: list[Style] = []
        if stripe_style and idx % 2 == 1:
            extras.append(stripe_style)
        render(row_node, extra=extras, prefer_height=compact_height)
        current_row += 1


def _table_size(node: TableNode) -> _Size:
    width = 0
    height = 0
    if node.header:
        width = max(width, len(node.header.cells))
        height += 1
    for row in node.rows:
        width = max(width, len(row.cells))
        height += 1
    return _Size(width=width, height=height)


def _layout_item(
    item: SheetComponent, start_row: int, start_col: int
) -> tuple[list[_Placement], _Size]:
    if isinstance(item, CellNode):
        size = _Size(width=1, height=1)
        return ([_Placement(row=start_row, col=start_col, item=item)], size)
    elif isinstance(item, RowNode):
        size = _Size(width=len(item.cells), height=1)
        return ([_Placement(row=start_row, col=start_col, item=item)], size)
    elif isinstance(item, ColumnNode):
        size = _Size(width=1, height=len(item.cells))
        return ([_Placement(row=start_row, col=start_col, item=item)], size)
    elif isinstance(item, TableNode):
        size = _table_size(item)
        return ([_Placement(row=start_row, col=start_col, item=item)], size)
    elif isinstance(item, SpacerNode):
        size = _Size(width=0, height=item.rows)
        return ([_Placement(row=start_row, col=start_col, item=item)], size)
    elif isinstance(item, VerticalStackNode):
        placements: list[_Placement] = []  # pyright: ignore[reportRedeclaration]
        row_cursor = start_row
        max_width = 0
        for idx, child in enumerate(item.items):
            child_placements, child_size = _layout_item(child, row_cursor, start_col)
            placements.extend(child_placements)
            row_cursor += child_size.height
            if idx < len(item.items) - 1:
                row_cursor += item.gap
            max_width = max(max_width, child_size.width)
        height = row_cursor - start_row
        return placements, _Size(width=max_width, height=height)
    elif isinstance(item, HorizontalStackNode):
        placements: list[_Placement] = []
        col_cursor = start_col
        max_height = 0
        for idx, child in enumerate(item.items):
            child_placements, child_size = _layout_item(child, start_row, col_cursor)
            placements.extend(child_placements)
            col_cursor += child_size.width
            if idx < len(item.items) - 1:
                col_cursor += item.gap
            max_height = max(max_height, child_size.height)
        width = col_cursor - start_col
        return placements, _Size(width=width, height=max_height)
    else:
        assert_never(item)


def _apply_dimensions(
    ws, col_widths: Mapping[int, float], row_heights: Mapping[int, float]
) -> None:
    for column_index, width in col_widths.items():
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = max(width, 8.0)
    for row_index, height in row_heights.items():
        ws.row_dimensions[row_index].height = height


def render_sheet(ws, node: SheetNode) -> None:
    col_widths: dict[int, float] = {}
    row_heights: dict[int, float] = {}
    placements: list[_Placement] = []
    row_cursor = 1

    for item in node.items:
        item_placements, size = _layout_item(item, row_cursor, 1)
        placements.extend(item_placements)
        row_cursor += size.height

    for placement in placements:
        target = placement.item
        if isinstance(target, CellNode):
            _render_cell(
                ws, target, placement.row, placement.col, col_widths, row_heights
            )
        elif isinstance(target, RowNode):
            _render_row(
                ws, target, placement.row, placement.col, col_widths, row_heights
            )
        elif isinstance(target, ColumnNode):
            _render_column(
                ws, target, placement.row, placement.col, col_widths, row_heights
            )
        elif isinstance(target, TableNode):
            _render_table(
                ws, target, placement.row, placement.col, col_widths, row_heights
            )
        elif isinstance(target, SpacerNode):
            height = (
                target.height if target.height is not None else _default_row_height()
            )
            for offset in range(target.rows):
                row_index = placement.row + offset
                row_heights[row_index] = max(row_heights.get(row_index, 0.0), height)
        else:
            assert_never(target)

    _apply_dimensions(ws, col_widths, row_heights)
