"""
Módulo para exportar diccionarios de datos a diferentes formatos (CSV, Excel, JSON).
Movido e integrado desde /src para consolidar toda la funcionalidad en el paquete scavengr.

Author: Json Rivera
Date: 2024-09-26
Version: 1.2
"""

import csv
import json
import os
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    EXCEL_SUPPORT = True
except ImportError:
    print("openpyxl no está instalado, la exportación a Excel no estará disponible")
    print("Para instalar: pip install openpyxl")
    EXCEL_SUPPORT = False


class OutputWriter:
    """Clase para escribir diccionarios de datos en diferentes formatos."""
    
    def __init__(
        self,
        output_path: str,
        output_format: str,
        encoding: Optional[str] = "utf-8-sig",
    ):
        """
        Inicializar el escritor de salida.
        
        Args:
            output_path: Ruta del archivo de salida
            output_format: Formato de salida (csv, json, excel)
            encoding: Codificación del archivo
        """
        self.output_path = output_path
        self.output_format = output_format
        self.encoding = encoding

    def _validate_output_permissions(self):
        """Validar permisos de escritura en la ruta de salida."""
        output_dir = os.path.dirname(self.output_path) or "."
        if not os.access(output_dir, os.W_OK):
            raise PermissionError(
                f"No se puede escribir en el directorio: {output_dir}"
            )

    def write(self, data: List[Dict[str, Any]]):
        """
        Escribir los datos en el formato especificado.
        
        Args:
            data: Lista de diccionarios con los datos a escribir
            
        Raises:
            ValueError: Si el formato no está soportado
            Exception: Si ocurre un error durante la escritura
        """
        self._validate_output_permissions()
        if self.output_format == "csv":
            self._write_csv(data)
        elif self.output_format == "json":
            self._write_json(data)
        elif self.output_format == "excel":
            if not EXCEL_SUPPORT:
                raise ImportError(
                    "La exportación a Excel requiere openpyxl. "
                    "Por favor, instale la dependencia con: pip install openpyxl"
                )
            self._write_excel(data)
        else:
            raise ValueError(f"Formato no soportado: {self.output_format}")

    def _write_csv(self, data: List[Dict[str, Any]]):
        """
        Escribir los datos en formato CSV.
        
        Args:
            data: Lista de diccionarios con los datos
            
        Raises:
            Exception: Si ocurre un error durante la escritura
        """
        try:
            with open(self.output_path, "w", newline="", encoding=self.encoding) as f:
                if data:  # Verificar que hay datos
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    for row in data:
                        corrected_row = {}
                        for key, value in row.items():
                            # Asegurarse de que los valores sean serializables en CSV
                            if isinstance(value, str):
                                corrected_row[key] = value
                            else:
                                corrected_row[key] = value
                        writer.writerow(corrected_row)
                else:
                    # Archivo vacío si no hay datos
                    f.write("")
        except Exception as e:
            raise Exception(f"Error escribiendo CSV: {str(e)}")

    def _write_json(self, data: List[Dict[str, Any]]):
        """
        Escribir los datos en formato JSON.
        
        Args:
            data: Lista de diccionarios con los datos
            
        Raises:
            Exception: Si ocurre un error durante la escritura
        """
        try:
            with open(self.output_path, "w", encoding=self.encoding) as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            raise Exception(f"Error escribiendo JSON: {str(e)}")

    def _write_excel(self, data: List[Dict[str, Any]]):
        """
        Escribir los datos en formato Excel usando solo openpyxl.
        
        Args:
            data: Lista de diccionarios con los datos
            
        Raises:
            Exception: Si ocurre un error durante la escritura
        """
        try:
            # Crear libro de trabajo y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Diccionario_Datos"
            
            if not data:
                # Si no hay datos, crear archivo vacío
                wb.save(self.output_path)
                return
            
            # Obtener las columnas del primer registro
            headers = list(data[0].keys())
            
            # Escribir encabezados
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Escribir datos
            for row_idx, row_data in enumerate(data, 2):  # Empezar en fila 2
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Configurar anchos de columnas
            column_widths = {
                "TABLA": 20,
                "NOMBRE_DATO": 25,
                "TIPO_DE_DATO": 15,
                "LONGITUD": 10,
                "DESCRIPCION": 40,
                "REGEX": 30,
                "MASCARA": 15,
                "CRITERIOS_CALIDAD": 60,
                "TABLA_MAESTRA": 20,
                "CAMPO_MAESTRO": 20,
                "TIPO_RELACION": 10,
                "ES_MAESTRA": 10,
                "ES_PK": 8,
                "ES_NULLABLE": 10,
                "EJEMPLOS_VALIDOS": 25,
                "FUENTE_CAPTURA": 15,
                "MODULO": 20,
                "SENSIBILIDAD": 12,
                "OBSERVACIONES": 40,
            }

            # Ajustar anchos de columnas
            for col_idx, header in enumerate(headers, 1):
                col_letter = self._get_column_letter(col_idx)
                if header in column_widths:
                    ws.column_dimensions[col_letter].width = column_widths[header]
                else:
                    ws.column_dimensions[col_letter].width = 15

            # Configurar wrap text para columnas largas
            wrap_columns = ["DESCRIPCION", "CRITERIOS_CALIDAD", "OBSERVACIONES"]

            for header in wrap_columns:
                if header in headers:
                    col_idx = headers.index(header) + 1

                    # Aplicar wrap text a todas las celdas de la columna
                    for row_idx in range(2, len(data) + 2):  # +2 porque la fila 1 es el header
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        ws.row_dimensions[row_idx].height = 25  # Altura ajustada

            # Ajustar altura de filas automáticamente (aproximación)
            for row_idx in range(2, len(data) + 2):
                max_line_count = 1
                for header in wrap_columns:
                    if header in headers:
                        col_idx = headers.index(header) + 1
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value and isinstance(cell_value, str):
                            line_count = cell_value.count("\n") + 1
                            max_line_count = max(max_line_count, line_count)

                # Ajustar altura basado en número de líneas
                ws.row_dimensions[row_idx].height = 15 * max_line_count
            
            # Guardar archivo
            wb.save(self.output_path)

        except Exception as e:
            raise Exception(f"Error escribiendo Excel: {str(e)}")

    def _get_column_letter(self, col_idx: int) -> str:
        """
        Convertir índice de columna a letra (1 = A, 2 = B, etc.).
        
        Args:
            col_idx: Índice de columna (base 1)
            
        Returns:
            str: Letra de la columna
        """
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result