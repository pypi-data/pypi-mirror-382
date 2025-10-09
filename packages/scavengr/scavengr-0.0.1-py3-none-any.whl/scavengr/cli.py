#!/usr/bin/env python3
"""
🗃️ Scavengr CLI - "Descubre lo que tus bases esconden."
===============================================

CLI principal para Scavengr, proporciona comandos intuitivos para:
- Extraer metadatos de bases de datos
- Validar archivos DBML
- Generar diccionarios de datos
- Crear informes de análisis

Comandos disponibles:
- scavengr extract -o <output>                  → extraer metadatos y generar DBML
- scavengr validate -i <dbml>                   → validar DBML existente
- scavengr dictionary -i <dbml> -o <outputd>    → exportar diccionario
- scavengr report -i <dbml> -o <output>         → generar informe básico

Author: Jason Rivera
Version: 0.0.1
"""

import sys
import argparse
import os
from pathlib import Path
import json

# Importar utilidades centralizadas
from scavengr.utils import (
    setup_logging,
    Commands,
    Formats,
    ScavengrError,
    ProcessingError,
    validate_file_exists,
    validate_output_format,
    validate_write_permissions,
    validate_input_file_format,
    provide_user_feedback,
)

# Agregar el directorio actual al path para importaciones
sys.path.append(os.path.dirname(os.path.abspath(__file__)))


# ===================================================================
# LAZY LOADING DE MÓDULOS PESADOS
# ===================================================================

def get_dbml_parser():
    """Lazy loading del parser DBML (usado por report_command)."""
    # from parsers.dbml_parser import DBMLParser
    from scavengr.infrastructure.parsers.dbml_parser import DBMLParser
    return DBMLParser

# Configuración de logging: usar el módulo centralizado en scavengr.utils
logger = setup_logging()


# ===================================================================
# CLASE PRINCIPAL DEL CLI
# ===================================================================

class ScavengrCLI:
    """Clase principal del CLI de Scavengr"""
    
    def __init__(self):
        self.version = "0.0.1"
    
    def extract_command(self, args):
        """Comando para extraer metadatos y generar DBML usando caso de uso."""
        try:
            logger.info("[EXTRACT] Iniciando extracción de metadatos...")
            
            # Cargar configuración
            from scavengr.config.env_config import EnvConfigManager
            env_manager = EnvConfigManager(getattr(args, 'env_file', None))
            
            # Validar configuración
            validation = env_manager.validate_config()
            if not validation['valid']:
                logger.error("[ERROR] Configuración inválida:")
                for issue in validation['issues']:
                    logger.error(f"  - {issue}")
                return False

            if validation['warnings']:
                for warning in validation['warnings']:
                    logger.warning(f"[WARNING] {warning}")
            
            db_config = env_manager.get_db_config()
            generation_config = env_manager.get_generation_config()
            
            logger.info(f"[INFO] Sistema: {generation_config['source_system']['name']}")
            logger.info(f"[INFO] BD configurada: {db_config['type']} en {db_config['host']}")

            # Determinar archivo de salida
            output_file = args.output or f"{generation_config['file_prefix']}_extracted.dbml"

            # Feedback al usuario
            provide_user_feedback(
                f"Conectando a {db_config['type']} en {db_config['host']}",
                "info"
            )
            
            # Ejecutar caso de uso
            from scavengr.application import ExtractMetadata
            use_case = ExtractMetadata(db_config, generation_config)
            result = use_case.execute(output_file)
            
            # Procesar resultado
            if result.success:
                provide_user_feedback(
                    f"Esquema extraído exitosamente: {result.tables_count} tablas procesadas",
                    "success"
                )
                logger.info(
                    f"[SUCCESS] Extracción completada: {result.output_path} "
                    f"({result.file_size:,} bytes)"
                )
                logger.info(
                    f"[STATS] {result.tables_count} tablas, "
                    f"{result.columns_count} columnas, "
                    f"{result.relationships_count} relaciones"
                )
                return True
            else:
                logger.error(f"[ERROR] Error en extracción: {result.error_message}")
                provide_user_feedback(
                    f"Error: {result.error_message}",
                    "error"
                )
                return False
                
        except Exception as e:
            logger.error(f"[ERROR] Error en extracción: {str(e)}")
            provide_user_feedback(f"Error inesperado: {str(e)}", "error")
            return False
    
    def validate_command(self, args):
        """Comando para validar archivos DBML usando caso de uso."""
        try:
            logger.info(f"[VALIDATE] Validando archivo DBML: {args.input}")
            
            # Validar que el archivo existe
            input_path = validate_file_exists(args.input, "Archivo DBML")
            provide_user_feedback(f"Validando: {input_path}", "info")
            
            # Ejecutar caso de uso
            from scavengr.application import ValidateDBML
            use_case = ValidateDBML()
            result = use_case.execute(str(input_path))
            
            # Procesar resultado
            if result.is_valid:
                logger.info("[SUCCESS] Archivo DBML válido")
                logger.info(
                    f"[STATS] {result.tables_count} tablas, "
                    f"{result.relationships_count} relaciones"
                )
                provide_user_feedback(
                    f"Validación exitosa: {result.tables_count} tablas encontradas",
                    "success"
                )
                
                # Mostrar advertencias si las hay
                if result.has_warnings():
                    logger.warning(f"[WARNING] {len(result.get_warnings())} advertencias encontradas:")
                    for warning in result.get_warnings():
                        logger.warning(f"  - {warning.message}")
                
                return True
            else:
                logger.error("[ERROR] Archivo DBML inválido")
                logger.error(f"[ERROR] {len(result.get_errors())} errores encontrados:")
                
                # Mostrar errores
                for error in result.get_errors():
                    error_msg = error.message
                    if error.line:
                        error_msg = f"Línea {error.line}: {error_msg}"
                    logger.error(f"  - {error_msg}")
                    provide_user_feedback(error_msg, "error")
                
                return False
                
        except Exception as e:
            logger.error(f"[ERROR] Error al validar DBML: {str(e)}")
            provide_user_feedback(f"Error inesperado: {str(e)}", "error")
            return False
    
    def dictionary_command(self, args):
        """Comando para generar diccionarios de datos usando caso de uso."""
        try:
            logger.info(f"[DICTIONARY] Generando diccionario desde: {args.input}")
            
            # Validar archivo de entrada
            input_path = validate_input_file_format(args.input)
            provide_user_feedback(f"Archivo de entrada validado: {input_path}", "success")
            
            # Validar permisos de escritura
            try:
                output_path = validate_write_permissions(args.output)
                provide_user_feedback(
                    f"Permisos de escritura verificados: {output_path.parent}",
                    "info"
                )
            except Exception as e:
                provide_user_feedback(
                    f"No se puede escribir en: {args.output}. Detalle: {str(e)}",
                    "error"
                )
                logger.error(f"[ERROR] Permisos de escritura: {str(e)}")
                return False
            
            # Determinar formato de salida
            output_format = validate_output_format(args.output, args.format)
            logger.info(f"[INFO] Formato de salida: {output_format.upper()}")
            
            # Cargar configuración para contexto
            from scavengr.config.env_config import EnvConfigManager
            env_manager = EnvConfigManager(getattr(args, 'env_file', None))
            generation_config = env_manager.get_generation_config()
            
            # Ejecutar caso de uso
            from scavengr.application import GenerateDictionary
            use_case = GenerateDictionary(config=generation_config)
            result = use_case.execute(
                input_path=str(input_path),
                output_path=args.output,
                output_format=output_format
            )
            
            # Procesar resultado
            if result.success:
                logger.info(f"[SUCCESS] Diccionario generado: {result.output_path}")
                logger.info(
                    f"[STATS] {result.entries_count} entradas, "
                    f"formato {result.format.upper()}, "
                    f"{result.file_size:,} bytes"
                )
                provide_user_feedback(
                    f"Diccionario generado exitosamente: {result.entries_count} entradas",
                    "success"
                )
                return True
            else:
                logger.error(f"[ERROR] Error generando diccionario: {result.error_message}")
                provide_user_feedback(
                    f"Error: {result.error_message}",
                    "error"
                )
                return False
            
        except Exception as e:
            logger.error(f"[ERROR] Error al generar diccionario: {str(e)}")
            provide_user_feedback(f"Error inesperado: {str(e)}", "error")
            return False
    
    def report_command(self, args):
        """Comando para generar informes avanzados con análisis estadístico completo."""
        try:
            logger.info(f"[REPORT] Generando informe desde: {args.input}")

            if not os.path.exists(args.input):
                logger.error(f"[ERROR] Archivo DBML no encontrado: {args.input}")
                return False

            # Parsear DBML (lazy loading)
            DBMLParser = get_dbml_parser()
            parser = DBMLParser(args.input)
            schema = parser.parse()

            if not schema:
                logger.error("[ERROR] No se pudo parsear el archivo DBML")
                return False

            # Generar estadísticas avanzadas usando StatisticsAnalyzerService
            from scavengr.core.services import StatisticsAnalyzerService
            stats_service = StatisticsAnalyzerService()
            advanced_stats = stats_service.analyze_schema(schema)

            # Metadatos del informe
            from datetime import datetime
            metadatos = {
                "version_scavengr": self.version,
                "fecha_generacion": datetime.now().isoformat(timespec='seconds'), # Formato ISO sin microsegundos: 2024-06-25T14:30:00
                "generado_en": str(Path().cwd()), 
                "archivo_origen": os.path.abspath(args.input),
                "archivo_salida": os.path.abspath(args.output),
                "usuario": os.getenv("USERNAME") or os.getenv("USER") or "desconocido",
            }

            # Crear directorio de salida si no existe
            output_dir = os.path.dirname(args.output)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            # Determinar formato de salida
            output_ext = Path(args.output).suffix.lower()

            if output_ext == '.xlsx':
                # Generar informe Excel avanzado
                self._write_advanced_excel_report(args.output, metadatos, advanced_stats)
            else:
                # Crear informe JSON completo
                report = {
                    "metadatos": metadatos,
                    "analisis": advanced_stats
                }
                with open(args.output, 'w', encoding='utf-8') as f:
                    json.dump(report, f, indent=2, ensure_ascii=False)

            # Mostrar resumen en consola
            resumen = advanced_stats['resumen_general']
            logger.info(f"[SUCCESS] Informe generado: {args.output}")
            logger.info(f"[STATS] Tablas analizadas: {resumen['total_tablas']}")
            logger.info(f"[STATS] Columnas analizadas: {resumen['total_columnas']}")
            logger.info(f"[STATS] Score de calidad: {advanced_stats['calidad_datos']['score_general']}%")
            return True

        except Exception as e:
            logger.error(f"[ERROR] Error al generar informe: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _write_advanced_excel_report(self, output_path: str, metadatos: dict, stats: dict):
        """Escribir informe avanzado en formato Excel con múltiples hojas."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, Reference

            wb = Workbook()
            
            # Estilos reutilizables
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # === HOJA 1: RESUMEN GENERAL ===
            ws_resumen = wb.active
            ws_resumen.title = "Resumen General"
            
            ws_resumen['A1'] = "📊 INFORME DE ANÁLISIS DE BASE DE DATOS"
            ws_resumen['A1'].font = Font(bold=True, size=18, color="FFFFFF")
            ws_resumen['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws_resumen.merge_cells('A1:C1')
            
            # Metadatos
            row = 3
            ws_resumen[f'A{row}'] = "Fecha de Generación:"
            ws_resumen[f'B{row}'] = metadatos['fecha_generacion']
            row += 1
            ws_resumen[f'A{row}'] = "Versión Scavengr:"
            ws_resumen[f'B{row}'] = metadatos['version_scavengr']
            row += 1
            ws_resumen[f'A{row}'] = "Usuario:"
            ws_resumen[f'B{row}'] = metadatos['usuario']
            
            # Resumen General
            row += 2
            resumen = stats['resumen_general']
            ws_resumen[f'A{row}'] = "MÉTRICAS GENERALES"
            ws_resumen[f'A{row}'].font = subheader_font
            row += 1
            
            metricas_generales = [
                ("📋 Total de Tablas", resumen['total_tablas']),
                ("📊 Total de Columnas", resumen['total_columnas']),
                ("🔗 Total de Relaciones", resumen['total_relaciones']),
                ("📈 Promedio Columnas/Tabla", resumen['promedio_columnas_por_tabla']),
                ("🏢 Tablas Maestras", resumen['tablas_maestras']),
                ("💼 Tablas Transaccionales", resumen['tablas_transaccionales']),
            ]
            
            for label, valor in metricas_generales:
                ws_resumen[f'A{row}'] = label
                ws_resumen[f'B{row}'] = valor
                ws_resumen[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Score de Calidad
            row += 1
            calidad = stats['calidad_datos']
            ws_resumen[f'A{row}'] = "SCORE DE CALIDAD"
            ws_resumen[f'A{row}'].font = subheader_font
            row += 1
            
            ws_resumen[f'A{row}'] = "⭐ Score General"
            ws_resumen[f'B{row}'] = f"{calidad['score_general']}%"
            ws_resumen[f'B{row}'].font = Font(bold=True, size=14, color="FF0000" if calidad['score_general'] < 50 else "00B050")
            
            ws_resumen.column_dimensions['A'].width = 30
            ws_resumen.column_dimensions['B'].width = 20
            
            # === HOJA 2: CALIDAD DE DATOS ===
            ws_calidad = wb.create_sheet("Calidad de Datos")
            ws_calidad['A1'] = "ANÁLISIS DE CALIDAD DE DATOS"
            ws_calidad['A1'].font = header_font
            ws_calidad['A1'].fill = header_fill
            ws_calidad.merge_cells('A1:C1')
            
            row = 3
            dimensiones = [
                ("Completitud", calidad['completitud']),
                ("Integridad Referencial", calidad['integridad_referencial']),
                ("Consistencia Nombres", calidad['consistencia_nombres']),
                ("Documentación", calidad['documentacion']),
            ]
            
            for dimension, datos in dimensiones:
                ws_calidad[f'A{row}'] = dimension
                ws_calidad[f'B{row}'] = f"{datos['score']}%"
                ws_calidad[f'A{row}'].font = Font(bold=True)
                row += 1
            
            ws_calidad.column_dimensions['A'].width = 30
            ws_calidad.column_dimensions['B'].width = 15
            
            # === HOJA 3: DISTRIBUCIÓN DE TIPOS ===
            ws_tipos = wb.create_sheet("Tipos de Datos")
            ws_tipos['A1'] = "DISTRIBUCIÓN DE TIPOS DE DATOS"
            ws_tipos['A1'].font = header_font
            ws_tipos['A1'].fill = header_fill
            ws_tipos.merge_cells('A1:C1')
            
            row = 3
            ws_tipos[f'A{row}'] = "Tipo de Dato"
            ws_tipos[f'B{row}'] = "Cantidad"
            ws_tipos[f'C{row}'] = "Porcentaje"
            for col in ['A', 'B', 'C']:
                ws_tipos[f'{col}{row}'].font = Font(bold=True)
                ws_tipos[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            distribucion = stats['distribucion_tipos']
            for item in distribucion['top_10_tipos']:
                ws_tipos[f'A{row}'] = item['tipo']
                ws_tipos[f'B{row}'] = item['cantidad']
                ws_tipos[f'C{row}'] = f"{item['cantidad'] / distribucion['distribucion_categorias']['texto']['cantidad'] * 100:.1f}%"
                row += 1
            
            ws_tipos.column_dimensions['A'].width = 25
            ws_tipos.column_dimensions['B'].width = 12
            ws_tipos.column_dimensions['C'].width = 15
            
            # === HOJA 4: SEGURIDAD ===
            ws_seguridad = wb.create_sheet("Seguridad")
            ws_seguridad['A1'] = "ANÁLISIS DE SEGURIDAD Y SENSIBILIDAD"
            ws_seguridad['A1'].font = header_font
            ws_seguridad['A1'].fill = header_fill
            ws_seguridad.merge_cells('A1:C1')
            
            row = 3
            seguridad = stats['seguridad_sensibilidad']
            ws_seguridad[f'A{row}'] = "Nivel"
            ws_seguridad[f'B{row}'] = "Cantidad"
            ws_seguridad[f'C{row}'] = "Porcentaje"
            for col in ['A', 'B', 'C']:
                ws_seguridad[f'{col}{row}'].font = Font(bold=True)
            row += 1
            
            for nivel, datos in seguridad['distribucion_sensibilidad'].items():
                ws_seguridad[f'A{row}'] = nivel
                ws_seguridad[f'B{row}'] = datos['cantidad']
                ws_seguridad[f'C{row}'] = f"{datos['porcentaje']}%"
                row += 1
            
            # === HOJA 5: RECOMENDACIONES ===
            ws_recs = wb.create_sheet("Recomendaciones")
            ws_recs['A1'] = "RECOMENDACIONES DE MEJORA"
            ws_recs['A1'].font = header_font
            ws_recs['A1'].fill = header_fill
            ws_recs.merge_cells('A1:D1')
            
            row = 3
            ws_recs[f'A{row}'] = "Prioridad"
            ws_recs[f'B{row}'] = "Categoría"
            ws_recs[f'C{row}'] = "Recomendación"
            ws_recs[f'D{row}'] = "Tablas Afectadas"
            for col in ['A', 'B', 'C', 'D']:
                ws_recs[f'{col}{row}'].font = Font(bold=True)
            row += 1
            
            for rec in stats['recomendaciones']:
                ws_recs[f'A{row}'] = rec['prioridad']
                ws_recs[f'B{row}'] = rec['categoria']
                ws_recs[f'C{row}'] = rec['recomendacion']
                ws_recs[f'D{row}'] = ", ".join(rec['tablas_afectadas'][:3]) if rec['tablas_afectadas'] else "-"
                
                # Color según prioridad
                if rec['prioridad'] == 'CRÍTICA':
                    ws_recs[f'A{row}'].font = Font(color="FF0000", bold=True)
                elif rec['prioridad'] == 'ALTA':
                    ws_recs[f'A{row}'].font = Font(color="FFC000", bold=True)
                
                row += 1
            
            ws_recs.column_dimensions['A'].width = 12
            ws_recs.column_dimensions['B'].width = 20
            ws_recs.column_dimensions['C'].width = 50
            ws_recs.column_dimensions['D'].width = 30
            
            wb.save(output_path)
            
        except ImportError:
            # Fallback: escribir como JSON si no hay openpyxl
            report = {
                "metadatos": metadatos,
                "analisis": stats
            }
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
    
    def _write_excel_report(self, output_path, stats):
        """Escribir informe en formato Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Informe Análisis"

            # Estilo de encabezado
            header_font = Font(bold=True, size=16, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Título del informe con estilo
            ws['A1'] = "📊 INFORME DE ANÁLISIS - SCAVENGR"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws.merge_cells('A1:B1')

            # Datos del informe
            row = 3
            data = [
                ("📈 Total de Tablas", stats['total_tables']),
                ("📋 Total de Columnas", stats['total_columns']),
                ("🔗 Total de Relaciones", stats['total_relations']),
                ("🔑 Tablas con Clave Primaria", stats['tables_with_primary_key']),
                ("📊 Porcentaje de Cobertura", f"{stats['coverage_percentage']}%"),
            ]

            for label, value in data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            # Tipos de datos más comunes
            if stats.get('top_data_types'):
                row += 1
                ws[f'A{row}'] = "🏷️ Tipos de Datos Más Comunes"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1

                for dtype, count in stats['top_data_types']:
                    ws[f'A{row}'] = f"  • {dtype}"
                    ws[f'B{row}'] = count
                    row += 1

            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20

            wb.save(output_path)
            
        except ImportError:
            # Fallback: escribir como JSON si no hay openpyxl
            report = {
                "version_scavengr": self.version,
                "estadisticas": stats
            }
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
    
    def _generate_basic_stats(self, schema):
        """Generar estadísticas básicas del schema"""
        tables = schema.get('tables', [])
        
        total_tables = len(tables)
        total_columns = sum(len(table.columns) for table in tables)
        total_relations = len(schema.get('relationships', []))
        
        # Estadísticas por tabla  
        tables_with_pk = sum(1 for table in tables if any(
            col.is_pk for col in table.columns
        ))
        
        # Tipos de datos más comunes
        data_types = {}
        for table in tables:
            for column in table.columns:
                dtype = column.type or 'unknown'
                data_types[dtype] = data_types.get(dtype, 0) + 1
        
        top_data_types = sorted(data_types.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return {
            "total_tables": total_tables,
            "total_columns": total_columns,
            "total_relations": total_relations,
            "tables_with_primary_key": tables_with_pk,
            "coverage_percentage": round((tables_with_pk / total_tables * 100) if total_tables > 0 else 0, 2),
            "top_data_types": top_data_types
        }


def setup_argument_parser() -> argparse.ArgumentParser:
    """
    Configura y retorna el parser de argumentos CLI.
    
    Returns:
        ArgumentParser configurado con todos los comandos y opciones.
    """
    cli = ScavengrCLI()

    # Parser principal
    parser = argparse.ArgumentParser(
        prog='scavengr',
        description='🗃️ Scavengr - Descubre lo que tus bases esconden',
        epilog='Para más información, visite: https://github.com/JasRockr/Scavengr'
    )

    parser.add_argument(
        '-v', '--version',
        action='version', 
        version=f'Scavengr {cli.version}',
        help='Muestra la versión actual de Scavengr y sale.'
    )

    # # Agregar argumento verbose
    # parser.add_argument(
    #     '--verbose',
    #     action='store_true',
    #     help='Habilita el modo detallado (DEBUG) para el logging.'
    # )

    # Subcommands
    subparsers = parser.add_subparsers(dest='command', help='Comandos disponibles')

    # Comando extract
    extract_parser = subparsers.add_parser(
        Commands.EXTRACT, 
        help='Extraer metadatos de base de datos y generar DBML'
    )
    extract_parser.add_argument(
        '-o', '--output',
        help="Archivo DBML de salida (por defecto: <prefijo>_extracted.dbml)"
    )
    extract_parser.add_argument(
        '--env-file',
        help='Archivo .env específico a usar (por defecto: .env)'
    )

    # Comando validate
    validate_parser = subparsers.add_parser(
        Commands.VALIDATE,
        help='Validar archivo DBML existente'
    )
    validate_parser.add_argument(
        '-i', '--input',
        required=True,
        help='Ruta al archivo DBML a validar'
    )

    # Comando dictionary
    dict_parser = subparsers.add_parser(
        Commands.DICTIONARY,
        help='Generar diccionario de datos desde DBML'
    )
    dict_parser.add_argument(
        '-i', '--input',
        required=True,
        help='Ruta al archivo DBML de entrada'
    )
    dict_parser.add_argument(
        '-o', '--output',
        required=True,
        help='Ruta del archivo de salida'
    )
    dict_parser.add_argument(
        '-f', '--format',
        choices=Formats.SUPPORTED,
        help='Formato de salida (se detecta automáticamente por extensión)'
    )
    
    # Comando report
    report_parser = subparsers.add_parser(
        Commands.REPORT,
        help='Generar informe con análisis desde archivo DBML'
    )
    report_parser.add_argument(
        '-i', '--input',
        required=True,
        help='Ruta al archivo DBML de entrada'
    )
    report_parser.add_argument(
        '-o', '--output',
        required=True,
        help='Ruta del archivo de informe de salida (JSON)'
    )
    
    return parser


def execute_command(command: str, args: argparse.Namespace) -> bool:
    """
    Dispatcher de comandos - ejecuta el comando correspondiente.
    
    Args:
        command: Nombre del comando a ejecutar
        args: Argumentos parseados
        
    Returns:
        bool: True si el comando se ejecutó exitosamente
        
    Raises:
        ProcessingError: Si hay error en el procesamiento
    """
    cli = ScavengrCLI()
    
    command_map = {
        Commands.EXTRACT: cli.extract_command,
        Commands.VALIDATE: cli.validate_command,
        Commands.DICTIONARY: cli.dictionary_command,
        Commands.REPORT: cli.report_command
    }
    
    command_func = command_map.get(command)
    if not command_func:
        raise ProcessingError("comando", f"Comando desconocido: {command}")
    
    return command_func(args)


def main():
    """Función principal del CLI - simplificada y modular."""
    parser = setup_argument_parser()
    args = parser.parse_args()
    
    # Configurar logging según el argumento verbose
    if args.verbose:
        setup_logging(verbose=True)
        logger.info("Modo detallado (DEBUG) activado")
    
    if not args.command:
        parser.print_help()
        return 1
    
    # Ejecutar comando usando el dispatcher
    try:
        success = execute_command(args.command, args)
        return 0 if success else 1
        
    except KeyboardInterrupt:
        logger.info("[CANCELLED] Operacion cancelada por el usuario")
        return 1
    except ScavengrError as e:
        logger.error(f"[ERROR] {str(e)}")
        return 1
    except Exception as e:
        logger.error(f"[ERROR] Error inesperado: {str(e)}")
        return 1


if __name__ == '__main__':
        sys.exit(main())