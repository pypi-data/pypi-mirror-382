"""
Excel表格处理器
基于Dict的Excel文件操作，提供类似MySQL的接口
"""

import os
import shutil
from datetime import datetime
from typing import Dict, List, Union, Optional, Any
from pathlib import Path
from loguru import logger
from .filter_processor import ExcelFilterProcessor
from .workbook_manager import WorkbookManager

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.debug("openpyxl库未安装，请运行: pip install openpyxl")
    raise ImportError("openpyxl库未安装，Excel功能不可用")


class ExcelHandler:
    """
    Excel表格处理器，提供类似MySQL的接口
    
    主要功能：
    - select() - 查询数据，支持where条件、字段选择、排序、分页
    - insert() - 插入数据，支持单条或批量
    - update() - 更新数据，支持where条件
    - delete() - 删除数据，支持where条件
    - 支持index虚拟字段（Excel行号）
    - 支持多工作表操作
    - 支持原生openpyxl对象操作
    """

    def __init__(self, excel_path: str, header_row: int = 1, sheet_name: Optional[str] = None,
                 auto_create_excel: bool = False, auto_sheet: bool = False, auto_backup: bool = False,
                 check_file_lock: bool = True):
        """
        初始化Excel处理器

        Args:
            excel_path: Excel文件路径
            header_row: 表头所在行号（1为首行）
            sheet_name: 工作表名，默认None表示使用活动工作表
            auto_create_excel: Excel文件不存在时是否自动创建
            auto_sheet: 指定的工作表不存在时是否自动创建
            auto_backup: 是否自动备份Excel文件
            check_file_lock: 是否检查文件占用状态，默认True检查
        """
        logger.debug("初始化Excel处理器，文件: {}，检查文件占用: {}", excel_path, check_file_lock)

        self.excel_path = str(Path(excel_path).resolve())
        self.header_row = header_row
        self.sheet_name = sheet_name
        self.auto_create_excel = auto_create_excel
        self.auto_sheet = auto_sheet
        self.auto_backup = auto_backup
        self.check_file_lock = check_file_lock
        
        # 数据缓存
        self.data: Optional[List[Dict[str, Any]]] = None
        self.headers: Optional[List[str]] = None
        self._modified = False
        
        # 组件初始化
        self._filter_processor = ExcelFilterProcessor()
        self._workbook_manager = WorkbookManager(self.excel_path, self.sheet_name)
        
        # 自动备份
        if self.auto_backup and os.path.exists(self.excel_path):
            self._create_backup()

        # 检查文件是否被占用（可选）
        if self.check_file_lock and os.path.exists(self.excel_path):
            self._workbook_manager.check_file_lock()

    def __del__(self):
        """析构函数，自动清理资源"""
        try:
            self.close()
        except:
            pass

    def __enter__(self):
        """上下文管理器入口"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()

    def select(self, where: Optional[Dict[str, Any]] = None, fields: Optional[List[str]] = None,
               order_by: Optional[Dict[str, str]] = None, limit: Optional[int] = None,
               sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        查询数据（类似MySQL的SELECT）
        
        Args:
            where: 查询条件字典，支持操作符（如 {"age__gte": 18}）
            fields: 要查询的字段列表，None表示查询所有字段
            order_by: 排序字典，如 {"age": "DESC", "name": "ASC"}
            limit: 限制返回记录数
            sheet_name: 指定工作表名，None使用初始化时的工作表
            
        Returns:
            List[Dict]: 查询结果列表，每个字典包含'index'字段表示Excel行号
            
        Example:
            # 查询所有数据
            data = handler.select()
            
            # 条件查询
            data = handler.select(where={"age__gte": 18, "status": "active"})
            
            # 字段选择和排序
            data = handler.select(fields=["name", "age"], order_by={"age": "DESC"})
            
            # 分页查询
            data = handler.select(limit=10)
            
            # 行号条件（index为虚拟字段）
            data = handler.select(where={"index__gte": 5})
        """
        logger.debug("查询Excel数据，工作表: {}，条件: {}", sheet_name or self.sheet_name, where)
        
        # 处理工作表切换
        original_sheet = self.sheet_name
        if sheet_name is not None:
            self.sheet_name = sheet_name
            self.data = None  # 重置数据以强制重新读取
        
        try:
            # 确保数据已加载
            if self.data is None:
                self._read_file()
            
            result = self.data.copy() if self.data else []
            
            # 应用where条件过滤
            if where:
                result = self._filter_processor.filter_data(result, where)
            
            # 字段选择
            if fields:
                # 确保index字段总是包含（如果原数据有的话）
                if result and 'index' in result[0] and 'index' not in fields:
                    fields = ['index'] + fields
                
                filtered_result = []
                for row in result:
                    filtered_row = {field: row.get(field) for field in fields if field in row}
                    filtered_result.append(filtered_row)
                result = filtered_result
            
            # 排序
            if order_by:
                result = self._sort_data(result, order_by)
            
            # 限制记录数
            if limit is not None and limit > 0:
                result = result[:limit]
            
            logger.debug("查询完成，返回记录数: {}，字段: {}，排序: {}，限制: {}",
                        len(result), fields or "全部", order_by, limit)
            if result and len(result) <= 5:  # 少量数据时显示具体内容
                logger.debug("查询结果: {}", result)
            return result
            
        finally:
            # 如果是临时切换，恢复原来的设置
            if sheet_name is not None:
                self.sheet_name = original_sheet
                self.data = None

    def insert(self, data: Union[Dict[str, Any], List[Dict[str, Any]]], 
               sheet_name: Optional[str] = None, auto_save: bool = True) -> Union[int, List[int]]:
        """
        插入数据（类似MySQL的INSERT）
        
        Args:
            data: 要插入的数据字典或字典列表
            sheet_name: 指定工作表名，None使用初始化时的工作表
            auto_save: 是否自动保存
            
        Returns:
            Union[int, List[int]]: 单条插入返回行号，批量插入返回行号列表
            
        Example:
            # 插入单条数据
            row_id = handler.insert({"name": "张三", "age": 25})
            
            # 批量插入
            row_ids = handler.insert([
                {"name": "李四", "age": 30},
                {"name": "王五", "age": 28}
            ])
        """
        logger.debug("插入Excel数据，工作表: {}，数据类型: {}，记录数: {}",
                    sheet_name or self.sheet_name,
                    "单条" if isinstance(data, dict) else "批量",
                    1 if isinstance(data, dict) else len(data))
        
        # 处理工作表切换
        if sheet_name is not None:
            temp_handler = ExcelHandler(
                excel_path=self.excel_path,
                header_row=self.header_row,
                sheet_name=sheet_name,
                auto_create_excel=self.auto_create_excel,
                auto_sheet=self.auto_sheet,
                auto_backup=False,  # 避免重复备份
                check_file_lock=self.check_file_lock  # 继承文件锁检查设置
            )
            return temp_handler.insert(data, auto_save=auto_save)
        
        # 标准化输入
        if isinstance(data, dict):
            data_list = [data]
            is_single = True
        else:
            data_list = data
            is_single = False
        
        if not data_list:
            logger.debug("插入数据不能为空")
            raise ValueError("插入数据不能为空")
        
        # 移除index字段（如果存在）
        data_list = [{k: v for k, v in d.items() if k != 'index'} for d in data_list]

        # 记录插入的数据详情
        if len(data_list) <= 3:  # 少量数据时显示具体内容
            logger.debug("插入数据详情: {}", data_list)
        else:
            logger.debug("插入数据示例（前3条）: {}", data_list[:3])
        
        # 确保已读取数据
        if self.data is None:
            self._read_file()
        
        # 添加数据并分配行号
        if not self.data:
            self.data = []
        
        # 更新表头
        all_keys = set()
        for item in data_list:
            all_keys.update(item.keys())
        
        if not self.headers:
            self.headers = list(all_keys)
        else:
            for key in all_keys:
                if key not in self.headers:
                    self.headers.append(key)
        
        # 插入数据并分配行号
        insert_ids = []
        start_index = self.header_row + 1 + len(self.data)
        
        for i, item in enumerate(data_list):
            row_data = item.copy()
            row_index = start_index + i
            row_data['index'] = row_index
            self.data.append(row_data)
            insert_ids.append(row_index)
        
        self._modified = True
        
        if auto_save:
            self.save()
        
        logger.debug("插入完成，插入记录数: {}，分配行号: {}", len(data_list), insert_ids)
        if len(insert_ids) <= 5:  # 少量数据时显示行号详情
            logger.debug("插入行号详情: {}", insert_ids)
        return insert_ids[0] if is_single else insert_ids

    def update(self, data: Dict[str, Any], where: Dict[str, Any],
               sheet_name: Optional[str] = None, auto_save: bool = True) -> int:
        """
        更新数据（类似MySQL的UPDATE）

        Args:
            data: 要更新的数据字典
            where: 更新条件字典
            sheet_name: 指定工作表名，None使用初始化时的工作表
            auto_save: 是否自动保存

        Returns:
            int: 更新的行数

        Example:
            # 根据条件更新
            count = handler.update({"status": "completed"}, where={"name": "张三"})

            # 根据行号更新
            count = handler.update({"age": 26}, where={"index": 2})

            # 复杂条件更新
            count = handler.update({"status": "active"}, where={"age__gte": 18})
        """
        logger.debug("更新Excel数据，工作表: {}，更新数据: {}，条件: {}",
                    sheet_name or self.sheet_name, data, where)

        # 处理工作表切换
        if sheet_name is not None:
            temp_handler = ExcelHandler(
                excel_path=self.excel_path,
                header_row=self.header_row,
                sheet_name=sheet_name,
                auto_create_excel=self.auto_create_excel,
                auto_sheet=self.auto_sheet,
                auto_backup=False,  # 避免重复备份
                check_file_lock=self.check_file_lock  # 继承文件锁检查设置
            )
            return temp_handler.update(data, where, auto_save=auto_save)

        if not data:
            logger.debug("更新数据不能为空")
            raise ValueError("更新数据不能为空")

        if not where:
            logger.debug("更新条件不能为空")
            raise ValueError("更新条件不能为空")

        # 移除index字段（如果存在）
        update_data = {k: v for k, v in data.items() if k != 'index'}

        # 确保已读取数据
        if self.data is None:
            self._read_file()

        if not self.data:
            logger.debug("没有数据可更新")
            return 0

        # 查找匹配的行并更新
        updated_count = 0
        updated_rows = []
        for row in self.data:
            if self._filter_processor._match_where_conditions(row, where):
                old_values = {key: row.get(key) for key in update_data.keys()}
                for key, value in update_data.items():
                    row[key] = value
                    # 更新表头
                    if key not in self.headers:
                        self.headers.append(key)
                updated_rows.append({
                    'index': row.get('index'),
                    'old': old_values,
                    'new': {key: value for key, value in update_data.items()}
                })
                updated_count += 1

        if updated_count > 0:
            self._modified = True
            if auto_save:
                self.save()

        logger.debug("更新完成，更新行数: {}", updated_count)
        if updated_rows and len(updated_rows) <= 5:  # 少量数据时显示更新详情
            logger.debug("更新详情: {}", updated_rows)
        elif updated_rows:
            logger.debug("更新详情（前3条）: {}", updated_rows[:3])
        return updated_count

    def delete(self, where: Dict[str, Any], sheet_name: Optional[str] = None, auto_save: bool = True) -> int:
        """
        删除数据（类似MySQL的DELETE）

        Args:
            where: 删除条件字典
            sheet_name: 指定工作表名，None使用初始化时的工作表
            auto_save: 是否自动保存

        Returns:
            int: 删除的行数

        Example:
            # 根据条件删除
            count = handler.delete(where={"status": "inactive"})

            # 根据行号删除
            count = handler.delete(where={"index": 5})

            # 复杂条件删除
            count = handler.delete(where={"age__lt": 18})
        """
        logger.debug("删除Excel数据，工作表: {}，条件: {}", sheet_name or self.sheet_name, where)

        # 处理工作表切换
        if sheet_name is not None:
            temp_handler = ExcelHandler(
                excel_path=self.excel_path,
                header_row=self.header_row,
                sheet_name=sheet_name,
                auto_create_excel=self.auto_create_excel,
                auto_sheet=self.auto_sheet,
                auto_backup=False,  # 避免重复备份
                check_file_lock=self.check_file_lock  # 继承文件锁检查设置
            )
            return temp_handler.delete(where, auto_save=auto_save)

        if not where:
            logger.debug("删除条件不能为空")
            raise ValueError("删除条件不能为空")

        # 确保已读取数据
        if self.data is None:
            self._read_file()

        if not self.data:
            logger.debug("没有数据可删除")
            return 0

        # 查找匹配的行（要删除的行）
        original_count = len(self.data)
        deleted_rows = []
        remaining_rows = []

        for row in self.data:
            if self._filter_processor._match_where_conditions(row, where):
                deleted_rows.append({
                    'index': row.get('index'),
                    'data': {k: v for k, v in row.items() if k != 'index'}
                })
            else:
                remaining_rows.append(row)

        self.data = remaining_rows
        deleted_count = len(deleted_rows)

        if deleted_count > 0:
            # 重新分配行号
            for i, row in enumerate(self.data):
                row['index'] = self.header_row + 1 + i

            self._modified = True
            if auto_save:
                self.save()

        logger.debug("删除完成，删除行数: {}", deleted_count)
        if deleted_rows and len(deleted_rows) <= 5:  # 少量数据时显示删除详情
            logger.debug("删除详情: {}", deleted_rows)
        elif deleted_rows:
            logger.debug("删除详情（前3条）: {}", deleted_rows[:3])
        return deleted_count

    def exists(self, where: Dict[str, Any], sheet_name: Optional[str] = None) -> bool:
        """
        判断是否存在满足条件的数据

        Args:
            where: 查询条件字典
            sheet_name: 指定工作表名，None使用初始化时的工作表

        Returns:
            bool: 是否存在满足条件的数据

        Example:
            # 判断是否存在指定用户
            exists = handler.exists(where={"name": "张三"})

            # 判断是否存在高薪员工
            exists = handler.exists(where={"salary>=": 10000})

            # 判断指定行号是否存在
            exists = handler.exists(where={"index": 5})
        """
        logger.debug("检查数据是否存在，工作表: {}，条件: {}", sheet_name or self.sheet_name, where)

        if not where:
            logger.debug("查询条件不能为空")
            raise ValueError("查询条件不能为空")

        # 使用select方法查询，限制返回1条记录提高性能
        result = self.select(where=where, limit=1, sheet_name=sheet_name)
        exists = len(result) > 0

        logger.debug("数据存在性检查完成，结果: {}", exists)
        return exists

    def save(self) -> None:
        """保存数据到Excel文件"""
        if self.data is None:
            return

        try:
            # 检查文件锁定（可选）
            if self.check_file_lock and os.path.exists(self.excel_path):
                self._workbook_manager.check_file_lock()

            # 保存到Excel
            self._save_to_excel()
            self._modified = False
            logger.debug("Excel文件保存成功: {}，数据行数: {}，表头字段: {}",
                        self.excel_path, len(self.data) if self.data else 0,
                        len(self.headers) if self.headers else 0)

        except Exception as e:
            logger.debug("保存Excel文件失败: {}", str(e))
            raise IOError(f"保存Excel文件时出错: {str(e)}")

    def close(self):
        """关闭工作簿并释放资源"""
        self._workbook_manager.close_workbook()

    def get_workbook(self):
        """
        获取原生openpyxl工作簿对象

        Returns:
            openpyxl.Workbook: 工作簿对象
        """
        # 如果有未保存的修改，先保存到文件
        if self._modified:
            self.save()

        return self._workbook_manager.get_workbook()

    def get_worksheet(self, sheet_name: Optional[str] = None):
        """
        获取原生openpyxl工作表对象

        Args:
            sheet_name: 工作表名称，None表示使用当前工作表

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: 工作表对象
        """
        # 如果有未保存的修改，先保存到文件
        if self._modified:
            self.save()

        return self._workbook_manager.get_worksheet(sheet_name)

    def sync_from_workbook(self):
        """
        自动保存原生修改并同步回ExcelHandler

        此方法会：
        - 自动保存当前原生工作簿的修改
        - 重新从文件读取数据同步到ExcelHandler
        """
        # 保存原生工作簿修改
        self._workbook_manager.save_workbook()

        # 重新从文件读取数据
        self._read_file()
        self._modified = False

    def _sort_data(self, data: List[Dict[str, Any]], order_by: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        对数据进行排序

        Args:
            data: 要排序的数据列表
            order_by: 排序字典，如 {"age": "DESC", "name": "ASC"}

        Returns:
            List[Dict]: 排序后的数据列表
        """
        if not data or not order_by:
            return data

        # 多字段排序
        sorted_data = data.copy()
        for field, direction in reversed(list(order_by.items())):
            reverse = direction.upper() == 'DESC'
            sorted_data.sort(key=lambda x: (x.get(field) is None, x.get(field)), reverse=reverse)

        return sorted_data

    def _read_file(self) -> None:
        """从文件读取数据"""
        # 检查文件是否存在，不存在则根据auto_create_excel决定是否创建
        if not os.path.exists(self.excel_path):
            if not self.auto_create_excel:
                logger.debug("Excel文件不存在且auto_create_excel=False: {}", self.excel_path)
                raise FileNotFoundError(f"Excel文件不存在: {self.excel_path}")
            self._create_new_file()

        try:
            # 加载工作簿
            workbook = load_workbook(self.excel_path, data_only=True)

            # 确定工作表
            if self.sheet_name is None:
                # 没有指定工作表，使用活动工作表
                ws = workbook.active
                self.sheet_name = ws.title
            elif self.sheet_name not in workbook.sheetnames:
                # 指定的工作表不存在
                if not self.auto_sheet:
                    logger.debug("工作表不存在且auto_sheet=False: {}", self.sheet_name)
                    raise ValueError(f"工作表 '{self.sheet_name}' 不存在")
                # 自动创建工作表
                logger.debug("工作表不存在，正在创建: {}", self.sheet_name)
                ws = workbook.create_sheet(title=self.sheet_name)
                # 保存工作簿以确保新工作表被写入文件
                workbook.save(self.excel_path)
            else:
                # 工作表存在，直接使用
                ws = workbook[self.sheet_name]

            # 获取所有行
            rows = list(ws.rows)
            if not rows:
                self.headers = []
                self.data = []
                workbook.close()
                return

            # 处理表头
            header_row_idx = self.header_row - 1
            if header_row_idx < 0 or header_row_idx >= len(rows):
                logger.debug("无效的header_row值: {}", self.header_row)
                raise ValueError(f"无效的header_row值: {self.header_row}")

            self._process_headers(rows[header_row_idx])
            self._read_data_rows(rows, header_row_idx)

            workbook.close()

        except (FileNotFoundError, ValueError) as e:
            # 重新抛出这些特定的异常
            raise e
        except Exception as e:
            logger.debug("读取Excel文件失败: {}", str(e))
            self.headers = []
            self.data = []

    def _create_new_file(self) -> None:
        """创建新Excel文件"""
        logger.debug("创建新Excel文件: {}", self.excel_path)
        os.makedirs(os.path.dirname(os.path.abspath(self.excel_path)), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        if self.sheet_name:
            ws.title = self.sheet_name

        wb.save(self.excel_path)
        wb.close()

    def _process_headers(self, header_cells) -> None:
        """处理表头单元格"""
        self.headers = []
        for cell in header_cells:
            header_text = ""
            if cell.value is not None:
                header_text = str(cell.value).strip()

            if not header_text:
                header_text = f"列{get_column_letter(cell.column)}"

            self.headers.append(header_text)

        # 处理重复列名
        if len(self.headers) != len(set(self.headers)):
            seen = {}
            for i, header in enumerate(self.headers):
                if header in seen:
                    seen[header] += 1
                    self.headers[i] = f"{header}_{seen[header]}"
                else:
                    seen[header] = 1

    def _read_data_rows(self, rows, header_row_idx: int) -> None:
        """读取数据行"""
        self.data = []
        for row_idx in range(header_row_idx + 1, len(rows)):
            row_data = {}
            data_row = rows[row_idx]

            for col_idx, cell in enumerate(data_row):
                if col_idx < len(self.headers):
                    value = cell.value
                    if isinstance(value, str):
                        value = value.strip()
                    row_data[self.headers[col_idx]] = value

            # 添加index字段 - 对应Excel中的实际行号
            row_data['index'] = self.header_row + row_idx - header_row_idx

            # 检查是否为空行
            if any(v for k, v in row_data.items() if k != 'index' and v not in (None, "")):
                self.data.append(row_data)

    def _save_to_excel(self) -> None:
        """保存逻辑"""
        file_exists = os.path.exists(self.excel_path)

        if file_exists:
            wb = load_workbook(self.excel_path)
            # 读取标题行前的内容
            pre_header_data = {}
            if self.sheet_name in wb.sheetnames and self.header_row > 1:
                ws_old = wb[self.sheet_name]
                for row_idx in range(1, self.header_row):
                    for col_idx in range(1, 50):  # 检查前50列
                        cell = ws_old.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            pre_header_data.setdefault(row_idx, {})[col_idx] = cell.value

            # 删除旧sheet，创建新sheet
            if self.sheet_name in wb.sheetnames:
                del wb[self.sheet_name]
            ws = wb.create_sheet(title=self.sheet_name)

            # 恢复标题行前的内容
            for row_idx, row_data in pre_header_data.items():
                for col_idx, value in row_data.items():
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            # 创建新文件
            os.makedirs(os.path.dirname(os.path.abspath(self.excel_path)), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name or 'Sheet1'

        # 写入表头和数据
        headers_to_write = [h for h in self.headers if h != 'index']
        for col_idx, header in enumerate(headers_to_write, 1):
            ws.cell(row=self.header_row, column=col_idx, value=header)

        # 写入数据行
        for row_idx, row_data in enumerate(self.data, self.header_row + 1):
            for col_idx, header in enumerate(headers_to_write, 1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(self.excel_path)
        wb.close()

    def _create_backup(self) -> str:
        """
        创建Excel文件的备份

        Returns:
            str: 备份文件的完整路径
        """
        try:
            # 获取源文件信息
            source_dir = os.path.dirname(os.path.abspath(self.excel_path))
            source_filename = os.path.basename(self.excel_path)
            filename_without_ext, file_ext = os.path.splitext(source_filename)

            # 创建备份文件夹
            backup_dir = os.path.join(source_dir, "备份")
            os.makedirs(backup_dir, exist_ok=True)

            # 生成时间戳
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 生成备份文件名
            backup_filename = f"{filename_without_ext}_{timestamp}{file_ext}"
            backup_path = os.path.join(backup_dir, backup_filename)

            # 复制文件
            shutil.copy2(self.excel_path, backup_path)

            logger.debug("Excel文件已备份至: {}", backup_path)
            return backup_path

        except Exception as e:
            logger.debug("备份Excel文件失败: {}", str(e))
            raise Exception(f"备份操作失败: {str(e)}")
