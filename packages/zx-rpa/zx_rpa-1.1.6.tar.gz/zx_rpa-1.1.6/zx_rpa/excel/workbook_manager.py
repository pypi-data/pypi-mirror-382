"""
Excel工作簿管理器
处理原生openpyxl对象的操作和同步
"""

import os
from typing import Optional, Any
from loguru import logger

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.debug("openpyxl库未安装，请运行: pip install openpyxl")
    raise ImportError("openpyxl库未安装，Excel功能不可用")


class WorkbookManager:
    """Excel工作簿管理器，处理原生openpyxl对象"""

    def __init__(self, excel_path: str, sheet_name: Optional[str] = None):
        """
        初始化工作簿管理器
        
        Args:
            excel_path: Excel文件路径
            sheet_name: 工作表名称
        """
        logger.debug("初始化工作簿管理器，文件: {}", excel_path)
        
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.workbook = None
        self._ws = None

    def get_workbook(self):
        """
        获取原生openpyxl工作簿对象
        
        Returns:
            openpyxl.Workbook: 工作簿对象
        """
        logger.debug("获取原生工作簿对象")
        
        # 关闭之前的workbook避免冲突
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
        
        # 重新加载文件以获取最新状态的workbook
        if not os.path.exists(self.excel_path):
            logger.debug("Excel文件不存在: {}", self.excel_path)
            raise FileNotFoundError(f"Excel文件不存在: {self.excel_path}")
        
        self.workbook = load_workbook(self.excel_path)
        
        # 设置工作表
        if self.sheet_name and self.sheet_name in self.workbook.sheetnames:
            self._ws = self.workbook[self.sheet_name]
        else:
            self._ws = self.workbook.active
            self.sheet_name = self._ws.title
        
        return self.workbook

    def get_worksheet(self, sheet_name: Optional[str] = None):
        """
        获取原生openpyxl工作表对象
        
        Args:
            sheet_name: 工作表名称，None表示使用当前工作表
            
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: 工作表对象
        """
        logger.debug("获取原生工作表对象，工作表: {}", sheet_name or self.sheet_name)
        
        # 确保workbook是最新的
        if self.workbook is None:
            self.get_workbook()
        
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                logger.debug("工作表不存在: {}", sheet_name)
                raise ValueError(f"工作表不存在: {sheet_name}")
            return self.workbook[sheet_name]
        
        return self._ws

    def save_workbook(self):
        """
        保存原生工作簿的修改
        """
        logger.debug("保存原生工作簿修改")
        
        if self.workbook:
            try:
                self.workbook.save(self.excel_path)
                logger.debug("原生工作簿修改已保存")
            except Exception as e:
                logger.debug("保存原生修改失败: {}", str(e))
                raise Exception(f"保存原生修改失败: {str(e)}")
        else:
            logger.debug("没有工作簿对象需要保存")

    def close_workbook(self):
        """关闭工作簿并释放资源"""
        if self.workbook:
            try:
                self.workbook.close()
                logger.debug("工作簿已关闭")
            except:
                pass
            self.workbook = None
            self._ws = None

    def create_sheet(self, sheet_name: str, index: Optional[int] = None):
        """
        创建新工作表
        
        Args:
            sheet_name: 工作表名称
            index: 插入位置，None表示添加到末尾
            
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: 新创建的工作表对象
        """
        logger.debug("创建新工作表: {}", sheet_name)
        
        if self.workbook is None:
            self.get_workbook()
        
        if sheet_name in self.workbook.sheetnames:
            logger.debug("工作表已存在: {}", sheet_name)
            raise ValueError(f"工作表已存在: {sheet_name}")
        
        if index is not None:
            ws = self.workbook.create_sheet(title=sheet_name, index=index)
        else:
            ws = self.workbook.create_sheet(title=sheet_name)
        
        logger.debug("工作表创建成功: {}", sheet_name)
        return ws

    def delete_sheet(self, sheet_name: str):
        """
        删除工作表
        
        Args:
            sheet_name: 要删除的工作表名称
        """
        logger.debug("删除工作表: {}", sheet_name)
        
        if self.workbook is None:
            self.get_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            logger.debug("工作表不存在: {}", sheet_name)
            raise ValueError(f"工作表不存在: {sheet_name}")
        
        if len(self.workbook.sheetnames) <= 1:
            logger.debug("不能删除最后一个工作表")
            raise ValueError("不能删除最后一个工作表")
        
        del self.workbook[sheet_name]
        logger.debug("工作表删除成功: {}", sheet_name)

    def get_sheet_names(self) -> list:
        """
        获取所有工作表名称
        
        Returns:
            list: 工作表名称列表
        """
        if self.workbook is None:
            self.get_workbook()
        
        return self.workbook.sheetnames

    def copy_sheet(self, source_sheet: str, target_sheet: str):
        """
        复制工作表
        
        Args:
            source_sheet: 源工作表名称
            target_sheet: 目标工作表名称
        """
        logger.debug("复制工作表: {} -> {}", source_sheet, target_sheet)
        
        if self.workbook is None:
            self.get_workbook()
        
        if source_sheet not in self.workbook.sheetnames:
            logger.debug("源工作表不存在: {}", source_sheet)
            raise ValueError(f"源工作表不存在: {source_sheet}")
        
        if target_sheet in self.workbook.sheetnames:
            logger.debug("目标工作表已存在: {}", target_sheet)
            raise ValueError(f"目标工作表已存在: {target_sheet}")
        
        source_ws = self.workbook[source_sheet]
        target_ws = self.workbook.copy_worksheet(source_ws)
        target_ws.title = target_sheet
        
        logger.debug("工作表复制成功: {} -> {}", source_sheet, target_sheet)

    def check_file_lock(self) -> bool:
        """检查文件是否被占用"""
        if not os.path.exists(self.excel_path):
            logger.debug("文件不存在: {}", self.excel_path)
            raise FileNotFoundError(f"文件不存在: {self.excel_path}")
        
        import sys
        if sys.platform == "win32":
            try:
                import msvcrt
                with open(self.excel_path, 'r+b') as f:
                    try:
                        msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                        msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
                        return True
                    except OSError:
                        logger.debug("文件被占用，请关闭文件: {}", self.excel_path)
                        raise Exception("文件被占用，请关闭文件")
            except Exception:
                logger.debug("文件被占用，请关闭文件: {}", self.excel_path)
                raise Exception("文件被占用，请关闭文件")
        return True
