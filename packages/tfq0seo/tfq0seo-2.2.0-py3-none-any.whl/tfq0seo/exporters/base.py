"""
Export manager for different output formats
"""
import json
import csv
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import jinja2
from datetime import datetime
import logging
import os
import xml.etree.ElementTree as ET
from xml.dom import minidom
import re
from html import escape

# Fix encoding issues on Windows
import sys
import locale
if sys.platform == 'win32':
    # Set UTF-8 as default encoding
    import codecs
    # Set environment variable for UTF-8
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    # Configure stdout/stderr for UTF-8
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, errors='replace')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, errors='replace')

logger = logging.getLogger(__name__)

class ExportManager:
    """Manager for exporting SEO analysis results"""
    
    def __init__(self):
        self.template_dir = Path(__file__).parent.parent / 'templates'
        self.jinja_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(str(self.template_dir)),
            autoescape=jinja2.select_autoescape(['html', 'xml']),
            enable_async=False,
            finalize=lambda x: x if x is not None else ''
        )
        # Ensure template directory exists
        self.template_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_default_template()
    
    def _validate_output_path(self, path: str) -> Path:
        """Validate and prepare output path"""
        output_path = Path(path)
        
        # Create directory if it doesn't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Check write permissions
        try:
            output_path.touch(exist_ok=True)
        except PermissionError:
            raise PermissionError(f"Cannot write to {path}")
        
        return output_path
    
    def _sanitize_data(self, data: Any) -> Any:
        """Sanitize data for safe export"""
        if isinstance(data, dict):
            return {k: self._sanitize_data(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._sanitize_data(item) for item in data]
        elif isinstance(data, str):
            # Remove control characters
            return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', data)
        else:
            return data
    
    def _handle_encoding_safe(self, data: Any) -> Any:
        """Handle encoding issues by converting problematic characters"""
        if isinstance(data, dict):
            return {k: self._handle_encoding_safe(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._handle_encoding_safe(item) for item in data]
        elif isinstance(data, str):
            try:
                # Try to encode/decode to handle problematic characters
                return data.encode('utf-8', errors='replace').decode('utf-8')
            except:
                # Fallback to ASCII-safe representation
                return data.encode('ascii', errors='xmlcharrefreplace').decode('ascii')
        else:
            return data
    
    def export_json(self, data: Dict[str, Any], output_path: str, pretty: bool = True) -> None:
        """Export results as JSON with optional streaming for large datasets"""
        try:
            path = self._validate_output_path(output_path)
            sanitized_data = self._sanitize_data(data)
            
            # Check if data is large (>100MB estimated)
            if self._estimate_size(sanitized_data) > 100 * 1024 * 1024:
                # Stream large datasets
                self._export_json_streaming(sanitized_data, path)
            else:
                # Regular export
                with open(path, 'w', encoding='utf-8') as f:
                    if pretty:
                        json.dump(sanitized_data, f, indent=2, ensure_ascii=False)
                    else:
                        json.dump(sanitized_data, f, ensure_ascii=False)
            
            logger.info(f"JSON export completed: {path}")
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            raise
    
    def _export_json_streaming(self, data: Dict[str, Any], path: Path) -> None:
        """Stream large JSON datasets"""
        with open(path, 'w', encoding='utf-8') as f:
            f.write('{\n')
            
            items = list(data.items())
            for i, (key, value) in enumerate(items):
                f.write(f'  "{key}": ')
                
                if isinstance(value, list) and len(value) > 1000:
                    # Stream large lists
                    f.write('[\n')
                    for j, item in enumerate(value):
                        f.write('    ')
                        json.dump(item, f, ensure_ascii=False)
                        if j < len(value) - 1:
                            f.write(',')
                        f.write('\n')
                    f.write('  ]')
                else:
                    json.dump(value, f, indent=2, ensure_ascii=False)
                
                if i < len(items) - 1:
                    f.write(',')
                f.write('\n')
            
            f.write('}\n')
    
    def export_csv(self, data: Dict[str, Any], output_path: str) -> None:
        """Export results as CSV with improved handling"""
        try:
            path = self._validate_output_path(output_path)
            
            # Check if this is single page analysis or crawl data
            is_single_page = 'pages' not in data and 'url' in data
            
            if is_single_page:
                pages = [data]
            else:
                pages = data.get('pages', [])
            
            if not pages:
                logger.warning("No pages to export to CSV")
                return
            
            # Flatten page data for CSV
            rows = []
            for page in pages:
                row = self._flatten_page_data(page)
                rows.append(row)
            
            # Write CSV with all fields
            if rows:
                fieldnames = self._get_all_fieldnames(rows)
                
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:  # UTF-8 with BOM for Excel
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
            
            logger.info(f"CSV export completed: {path}")
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            raise
    
    def _flatten_page_data(self, page: Dict) -> Dict:
        """Flatten nested page data for CSV export"""
        flat = {
            'url': page.get('url', ''),
            'final_url': page.get('final_url', ''),
            'title': page.get('title', ''),
            'title_length': page.get('meta_tags', {}).get('title_length', 0),
            'description': page.get('meta_description', ''),
            'description_length': page.get('meta_tags', {}).get('description_length', 0),
            'status_code': page.get('status_code', 0),
            'load_time': page.get('load_time', 0),
            'word_count': page.get('content', {}).get('word_count', 0),
            'readability_score': page.get('content', {}).get('readability_scores', {}).get('flesch_reading_ease', 0),
            'h1_count': page.get('meta_tags', {}).get('h1_count', 0),
            'h2_count': page.get('meta_tags', {}).get('heading_structure', {}).get('h2', 0),
            'internal_links': page.get('links', {}).get('internal_links', 0),
            'external_links': page.get('links', {}).get('external_links', 0),
            'broken_links': page.get('links', {}).get('broken_links', 0),
            'images_total': page.get('content', {}).get('images', {}).get('total', 0),
            'images_without_alt': page.get('content', {}).get('images', {}).get('without_alt', 0),
            'score': page.get('score', 0),
            'score_meta': page.get('score_breakdown', {}).get('meta_tags', 0),
            'score_content': page.get('score_breakdown', {}).get('content', 0),
            'score_technical': page.get('score_breakdown', {}).get('technical', 0),
            'score_performance': page.get('score_breakdown', {}).get('performance', 0),
            'score_links': page.get('score_breakdown', {}).get('links', 0),
            'issues_critical': sum(1 for i in page.get('issues', []) if i.get('severity') == 'critical'),
            'issues_warning': sum(1 for i in page.get('issues', []) if i.get('severity') == 'warning'),
            'issues_notice': sum(1 for i in page.get('issues', []) if i.get('severity') == 'notice'),
            'https': page.get('technical', {}).get('https', False),
            'mobile_friendly': page.get('technical', {}).get('mobile_friendly', False),
            'canonical_url': page.get('technical', {}).get('canonical_url', ''),
            'has_structured_data': len(page.get('technical', {}).get('structured_data', [])) > 0,
            'core_web_vitals_fcp': page.get('performance', {}).get('core_web_vitals', {}).get('FCP', 0),
            'core_web_vitals_lcp': page.get('performance', {}).get('core_web_vitals', {}).get('LCP', 0),
            'core_web_vitals_cls': page.get('performance', {}).get('core_web_vitals', {}).get('CLS', 0),
        }
        
        # Add top 3 issues
        issues = page.get('issues', [])
        for i in range(3):
            if i < len(issues):
                flat[f'issue_{i+1}_type'] = issues[i].get('type', '')
                flat[f'issue_{i+1}_severity'] = issues[i].get('severity', '')
                flat[f'issue_{i+1}_message'] = issues[i].get('message', '')[:100]  # Truncate long messages
        
        return flat
    
    def _get_all_fieldnames(self, rows: List[Dict]) -> List[str]:
        """Get all unique fieldnames from rows"""
        fieldnames = []
        seen = set()
        
        # Define preferred order
        preferred_order = [
            'url', 'final_url', 'title', 'title_length', 'description', 'description_length',
            'status_code', 'load_time', 'score', 'score_meta', 'score_content', 'score_technical',
            'score_performance', 'score_links', 'word_count', 'readability_score',
            'h1_count', 'h2_count', 'internal_links', 'external_links', 'broken_links',
            'images_total', 'images_without_alt', 'issues_critical', 'issues_warning', 'issues_notice',
            'https', 'mobile_friendly', 'canonical_url', 'has_structured_data'
        ]
        
        # Add preferred fields first
        for field in preferred_order:
            if any(field in row for row in rows):
                fieldnames.append(field)
                seen.add(field)
        
        # Add remaining fields
        for row in rows:
            for field in row:
                if field not in seen:
                    fieldnames.append(field)
                    seen.add(field)
        
        return fieldnames
    
    def export_xlsx(self, data: Dict[str, Any], output_path: str) -> None:
        """Export results as Excel file with multiple sheets and formatting"""
        try:
            path = self._validate_output_path(output_path)
            
            # Check if this is single page analysis or crawl data
            is_single_page = 'pages' not in data and 'url' in data
            
            if is_single_page:
                # Convert single page data to crawl format
                page_data = data
                data = {
                    'config': {'url': page_data.get('url', 'Unknown')},
                    'pages': [page_data],
                    'summary': {
                        'total_pages': 1,
                        'average_seo_score': page_data.get('score', 0),
                        'average_load_time': page_data.get('load_time', 0),
                        'average_word_count': page_data.get('content', {}).get('word_count', 0)
                    },
                    'issues': {
                        'critical': sum(1 for i in page_data.get('issues', []) if i.get('severity') == 'critical'),
                        'warnings': sum(1 for i in page_data.get('issues', []) if i.get('severity') == 'warning'),
                        'notices': sum(1 for i in page_data.get('issues', []) if i.get('severity') == 'notice')
                    }
                }
            
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = self._prepare_summary_sheet(data)
                if summary_data:
                    summary_df = pd.DataFrame([summary_data])
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Format summary sheet
                    self._format_summary_sheet(writer, 'Summary')
                
                # Pages sheet
                pages_df = self._prepare_pages_dataframe(data.get('pages', []))
                if not pages_df.empty:
                    pages_df.to_excel(writer, sheet_name='Pages', index=False)
                    self._format_pages_sheet(writer, 'Pages')
                
                # Issues sheet
                issues_df = self._prepare_issues_dataframe(data.get('pages', []))
                if not issues_df.empty:
                    issues_df.to_excel(writer, sheet_name='Issues', index=False)
                    self._format_issues_sheet(writer, 'Issues')
                
                # Top issues sheet
                top_issues = data.get('summary', {}).get('top_issues', [])
                if top_issues:
                    top_issues_df = pd.DataFrame(top_issues)
                    if 'example_pages' in top_issues_df.columns:
                        top_issues_df['example_pages'] = top_issues_df['example_pages'].apply(
                            lambda x: ', '.join(x[:3]) if isinstance(x, list) else x
                        )
                    top_issues_df.to_excel(writer, sheet_name='Top Issues', index=False)
                    self._format_issues_sheet(writer, 'Top Issues')
                
                # Performance metrics sheet
                perf_df = self._prepare_performance_dataframe(data.get('pages', []))
                if not perf_df.empty:
                    perf_df.to_excel(writer, sheet_name='Performance', index=False)
                    self._format_performance_sheet(writer, 'Performance')
                
                # Add charts if data is available
                if len(data.get('pages', [])) > 1:
                    self._add_charts(writer, data)
            
            logger.info(f"Excel export completed: {path}")
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise
    
    def _format_summary_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Format the summary sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:  # First row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_pages_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Format the pages sheet with conditional formatting"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply conditional formatting for scores
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        # Find score column
        for col in worksheet.columns:
            if col[0].value == 'SEO Score':
                col_letter = col[0].column_letter
                
                # Good scores (80+) - green
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='greaterThanOrEqual', formula=['80'],
                              fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
                )
                
                # Medium scores (60-79) - yellow
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='between', formula=['60', '79'],
                              fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
                )
                
                # Poor scores (<60) - red
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='lessThan', formula=['60'],
                              fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
                )
    
    def _format_issues_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Format the issues sheet"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply conditional formatting for severity
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        # Find severity column
        for col in worksheet.columns:
            if col[0].value == 'Severity':
                col_letter = col[0].column_letter
                
                # Critical - red
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='equal', formula=['"critical"'],
                              fill=PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'))
                )
                
                # Warning - orange
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='equal', formula=['"warning"'],
                              fill=PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'))
                )
                
                # Notice - yellow
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='equal', formula=['"notice"'],
                              fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
                )
    
    def _format_performance_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Format the performance sheet"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply conditional formatting for load times
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        # Find load time column
        for col in worksheet.columns:
            if col[0].value == 'Load Time (s)':
                col_letter = col[0].column_letter
                
                # Fast (<1.5s) - green
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='lessThan', formula=['1.5'],
                              fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
                )
                
                # Moderate (1.5-3s) - yellow
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='between', formula=['1.5', '3'],
                              fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
                )
                
                # Slow (>3s) - red
                worksheet.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}1000',
                    CellIsRule(operator='greaterThan', formula=['3'],
                              fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
                )
    
    def _add_charts(self, writer: pd.ExcelWriter, data: Dict[str, Any]) -> None:
        """Add charts to Excel workbook"""
        from openpyxl.chart import BarChart, PieChart, Reference
        
        workbook = writer.book
        
        # Create Charts sheet
        charts_sheet = workbook.create_sheet('Charts')
        
        # Score distribution chart
        pages = data.get('pages', [])
        if pages:
            score_bins = {'0-20': 0, '21-40': 0, '41-60': 0, '61-80': 0, '81-100': 0}
            for page in pages:
                score = page.get('score', 0)
                if score <= 20:
                    score_bins['0-20'] += 1
                elif score <= 40:
                    score_bins['21-40'] += 1
                elif score <= 60:
                    score_bins['41-60'] += 1
                elif score <= 80:
                    score_bins['61-80'] += 1
                else:
                    score_bins['81-100'] += 1
            
            # Write data
            row = 1
            charts_sheet.cell(row=row, column=1, value='Score Range')
            charts_sheet.cell(row=row, column=2, value='Count')
            
            for label, count in score_bins.items():
                row += 1
                charts_sheet.cell(row=row, column=1, value=label)
                charts_sheet.cell(row=row, column=2, value=count)
            
            # Create bar chart
            chart = BarChart()
            chart.title = "SEO Score Distribution"
            chart.x_axis.title = "Score Range"
            chart.y_axis.title = "Number of Pages"
            
            data = Reference(charts_sheet, min_col=2, min_row=1, max_row=row, max_col=2)
            categories = Reference(charts_sheet, min_col=1, min_row=2, max_row=row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            charts_sheet.add_chart(chart, "D1")
    
    def export_html(self, data: Dict[str, Any], output_path: str) -> None:
        """Export results as HTML report"""
        try:
            path = self._validate_output_path(output_path)
            
            # Check if this is single page analysis or crawl data
            is_single_page = 'pages' not in data and 'url' in data
            
            if is_single_page:
                # Convert single page data to crawl format
                page_data = data
                data = self._convert_single_page_to_crawl_format(page_data)
            
            # Determine which template to use based on whether this is an enhanced report
            is_enhanced_report = 'executive_summary' in data or 'action_plan' in data or data.get('report_version') == '2.0'
            template_name = 'enhanced_report.html' if is_enhanced_report else 'report.html'
            
            # Load template
            try:
                template = self.jinja_env.get_template(template_name)
            except Exception as e:
                logger.warning(f"Failed to load {template_name}: {e}. Falling back to report.html")
                template = self.jinja_env.get_template('report.html')
            
            # Prepare data - ensure all strings are properly encoded
            report_data = {
                'title': f"SEO Report - {data.get('config', {}).get('url', 'Website')}",
                'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'data': self._sanitize_data(data),
                'summary': data.get('summary', {}),
                'pages': data.get('pages', []),
                'issues': data.get('issues', {}),
                'config': data.get('config', {}),
                'is_single_page': is_single_page
            }
            
            # Add enhanced report data if present
            if is_enhanced_report:
                report_data.update({
                    'executive_summary': data.get('executive_summary', {}),
                    'action_plan': data.get('action_plan', {}),
                    'implementation_timeline': data.get('implementation_timeline', {}),
                    'expected_results': data.get('expected_results', {}),
                    'resources_needed': data.get('resources_needed', {}),
                    'analysis_data': data.get('analysis_data', data)
                })
            
            # Add charts data
            report_data['charts_data'] = self._prepare_charts_data(data)
            
            # Render template with proper encoding handling
            try:
                html_content = template.render(**report_data)
            except (UnicodeDecodeError, UnicodeEncodeError) as e:
                # Handle encoding issues by converting problematic characters
                logger.warning(f"Encoding issue encountered: {e}. Applying fallback encoding.")
                report_data = self._handle_encoding_safe(report_data)
                html_content = template.render(**report_data)
            
            # Write file - ensure proper UTF-8 encoding for Windows
            with open(path, 'w', encoding='utf-8', errors='replace') as f:
                f.write(html_content)
            
            logger.info(f"HTML export completed: {path}")
            
        except Exception as e:
            logger.error(f"HTML export failed: {e}")
            raise
    
    def export_xml_sitemap(self, data: Dict[str, Any], output_path: str) -> None:
        """Export URLs as XML sitemap"""
        try:
            path = self._validate_output_path(output_path)
            
            # Create root element
            urlset = ET.Element('urlset')
            urlset.set('xmlns', 'http://www.sitemaps.org/schemas/sitemap/0.9')
            
            # Add URLs
            pages = data.get('pages', [])
            for page in pages:
                if page.get('status_code') == 200:
                    url_elem = ET.SubElement(urlset, 'url')
                    
                    # Location (required)
                    loc = ET.SubElement(url_elem, 'loc')
                    loc.text = page.get('url', '')
                    
                    # Last modified (optional)
                    lastmod = ET.SubElement(url_elem, 'lastmod')
                    lastmod.text = datetime.now().strftime('%Y-%m-%d')
                    
                    # Priority based on SEO score
                    priority = ET.SubElement(url_elem, 'priority')
                    score = page.get('score', 50)
                    priority.text = str(round(score / 100, 1))
            
            # Pretty print XML
            xml_string = minidom.parseString(ET.tostring(urlset)).toprettyxml(indent='  ')
            
            # Write file
            with open(path, 'w', encoding='utf-8') as f:
                f.write(xml_string)
            
            logger.info(f"XML sitemap export completed: {path}")
            
        except Exception as e:
            logger.error(f"XML sitemap export failed: {e}")
            raise
    
    def _prepare_summary_sheet(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare summary data for Excel"""
        summary = data.get('summary', {})
        config = data.get('config', {})
        issues = data.get('issues', {})
        
        # Count issues from pages if not in summary
        if not issues and data.get('pages'):
            issues = self._count_issues_from_pages(data['pages'])
        
        return {
            'Website': config.get('url', ''),
            'Analysis Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Total Pages': summary.get('total_pages', len(data.get('pages', []))),
            'Valid Pages': summary.get('valid_pages', summary.get('total_pages', 0)),
            'Average Load Time (s)': round(summary.get('average_load_time', 0), 2),
            'Average Word Count': summary.get('average_word_count', 0),
            'Average SEO Score': round(summary.get('average_seo_score', 0), 1),
            'Critical Issues': issues.get('critical', 0),
            'Warnings': issues.get('warnings', issues.get('warning', 0)),
            'Notices': issues.get('notices', issues.get('notice', 0)),
            'Pages Missing Title': summary.get('pages_missing_title', 0),
            'Pages Missing Description': summary.get('pages_missing_description', 0),
            'Slow Loading Pages': summary.get('slow_loading_pages', 0),
            'Thin Content Pages': summary.get('thin_content_pages', 0),
            'HTTPS Pages': summary.get('https_pages', 0),
            'Mobile Friendly Pages': summary.get('mobile_friendly_pages', 0)
        }
    
    def _prepare_pages_dataframe(self, pages: List[Dict]) -> pd.DataFrame:
        """Prepare pages data for Excel"""
        if not pages:
            return pd.DataFrame()
        
        rows = []
        for page in pages:
            if page.get('error'):
                # Skip error pages or include minimal info
                continue
            
            row = {
                'URL': page.get('url', ''),
                'Title': page.get('title', '')[:100],  # Truncate long titles
                'Title Length': page.get('meta_tags', {}).get('title_length', 0),
                'Description': page.get('meta_description', '')[:160],  # Truncate
                'Description Length': page.get('meta_tags', {}).get('description_length', 0),
                'Status Code': page.get('status_code', 0),
                'Load Time (s)': round(page.get('load_time', 0), 2),
                'Word Count': page.get('content', {}).get('word_count', 0),
                'Readability Score': round(page.get('content', {}).get('readability_scores', {}).get('flesch_reading_ease', 0), 1),
                'H1 Count': page.get('meta_tags', {}).get('h1_count', 0),
                'Internal Links': page.get('links', {}).get('internal_links', 0),
                'External Links': page.get('links', {}).get('external_links', 0),
                'Broken Links': page.get('links', {}).get('broken_links', 0),
                'Images': page.get('content', {}).get('images', {}).get('total', 0),
                'Images Without Alt': page.get('content', {}).get('images', {}).get('without_alt', 0),
                'SEO Score': round(page.get('score', 0), 1),
                'Issues': len(page.get('issues', [])),
                'HTTPS': 'Yes' if page.get('technical', {}).get('https', False) else 'No',
                'Mobile Friendly': 'Yes' if page.get('technical', {}).get('mobile_friendly', False) else 'No'
            }
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _prepare_issues_dataframe(self, pages: List[Dict]) -> pd.DataFrame:
        """Prepare issues data for Excel"""
        all_issues = []
        
        for page in pages:
            url = page.get('url', '')
            for issue in page.get('issues', []):
                all_issues.append({
                    'URL': url,
                    'Type': issue.get('type', ''),
                    'Severity': issue.get('severity', ''),
                    'Message': issue.get('message', ''),
                    'Category': self._get_issue_category(issue.get('type', ''))
                })
        
        df = pd.DataFrame(all_issues)
        
        # Sort by severity
        severity_order = {'critical': 0, 'warning': 1, 'notice': 2}
        if not df.empty:
            df['severity_order'] = df['Severity'].map(severity_order)
            df = df.sort_values(['severity_order', 'Type', 'URL'])
            df = df.drop('severity_order', axis=1)
        
        return df
    
    def _prepare_performance_dataframe(self, pages: List[Dict]) -> pd.DataFrame:
        """Prepare performance data for Excel"""
        if not pages:
            return pd.DataFrame()
        
        rows = []
        for page in pages:
            if page.get('error'):
                continue
            
            perf = page.get('performance', {})
            cwv = perf.get('core_web_vitals', {})
            
            row = {
                'URL': page.get('url', ''),
                'Load Time (s)': round(page.get('load_time', 0), 2),
                'Performance Score': round(perf.get('performance_score', 0), 1),
                'FCP (s)': round(cwv.get('FCP', 0), 2),
                'LCP (s)': round(cwv.get('LCP', 0), 2),
                'CLS': round(cwv.get('CLS', 0), 3),
                'FID (ms)': round(cwv.get('FID', 0), 0),
                'TTI (s)': round(cwv.get('TTI', 0), 2),
                'Total Resources': perf.get('resource_summary', {}).get('total_resources', 0),
                'Total Size (KB)': round(perf.get('resource_summary', {}).get('total_size', 0) / 1024, 1),
                'Compressed': 'Yes' if page.get('technical', {}).get('compression') else 'No',
                'Cache Control': page.get('technical', {}).get('cache_control', 'None')[:50]
            }
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _get_issue_category(self, issue_type: str) -> str:
        """Categorize issue types"""
        categories = {
            'meta': ['missing_title', 'short_title', 'long_title', 'missing_description', 
                    'short_description', 'long_description', 'duplicate_title', 'duplicate_description'],
            'content': ['thin_content', 'low_readability', 'missing_h1', 'multiple_h1', 
                       'keyword_stuffing', 'missing_keywords'],
            'technical': ['no_https', 'no_compression', 'no_viewport', 'missing_canonical',
                         'invalid_canonical', 'noindex_directive'],
            'performance': ['slow_load_time', 'large_page_size', 'too_many_resources',
                           'render_blocking_resources', 'no_caching'],
            'links': ['broken_links', 'no_internal_links', 'too_many_external_links',
                     'missing_anchor_text', 'nofollow_all_links'],
            'images': ['missing_alt_text', 'large_images', 'wrong_format', 'missing_dimensions']
        }
        
        for category, types in categories.items():
            if any(t in issue_type for t in types):
                return category
        
        return 'other'
    
    def _convert_single_page_to_crawl_format(self, page_data: Dict) -> Dict:
        """Convert single page data to crawl format for consistent handling"""
        issues = self._count_issues_from_pages([page_data])
        
        return {
            'config': {'url': page_data.get('url', 'Unknown')},
            'pages': [page_data],
            'summary': {
                'total_pages': 1,
                'valid_pages': 1 if page_data.get('status_code') == 200 else 0,
                'average_seo_score': page_data.get('score', 0),
                'average_load_time': page_data.get('load_time', 0),
                'average_word_count': page_data.get('content', {}).get('word_count', 0),
                'pages_missing_title': 1 if not page_data.get('title') else 0,
                'pages_missing_description': 1 if not page_data.get('meta_description') else 0,
                'https_pages': 1 if page_data.get('technical', {}).get('https') else 0,
                'mobile_friendly_pages': 1 if page_data.get('technical', {}).get('mobile_friendly') else 0,
                'top_issues': self._extract_top_issues([page_data])
            },
            'issues': issues
        }
    
    def _count_issues_from_pages(self, pages: List[Dict]) -> Dict[str, int]:
        """Count issues by severity from pages"""
        counts = {'critical': 0, 'warning': 0, 'notice': 0}
        
        for page in pages:
            for issue in page.get('issues', []):
                severity = issue.get('severity', 'notice')
                if severity in counts:
                    counts[severity] += 1
        
        return counts
    
    def _extract_top_issues(self, pages: List[Dict]) -> List[Dict]:
        """Extract top issues from pages"""
        issue_counts = {}
        
        for page in pages:
            for issue in page.get('issues', []):
                key = (issue.get('type', ''), issue.get('message', ''), issue.get('severity', ''))
                if key not in issue_counts:
                    issue_counts[key] = 0
                issue_counts[key] += 1
        
        # Convert to list and sort
        top_issues = []
        for (issue_type, message, severity), count in issue_counts.items():
            top_issues.append({
                'type': issue_type,
                'message': message,
                'severity': severity,
                'count': count
            })
        
        # Sort by severity and count
        severity_weight = {'critical': 3, 'warning': 2, 'notice': 1}
        top_issues.sort(key=lambda x: (severity_weight.get(x['severity'], 0), x['count']), reverse=True)
        
        return top_issues[:10]
    
    def _prepare_charts_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for JavaScript charts in HTML"""
        charts_data = {}
        pages = data.get('pages', [])
        
        if pages:
            # Score distribution
            score_bins = [0, 0, 0, 0, 0]  # 0-20, 21-40, 41-60, 61-80, 81-100
            for page in pages:
                score = page.get('score', 0)
                if score <= 20:
                    score_bins[0] += 1
                elif score <= 40:
                    score_bins[1] += 1
                elif score <= 60:
                    score_bins[2] += 1
                elif score <= 80:
                    score_bins[3] += 1
                else:
                    score_bins[4] += 1
            
            charts_data['score_distribution'] = {
                'labels': ['0-20', '21-40', '41-60', '61-80', '81-100'],
                'data': score_bins
            }
            
            # Performance distribution
            perf_bins = [0, 0, 0]  # Fast, Moderate, Slow
            for page in pages:
                load_time = page.get('load_time', 0)
                if load_time < 1.5:
                    perf_bins[0] += 1
                elif load_time < 3:
                    perf_bins[1] += 1
                else:
                    perf_bins[2] += 1
            
            charts_data['performance_distribution'] = {
                'labels': ['Fast (<1.5s)', 'Moderate (1.5-3s)', 'Slow (>3s)'],
                'data': perf_bins
            }
            
            # Issue severity distribution
            issues = self._count_issues_from_pages(pages)
            charts_data['issues_distribution'] = {
                'labels': ['Critical', 'Warning', 'Notice'],
                'data': [issues.get('critical', 0), issues.get('warning', 0), issues.get('notice', 0)]
            }
        
        return charts_data
    
    def _estimate_size(self, data: Any) -> int:
        """Estimate the size of data in bytes"""
        return len(json.dumps(data, default=str).encode('utf-8'))
    
    def _ensure_default_template(self) -> None:
        """Ensure default HTML template exists"""
        template_path = self.template_dir / 'report.html'
        if not template_path.exists():
            self._create_default_template()
    
    def _create_default_template(self) -> None:
        """Create default HTML template"""
        template_content = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        :root {
            --primary-color: #2c3e50;
            --secondary-color: #3498db;
            --success-color: #2ecc71;
            --warning-color: #f39c12;
            --danger-color: #e74c3c;
            --bg-color: #ecf0f1;
            --text-color: #2c3e50;
            --border-color: #bdc3c7;
        }
        
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            color: var(--text-color);
            background: var(--bg-color);
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
        }
        
        .header {
            background: var(--primary-color);
            color: white;
            padding: 40px;
            border-radius: 8px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        
        .header h1 {
            margin-bottom: 10px;
            font-size: 2.5em;
        }
        
        .header .meta {
            opacity: 0.9;
            font-size: 1.1em;
        }
        
        .summary-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }
        
        .card {
            background: white;
            padding: 25px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            transition: transform 0.2s, box-shadow 0.2s;
        }
        
        .card:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.15);
        }
        
        .card h3 {
            color: var(--primary-color);
            margin-bottom: 15px;
            font-size: 1.1em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        .metric {
            font-size: 2.5em;
            font-weight: bold;
            color: var(--secondary-color);
            margin-bottom: 5px;
        }
        
        .metric-label {
            color: #7f8c8d;
            font-size: 0.9em;
        }
        
        .issues-summary {
            display: flex;
            gap: 20px;
            margin: 20px 0;
            flex-wrap: wrap;
        }
        
        .issue-badge {
            padding: 12px 24px;
            border-radius: 5px;
            color: white;
            font-weight: bold;
            font-size: 1.1em;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        
        .critical { background: var(--danger-color); }
        .warning { background: var(--warning-color); }
        .notice { background: #95a5a6; }
        
        .score {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 4px;
            font-weight: bold;
            font-size: 0.9em;
        }
        
        .score-good { background: var(--success-color); color: white; }
        .score-medium { background: var(--warning-color); color: white; }
        .score-poor { background: var(--danger-color); color: white; }
        
        table {
            width: 100%;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 20px 0;
            border-collapse: collapse;
        }
        
        th, td {
            padding: 15px;
            text-align: left;
            border-bottom: 1px solid var(--border-color);
        }
        
        th {
            background: #34495e;
            color: white;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.9em;
            letter-spacing: 0.5px;
        }
        
        tr:hover {
            background: #f8f9fa;
        }
        
        tr:last-child td {
            border-bottom: none;
        }
        
        .chart-container {
            background: white;
            padding: 25px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 20px 0;
        }
        
        .chart-container h2 {
            margin-bottom: 20px;
            color: var(--primary-color);
        }
        
        .tabs {
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 30px 0;
            overflow: hidden;
        }
        
        .tab-buttons {
            display: flex;
            background: #f8f9fa;
            border-bottom: 2px solid var(--border-color);
        }
        
        .tab-button {
            padding: 15px 30px;
            background: none;
            border: none;
            cursor: pointer;
            font-size: 1em;
            font-weight: 500;
            color: #7f8c8d;
            transition: all 0.3s;
            position: relative;
        }
        
        .tab-button:hover {
            color: var(--primary-color);
            background: rgba(52, 152, 219, 0.1);
        }
        
        .tab-button.active {
            color: var(--secondary-color);
            background: white;
        }
        
        .tab-button.active::after {
            content: '';
            position: absolute;
            bottom: -2px;
            left: 0;
            right: 0;
            height: 2px;
            background: var(--secondary-color);
        }
        
        .tab-content {
            padding: 30px;
            display: none;
        }
        
        .tab-content.active {
            display: block;
        }
        
        .progress-bar {
            background: #e0e0e0;
            border-radius: 10px;
            height: 20px;
            overflow: hidden;
            margin: 10px 0;
        }
        
        .progress-fill {
            height: 100%;
            background: var(--success-color);
            transition: width 0.3s;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-size: 0.8em;
            font-weight: bold;
        }
        
        .recommendations {
            background: #f8f9fa;
            border-left: 4px solid var(--secondary-color);
            padding: 20px;
            margin: 20px 0;
            border-radius: 4px;
        }
        
        .recommendations h3 {
            color: var(--secondary-color);
            margin-bottom: 15px;
        }
        
        .recommendations ul {
            list-style: none;
            padding-left: 0;
        }
        
        .recommendations li {
            padding: 10px 0;
            border-bottom: 1px solid #e0e0e0;
        }
        
        .recommendations li:last-child {
            border-bottom: none;
        }
        
        .footer {
            text-align: center;
            padding: 40px;
            color: #7f8c8d;
            font-size: 0.9em;
        }
        
        @media (max-width: 768px) {
            .summary-grid {
                grid-template-columns: 1fr;
            }
            
            .tab-buttons {
                flex-direction: column;
            }
            
            .tab-button {
                width: 100%;
                text-align: left;
            }
            
            table {
                font-size: 0.9em;
            }
            
            th, td {
                padding: 10px;
            }
        }
        
        /* Print styles */
        @media print {
            body {
                background: white;
            }
            
            .card, table, .chart-container {
                box-shadow: none;
                border: 1px solid #ddd;
            }
            
            .tab-content {
                display: block !important;
                page-break-inside: avoid;
            }
        }
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{{ title }}</h1>
            <div class="meta">
                <p>Generated: {{ generated_at }}</p>
                <p>Website: {{ config.url }}</p>
                {% if is_single_page %}
                <p><strong>Single Page Analysis</strong></p>
                {% else %}
                <p>Total Pages Analyzed: {{ summary.total_pages }}</p>
                {% endif %}
            </div>
        </div>
        
        <div class="summary-grid">
            <div class="card">
                <h3>SEO Score</h3>
                <div class="metric">{{ "%.1f"|format(summary.average_seo_score) }}</div>
                <div class="metric-label">Average Score</div>
                <div class="progress-bar">
                    <div class="progress-fill" style="width: {{ summary.average_seo_score }}%; background: {% if summary.average_seo_score >= 80 %}var(--success-color){% elif summary.average_seo_score >= 60 %}var(--warning-color){% else %}var(--danger-color){% endif %}">
                        {{ "%.1f"|format(summary.average_seo_score) }}%
                    </div>
                </div>
            </div>
            
            <div class="card">
                <h3>Performance</h3>
                <div class="metric">{{ "%.2f"|format(summary.average_load_time) }}s</div>
                <div class="metric-label">Average Load Time</div>
                {% if summary.performance_summary %}
                <div style="margin-top: 15px;">
                    <small>Fast: {{ summary.performance_summary.fast }} | Moderate: {{ summary.performance_summary.moderate }} | Slow: {{ summary.performance_summary.slow }}</small>
                </div>
                {% endif %}
            </div>
            
            <div class="card">
                <h3>Content</h3>
                <div class="metric">{{ summary.average_word_count }}</div>
                <div class="metric-label">Average Word Count</div>
                <div style="margin-top: 15px;">
                    <small>Thin Content Pages: {{ summary.thin_content_pages }}</small>
                </div>
            </div>
            
            <div class="card">
                <h3>Technical</h3>
                <div style="margin-top: 10px;">
                    <p><strong>HTTPS:</strong> {{ summary.https_pages }}/{{ summary.total_pages }}</p>
                    <p><strong>Mobile Friendly:</strong> {{ summary.mobile_friendly_pages }}/{{ summary.total_pages }}</p>
                    <p><strong>Missing Titles:</strong> {{ summary.pages_missing_title }}</p>
                    <p><strong>Missing Descriptions:</strong> {{ summary.pages_missing_description }}</p>
                </div>
            </div>
        </div>
        
        <div class="card">
            <h3>Issues Summary</h3>
            <div class="issues-summary">
                <div class="issue-badge critical">
                    <span style="font-size: 1.5em;">⚠️</span>
                    {{ issues.critical or issues.get('critical', 0) }} Critical
                </div>
                <div class="issue-badge warning">
                    <span style="font-size: 1.5em;">⚡</span>
                    {{ issues.warnings or issues.get('warning', 0) }} Warnings
                </div>
                <div class="issue-badge notice">
                    <span style="font-size: 1.5em;">ℹ️</span>
                    {{ issues.notices or issues.get('notice', 0) }} Notices
                </div>
            </div>
        </div>
        
        {% if charts_data %}
        <div class="chart-container">
            <h2>SEO Score Distribution</h2>
            <canvas id="scoreChart"></canvas>
        </div>
        
        <div class="chart-container">
            <h2>Performance Distribution</h2>
            <canvas id="performanceChart"></canvas>
        </div>
        {% endif %}
        
        <div class="tabs">
            <div class="tab-buttons">
                <button class="tab-button active" onclick="showTab('pages')">Pages Analysis</button>
                <button class="tab-button" onclick="showTab('issues')">All Issues</button>
                {% if summary.top_issues %}
                <button class="tab-button" onclick="showTab('top-issues')">Top Issues</button>
                {% endif %}
                <button class="tab-button" onclick="showTab('recommendations')">Recommendations</button>
            </div>
            
            <div id="pages" class="tab-content active">
                <h2>Pages Analysis</h2>
                <div style="overflow-x: auto;">
                    <table>
                        <thead>
                            <tr>
                                <th>URL</th>
                                <th>Title</th>
                                <th>Load Time</th>
                                <th>Word Count</th>
                                <th>Score</th>
                                <th>Issues</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for page in pages[:50] %}
                            <tr>
                                <td><a href="{{ page.get('url', '') }}" target="_blank" style="color: var(--secondary-color); text-decoration: none;">{{ page.get('url', '')[:60] }}{% if page.get('url', '')|length > 60 %}...{% endif %}</a></td>
                                <td>{{ page.get('title', '')[:40] }}{% if page.get('title', '')|length > 40 %}...{% endif %}</td>
                                <td>{{ "%.2f"|format(page.get('load_time', 0)) }}s</td>
                                <td>{{ page.get('content', {}).get('word_count', 0) }}</td>
                                <td>
                                    <span class="score {% if page.get('score', 0) >= 80 %}score-good{% elif page.get('score', 0) >= 60 %}score-medium{% else %}score-poor{% endif %}">
                                        {{ "%.1f"|format(page.get('score', 0)) }}
                                    </span>
                                </td>
                                <td>{{ page.get('issues', [])|length }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                    {% if pages|length > 50 %}
                    <p style="text-align: center; margin-top: 20px; color: #7f8c8d;">
                        Showing first 50 of {{ pages|length }} pages
                    </p>
                    {% endif %}
                </div>
            </div>
            
            <div id="issues" class="tab-content">
                <h2>All Issues</h2>
                <div style="overflow-x: auto;">
                    <table>
                        <thead>
                            <tr>
                                <th>Page</th>
                                <th>Type</th>
                                <th>Severity</th>
                                <th>Message</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for page in pages %}
                                {% for issue in page.get('issues', [])[:5] %}
                                <tr>
                                    <td><a href="{{ page.get('url', '') }}" target="_blank" style="color: var(--secondary-color); text-decoration: none;">{{ page.get('url', '')[:40] }}...</a></td>
                                    <td>{{ issue.get('type', '') }}</td>
                                    <td><span class="issue-badge {{ issue.get('severity', '') }}" style="font-size: 0.9em; padding: 4px 8px;">{{ issue.get('severity', '') }}</span></td>
                                    <td>{{ issue.get('message', '')[:80] }}{% if issue.get('message', '')|length > 80 %}...{% endif %}</td>
                                </tr>
                                {% endfor %}
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
            
            {% if summary.top_issues %}
            <div id="top-issues" class="tab-content">
                <h2>Top Issues</h2>
                <div style="overflow-x: auto;">
                    <table>
                        <thead>
                            <tr>
                                <th>Issue Type</th>
                                <th>Message</th>
                                <th>Severity</th>
                                <th>Occurrences</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for issue in summary.top_issues %}
                            <tr>
                                <td>{{ issue.get('type', '') }}</td>
                                <td>{{ issue.get('message', '') }}</td>
                                <td><span class="issue-badge {{ issue.get('severity', '') }}" style="font-size: 0.9em; padding: 4px 8px;">{{ issue.get('severity', '') }}</span></td>
                                <td>{{ issue.get('count', 0) }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
            {% endif %}
            
            <div id="recommendations" class="tab-content">
                <h2>Recommendations</h2>
                <div class="recommendations">
                    <h3>🎯 High Priority</h3>
                    <ul>
                        {% if summary.pages_missing_title > 0 %}
                        <li>Add unique, descriptive title tags to {{ summary.pages_missing_title }} pages</li>
                        {% endif %}
                        {% if summary.pages_missing_description > 0 %}
                        <li>Add compelling meta descriptions to {{ summary.pages_missing_description }} pages</li>
                        {% endif %}
                        {% if summary.https_pages < summary.total_pages %}
                        <li>Migrate {{ summary.total_pages - summary.https_pages }} pages to HTTPS for security</li>
                        {% endif %}
                        {% if summary.slow_loading_pages > 0 %}
                        <li>Optimize {{ summary.slow_loading_pages }} slow-loading pages (>3s load time)</li>
                        {% endif %}
                        {% if summary.thin_content_pages > 0 %}
                        <li>Expand content on {{ summary.thin_content_pages }} pages with less than 300 words</li>
                        {% endif %}
                    </ul>
                </div>
                
                <div class="recommendations">
                    <h3>💡 Medium Priority</h3>
                    <ul>
                        {% if summary.mobile_friendly_pages < summary.total_pages %}
                        <li>Improve mobile experience on {{ summary.total_pages - summary.mobile_friendly_pages }} pages</li>
                        {% endif %}
                        {% if issues.warning > 10 %}
                        <li>Address {{ issues.warning }} warning-level issues across the site</li>
                        {% endif %}
                        <li>Implement structured data markup for better search visibility</li>
                        <li>Optimize images with proper compression and formats</li>
                    </ul>
                </div>
                
                <div class="recommendations">
                    <h3>📋 Best Practices</h3>
                    <ul>
                        <li>Maintain consistent internal linking structure</li>
                        <li>Use descriptive anchor text for all links</li>
                        <li>Implement proper heading hierarchy (H1-H6)</li>
                        <li>Add alt text to all images for accessibility</li>
                        <li>Monitor and fix broken links regularly</li>
                    </ul>
                </div>
            </div>
        </div>
        
        <div class="footer">
            <p>Generated by tfq0seo - Professional SEO Analysis Toolkit</p>
            <p>Report created on {{ generated_at }}</p>
        </div>
    </div>
    
    <script>
        // Tab functionality
        function showTab(tabName) {
            const tabs = document.querySelectorAll('.tab-content');
            const buttons = document.querySelectorAll('.tab-button');
            
            tabs.forEach(tab => {
                tab.classList.remove('active');
            });
            
            buttons.forEach(button => {
                button.classList.remove('active');
            });
            
            document.getElementById(tabName).classList.add('active');
            event.target.classList.add('active');
        }
        
        {% if charts_data %}
        // Chart.js configuration
        const chartOptions = {
            responsive: true,
            maintainAspectRatio: false,
            plugins: {
                legend: {
                    position: 'bottom',
                    labels: {
                        padding: 20,
                        font: {
                            size: 14
                        }
                    }
                }
            }
        };
        
        // Score distribution chart
        {% if charts_data.score_distribution %}
        const scoreCtx = document.getElementById('scoreChart').getContext('2d');
        new Chart(scoreCtx, {
            type: 'bar',
            data: {
                labels: {{ charts_data.score_distribution.labels | tojson }},
                datasets: [{
                    label: 'Number of Pages',
                    data: {{ charts_data.score_distribution.data | tojson }},
                    backgroundColor: [
                        '#e74c3c',
                        '#e67e22',
                        '#f39c12',
                        '#f1c40f',
                        '#2ecc71'
                    ],
                    borderWidth: 0
                }]
            },
            options: {
                ...chartOptions,
                scales: {
                    y: {
                        beginAtZero: true,
                        ticks: {
                            stepSize: 1
                        }
                    }
                }
            }
        });
        {% endif %}
        
        // Performance distribution chart
        {% if charts_data.performance_distribution %}
        const perfCtx = document.getElementById('performanceChart').getContext('2d');
        new Chart(perfCtx, {
            type: 'doughnut',
            data: {
                labels: {{ charts_data.performance_distribution.labels | tojson }},
                datasets: [{
                    data: {{ charts_data.performance_distribution.data | tojson }},
                    backgroundColor: [
                        '#2ecc71',
                        '#f39c12',
                        '#e74c3c'
                    ],
                    borderWidth: 0
                }]
            },
            options: chartOptions
        });
        {% endif %}
        {% endif %}
    </script>
</body>
</html>'''
        
        template_path = self.template_dir / 'report.html'
        template_path.write_text(template_content) 