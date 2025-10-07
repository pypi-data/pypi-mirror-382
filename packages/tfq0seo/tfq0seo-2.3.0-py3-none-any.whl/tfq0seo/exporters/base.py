"""Export manager for various output formats - simple and direct implementation."""

import json
import csv
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from jinja2 import Template, FileSystemLoader, Environment


class ExportManager:
    """Manages export of analysis results to various formats."""
    
    def __init__(self, config: Optional[Any] = None):
        """Initialize export manager with configuration."""
        # Handle both dict and dataclass config objects
        if hasattr(config, '__dict__'):
            # It's a dataclass or object, convert to dict
            self.config = config.__dict__ if config else {}
        else:
            # It's already a dict or None
            self.config = config or {}
            
        self.output_dir = Path(self.config.get('output_directory', './reports'))
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Setup Jinja2 for HTML templates
        template_dir = Path(__file__).parent.parent / 'templates'
        self.jinja_env = Environment(loader=FileSystemLoader(str(template_dir)))
    
    def export(self, data: Dict[str, Any], format: str, output_path: Optional[str] = None) -> str:
        """Export data to specified format."""
        if output_path:
            output_file = Path(output_path)
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"seo_report_{timestamp}.{format}"
            output_file = self.output_dir / filename
        
        # Ensure parent directory exists
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        if format == 'json':
            return self.export_json(data, output_file)
        elif format == 'csv':
            return self.export_csv(data, output_file)
        elif format == 'xlsx':
            return self.export_xlsx(data, output_file)
        elif format == 'html':
            return self.export_html(data, output_file)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def export_json(self, data: Dict[str, Any], output_file: Path) -> str:
        """Export data as JSON."""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str, ensure_ascii=False)
        return str(output_file)
    
    def export_csv(self, data: Dict[str, Any], output_file: Path) -> str:
        """Export data as CSV."""
        # Flatten the data for CSV export
        rows = []
        
        if 'pages' in data:
            # Multi-page report
            for page in data['pages']:
                row = self._flatten_page_data(page)
                rows.append(row)
        else:
            # Single page report
            row = self._flatten_page_data(data)
            rows.append(row)
        
        if not rows:
            rows = [{'error': 'No data to export'}]
        
        # Write CSV
        fieldnames = set()
        for row in rows:
            fieldnames.update(row.keys())
        fieldnames = sorted(list(fieldnames))
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        return str(output_file)
    
    def export_xlsx(self, data: Dict[str, Any], output_file: Path) -> str:
        """Export data as Excel file."""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl not available
            output_file = output_file.with_suffix('.csv')
            return self.export_csv(data, output_file)
        
        wb = Workbook()
        
        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        self._write_summary_sheet(summary_sheet, data)
        
        # Issues sheet
        if 'issues' in data or ('pages' in data and data['pages']):
            issues_sheet = wb.create_sheet("Issues")
            self._write_issues_sheet(issues_sheet, data)
        
        # Pages sheet (if multi-page report)
        if 'pages' in data:
            pages_sheet = wb.create_sheet("Pages")
            self._write_pages_sheet(pages_sheet, data['pages'])
        
        # Recommendations sheet
        if 'recommendations' in data:
            rec_sheet = wb.create_sheet("Recommendations")
            self._write_recommendations_sheet(rec_sheet, data['recommendations'])
        
        wb.save(output_file)
        return str(output_file)
    
    def export_html(self, data: Dict[str, Any], output_file: Path) -> str:
        """Export data as HTML report."""
        # Select template based on data characteristics
        template_name = 'report.html'
        
        # Use optimized template if we have aggregated data or multiple pages
        if ('aggregated_issues' in data or 'executive_summary' in data or 
            'enhanced_recommendations' in data or 
            (data.get('summary', {}).get('total_pages', 0) > 1)):
            template_name = 'optimized_report.html'
        elif self.config.get('html_template') == 'enhanced':
            template_name = 'enhanced_report.html'
        
        try:
            template = self.jinja_env.get_template(template_name)
        except:
            # If template not found, try alternative or use basic
            try:
                template = self.jinja_env.get_template('report.html')
            except:
                template = Template(self._get_basic_html_template())
        
        # Prepare data for template
        template_data = self._prepare_html_data(data)
        
        # Render HTML
        html_content = template.render(**template_data)
        
        # Write to file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return str(output_file)
    
    def _flatten_page_data(self, page_data: Dict[str, Any]) -> Dict[str, Any]:
        """Flatten nested page data for CSV export."""
        flat = {
            'url': page_data.get('url', ''),
            'overall_score': page_data.get('overall_score', 0),
            'status_code': page_data.get('status_code', 0),
            'load_time': page_data.get('load_time', 0),
        }
        
        # Add scores from each analyzer
        for analyzer in ['seo', 'content', 'technical', 'performance', 'links']:
            if analyzer in page_data:
                flat[f'{analyzer}_score'] = page_data[analyzer].get('score', 0)
        
        # Add issue counts
        if 'issue_counts' in page_data:
            flat.update({f'issues_{k}': v for k, v in page_data['issue_counts'].items()})
        
        # Add some key data points
        if 'seo' in page_data and 'data' in page_data['seo']:
            seo_data = page_data['seo']['data']
            flat['title'] = seo_data.get('title', '')[:100]
            flat['description'] = seo_data.get('description', '')[:160]
        
        if 'content' in page_data and 'data' in page_data['content']:
            content_data = page_data['content']['data']
            flat['word_count'] = content_data.get('word_count', 0)
            flat['flesch_reading_ease'] = content_data.get('flesch_reading_ease', 0)
        
        if 'performance' in page_data and 'data' in page_data['performance']:
            perf_data = page_data['performance']['data']
            flat['total_resources'] = perf_data.get('total_resources', 0)
            flat['content_size_mb'] = perf_data.get('content_size_mb', 0)
        
        return flat
    
    def _write_summary_sheet(self, sheet, data: Dict[str, Any]):
        """Write summary data to Excel sheet."""
        # Title
        sheet['A1'] = 'SEO Analysis Report'
        sheet['A1'].font = Font(size=16, bold=True)
        
        row = 3
        
        # Overall score
        sheet[f'A{row}'] = 'Overall Score'
        sheet[f'B{row}'] = data.get('overall_score', 0)
        sheet[f'B{row}'].font = Font(bold=True, color='00FF00' if data.get('overall_score', 0) >= 80 else 'FF0000')
        row += 1
        
        # Category scores
        if 'category_scores' in data:
            sheet[f'A{row}'] = 'Category Scores'
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for category, score in data['category_scores'].items():
                sheet[f'A{row}'] = category.title()
                sheet[f'B{row}'] = score
                row += 1
        
        row += 1
        
        # Issue counts
        if 'issue_counts' in data:
            sheet[f'A{row}'] = 'Issue Summary'
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            sheet[f'A{row}'] = 'Critical'
            sheet[f'B{row}'] = data['issue_counts'].get('critical', 0)
            sheet[f'B{row}'].font = Font(color='FF0000')
            row += 1
            
            sheet[f'A{row}'] = 'Warnings'
            sheet[f'B{row}'] = data['issue_counts'].get('warning', 0)
            sheet[f'B{row}'].font = Font(color='FFA500')
            row += 1
            
            sheet[f'A{row}'] = 'Notices'
            sheet[f'B{row}'] = data['issue_counts'].get('notice', 0)
            sheet[f'B{row}'].font = Font(color='0000FF')
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _write_issues_sheet(self, sheet, data: Dict[str, Any]):
        """Write issues to Excel sheet."""
        # Headers
        headers = ['Severity', 'Category', 'Message', 'URL']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Collect all issues
        all_issues = []
        if 'issues' in data:
            for issue in data['issues']:
                all_issues.append({
                    'severity': issue.get('severity', ''),
                    'category': issue.get('category', ''),
                    'message': issue.get('message', ''),
                    'url': data.get('url', '')
                })
        elif 'pages' in data:
            for page in data['pages']:
                if 'issues' in page:
                    for issue in page['issues']:
                        all_issues.append({
                            'severity': issue.get('severity', ''),
                            'category': issue.get('category', ''),
                            'message': issue.get('message', ''),
                            'url': page.get('url', '')
                        })
        
        # Sort by severity
        severity_order = {'critical': 0, 'warning': 1, 'notice': 2}
        all_issues.sort(key=lambda x: severity_order.get(x['severity'], 3))
        
        # Write issues
        for row_num, issue in enumerate(all_issues, 2):
            sheet.cell(row=row_num, column=1, value=issue['severity'])
            sheet.cell(row=row_num, column=2, value=issue['category'])
            sheet.cell(row=row_num, column=3, value=issue['message'])
            sheet.cell(row=row_num, column=4, value=issue['url'])
            
            # Color code severity
            severity_cell = sheet.cell(row=row_num, column=1)
            if issue['severity'] == 'critical':
                severity_cell.font = Font(color='FF0000')
            elif issue['severity'] == 'warning':
                severity_cell.font = Font(color='FFA500')
            elif issue['severity'] == 'notice':
                severity_cell.font = Font(color='0000FF')
    
    def _write_pages_sheet(self, sheet, pages: List[Dict[str, Any]]):
        """Write pages data to Excel sheet."""
        if not pages:
            return
        
        # Headers
        headers = ['URL', 'Overall Score', 'Status', 'Load Time', 'SEO Score', 
                  'Content Score', 'Technical Score', 'Performance Score', 'Links Score',
                  'Critical Issues', 'Warnings', 'Notices']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Write page data
        for row_num, page in enumerate(pages, 2):
            sheet.cell(row=row_num, column=1, value=page.get('url', ''))
            sheet.cell(row=row_num, column=2, value=page.get('overall_score', 0))
            sheet.cell(row=row_num, column=3, value=page.get('status_code', 0))
            sheet.cell(row=row_num, column=4, value=page.get('load_time', 0))
            
            # Analyzer scores
            col = 5
            for analyzer in ['seo', 'content', 'technical', 'performance', 'links']:
                score = page.get(analyzer, {}).get('score', 0) if analyzer in page else 0
                sheet.cell(row=row_num, column=col, value=score)
                col += 1
            
            # Issue counts
            issue_counts = page.get('issue_counts', {})
            sheet.cell(row=row_num, column=10, value=issue_counts.get('critical', 0))
            sheet.cell(row=row_num, column=11, value=issue_counts.get('warning', 0))
            sheet.cell(row=row_num, column=12, value=issue_counts.get('notice', 0))
    
    def _write_recommendations_sheet(self, sheet, recommendations: List[str]):
        """Write recommendations to Excel sheet."""
        sheet['A1'] = 'Recommendations'
        sheet['A1'].font = Font(size=14, bold=True)
        
        for row_num, rec in enumerate(recommendations, 3):
            sheet[f'A{row_num}'] = f'{row_num - 2}.'
            sheet[f'B{row_num}'] = rec
    
    def _prepare_html_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for HTML template."""
        # Ensure all required fields exist
        prepared = {
            'url': data.get('url', 'Multiple Pages'),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'overall_score': data.get('overall_score', 0),
            'category_scores': data.get('category_scores', {}),
            'issues': data.get('issues', []),
            'issue_counts': data.get('issue_counts', {}),
            'recommendations': data.get('recommendations', []),
            'pages': data.get('pages', []),
        }
        
        # Handle nested structures from new report format
        # Check for scores nested structure
        if 'scores' in data:
            prepared['overall_score'] = data['scores'].get('overall', 0)
            prepared['category_scores'] = data['scores'].get('categories', {})
        
        # Check for issues nested structure
        if 'issues' in data and isinstance(data['issues'], dict):
            if 'counts' in data['issues']:
                prepared['issue_counts'] = data['issues']['counts']
        
        # Ensure scores exist
        if not prepared['category_scores']:
            prepared['category_scores'] = {
                'seo': 0,
                'content': 0,
                'technical': 0,
                'performance': 0,
                'links': 0
            }
        
        # Add new optimized data fields with defaults
        if 'aggregated_issues' in data:
            prepared['aggregated_issues'] = data['aggregated_issues']
        else:
            prepared['aggregated_issues'] = []
        
        if 'aggregation_stats' in data:
            prepared['aggregation_stats'] = data['aggregation_stats']
        else:
            prepared['aggregation_stats'] = {}
        
        # Handle recommendations - check for new format
        if 'recommendations' in data and isinstance(data['recommendations'], dict):
            if 'specific' in data['recommendations']:
                prepared['enhanced_recommendations'] = data['recommendations']['specific']
            if 'executive' in data['recommendations']:
                prepared['executive_summary'] = data['recommendations']['executive']
        else:
            # Try direct fields
            if 'enhanced_recommendations' in data:
                prepared['enhanced_recommendations'] = data['enhanced_recommendations']
            else:
                prepared['enhanced_recommendations'] = []
            
            if 'executive_summary' in data:
                prepared['executive_summary'] = data['executive_summary']
            else:
                # Create a default executive summary
                prepared['executive_summary'] = {
                    'overview': {
                        'total_pages_analyzed': data.get('summary', {}).get('total_pages', 1) if 'summary' in data else 1,
                        'overall_health_score': data.get('overall_score', 0)
                    },
                    'key_metrics': {
                        'average_score': data.get('overall_score', 0),
                        'critical_issues': data.get('issue_counts', {}).get('critical', 0),
                        'warnings': data.get('issue_counts', {}).get('warning', 0),
                        'notices': data.get('issue_counts', {}).get('notice', 0)
                    },
                    'top_issues': [],
                    'quick_wins': []
                }
        
        if 'performance_metrics' in data:
            prepared['performance_metrics'] = data['performance_metrics']
        else:
            prepared['performance_metrics'] = {}
        
        # Handle pages data
        if 'pages' in data and isinstance(data['pages'], dict):
            # New format with nested structure
            if 'summary' in data['pages']:
                prepared['pages_summary'] = data['pages']['summary']
            if 'detailed' in data['pages']:
                prepared['pages'] = data['pages']['detailed']
        else:
            # Old format compatibility
            if 'pages_summary' in data:
                prepared['pages_summary'] = data['pages_summary']
            else:
                prepared['pages_summary'] = []
        
        if 'pages_truncated' in data:
            prepared['pages_truncated'] = data['pages_truncated']
            prepared['pages_truncated_count'] = data.get('pages_truncated_count', 0)
        else:
            prepared['pages_truncated'] = False
            prepared['pages_truncated_count'] = 0
        
        if 'top_issues' in data:
            prepared['top_issues'] = data['top_issues']
        elif 'issues' in data and isinstance(data['issues'], dict) and 'top_issues' in data['issues']:
            prepared['top_issues'] = data['issues']['top_issues']
        else:
            prepared['top_issues'] = []
        
        # Add performance data
        if 'performance' in data:
            prepared['performance_data'] = data['performance']
        
        # Add summary if multi-page
        if 'summary' in data:
            prepared['summary'] = data['summary']
        
        # Handle issues - check if it's a dict (new format) or list (old format)
        severity_order = {'critical': 0, 'warning': 1, 'notice': 2}
        
        if isinstance(prepared['issues'], dict):
            # New format: issues is a dict with 'aggregated' key
            if 'aggregated' in prepared['issues']:
                issue_list = prepared['issues']['aggregated']
            elif 'top_issues' in prepared['issues']:
                issue_list = prepared['issues']['top_issues']
            else:
                issue_list = []
            
            # Update prepared['issues'] to be the list for template compatibility
            prepared['issues'] = issue_list
        else:
            # Old format: issues is already a list
            issue_list = prepared['issues']
        
        # Sort issues by severity
        if issue_list and isinstance(issue_list, list):
            issue_list.sort(key=lambda x: severity_order.get(x.get('severity', 'notice'), 3))
            prepared['issues'] = issue_list
        
        # Group issues by severity
        prepared['issues_by_severity'] = {
            'critical': [i for i in prepared['issues'] if i.get('severity') == 'critical'],
            'warning': [i for i in prepared['issues'] if i.get('severity') == 'warning'],
            'notice': [i for i in prepared['issues'] if i.get('severity') == 'notice']
        }
        
        return prepared
    
    def _get_basic_html_template(self) -> str:
        """Return a basic HTML template if template files are not available."""
        return '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SEO Analysis Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; }
        h1 { color: #333; }
        .score { font-size: 48px; font-weight: bold; }
        .score.good { color: #0cce6b; }
        .score.medium { color: #ffa400; }
        .score.poor { color: #ff4e42; }
        .issues { margin-top: 30px; }
        .issue { padding: 10px; margin: 10px 0; border-left: 4px solid; }
        .issue.critical { border-color: #ff4e42; background: #fff5f5; }
        .issue.warning { border-color: #ffa400; background: #fffaf0; }
        .issue.notice { border-color: #4285f4; background: #f0f8ff; }
    </style>
</head>
<body>
    <div class="container">
        <h1>SEO Analysis Report</h1>
        <p>URL: {{ url }}</p>
        <p>Generated: {{ timestamp }}</p>
        
        <h2>Overall Score</h2>
        <div class="score {% if overall_score >= 80 %}good{% elif overall_score >= 50 %}medium{% else %}poor{% endif %}">
            {{ overall_score|round|int }}/100
        </div>
        
        <h2>Issues ({{ issues|length }})</h2>
        <div class="issues">
            {% for issue in issues %}
            <div class="issue {{ issue.severity }}">
                <strong>{{ issue.category }}</strong>: {{ issue.message }}
            </div>
            {% endfor %}
        </div>
        
        <h2>Recommendations</h2>
        <ol>
            {% for rec in recommendations %}
            <li>{{ rec }}</li>
            {% endfor %}
        </ol>
    </div>
</body>
</html>'''
