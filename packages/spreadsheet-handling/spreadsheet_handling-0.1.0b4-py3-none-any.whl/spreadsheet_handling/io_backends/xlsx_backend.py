from __future__ import annotations

from pathlib import Path
from typing import Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Keep the base import so existing imports/typing remain valid.
from .base import BackendBase, BackendOptions


# ======================================================================================
# Styling helpers
# ======================================================================================

def _decorate_workbook(
        workbook_path: Path,
        *,
        auto_filter: bool = True,
        header_fill_rgb: str = "DDDDDD",
        freeze_header: bool = False,
) -> None:
    """
    Post-process a written XLSX file:
    - apply AutoFilter across the used range
    - color header row (row 1) with a light gray fill & bold font
    - optionally freeze the first row (pane below header)
    """
    wb = load_workbook(workbook_path)

    header_fill = PatternFill("solid", fgColor=header_fill_rgb)
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        # AutoFilter across used range
        if auto_filter and ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions  # e.g. "A1:D100"

        # Header styling (row 1)
        if ws.max_column:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font

        # Freeze pane below header
        if freeze_header:
            ws.freeze_panes = "A2"

    wb.save(workbook_path)


def _flatten_header_to_level0(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure we write a single header row:
    - if columns is a MultiIndex, take level 0
    - otherwise keep as-is
    """
    if isinstance(df.columns, pd.MultiIndex):
        df = df.copy()
        df.columns = [t[0] for t in df.columns.to_list()]
    return df


# ======================================================================================
# Excel backend (multi-sheet only)
# ======================================================================================

class ExcelBackend(BackendBase):
    """
    XLSX adapter backed by pandas + openpyxl.

    Public API (used by our router/tests):
      - write_multi(frames, path, options=None)
      - read_multi(path, header_levels, options=None)
    """

    def write_multi(
            self,
            frames: Dict[str, pd.DataFrame],
            path: str,
            options: BackendOptions | None = None,
    ) -> None:
        """
        Write multiple DataFrames to an XLSX:
        - one sheet per dict key
        - flatten MultiIndex headers to level 0 (single header row)
        - apply header styling (autofilter + gray + bold), optionally freeze header
        """
        out = Path(path).with_suffix(".xlsx")
        out.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(out, engine="openpyxl") as xw:
            for sheet_name, df in frames.items():
                sheet = (sheet_name or "Sheet")[:31]  # Excel sheet name limit
                df_out = _flatten_header_to_level0(df)
                df_out.to_excel(xw, sheet_name=sheet, index=False)

        # If you later extend BackendOptions with excel-related knobs,
        # you can thread them here. For now use reasonable defaults:
        _decorate_workbook(
            out,
            auto_filter=True,
            header_fill_rgb="DDDDDD",
            freeze_header=False,
        )

    def read_multi(
            self,
            path: str,
            header_levels: int,
            options: BackendOptions | None = None,
    ) -> Dict[str, pd.DataFrame]:
        """
        Read all sheets assuming a single header row (header=0).
        If header_levels > 1, lift columns to a MultiIndex (level-0 data,
        remaining levels empty strings).
        """
        p = Path(path)
        sheets = pd.read_excel(
            p, sheet_name=None, header=0, engine="openpyxl", dtype=str
        )
        out: Dict[str, pd.DataFrame] = {}
        levels = header_levels if (header_levels and header_levels > 1) else 1

        for name, df in sheets.items():
            df = df.where(pd.notnull(df), "")  # normalize NaNs to empty strings
            if not isinstance(df.columns, pd.MultiIndex) and levels > 1:
                tuples = [(c,) + ("",) * (levels - 1) for c in list(df.columns)]
                df = df.copy()
                df.columns = pd.MultiIndex.from_tuples(tuples)
            out[name] = df

        return out

# ------------------------------------------------------------------------------
# Legacy/test convenience shim
# ------------------------------------------------------------------------------

def write_xlsx(
        path: str,
        frames: Dict[str, pd.DataFrame],
        meta: Any,  # kept for signature compatibility; not used here
        ctx: Any,   # expected to hold ctx.app.excel.{auto_filter, header_fill_rgb, freeze_header}
) -> None:
    """
    Test-facing convenience used by unit tests:
    - writes XLSX with one sheet per frame (flattening header to level 0)
    - applies AutoFilter + gray/bold header
    - honors ctx.app.excel options if present
    """
    out = Path(path).with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    # 1) write via pandas (single header row)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        for sheet_name, df in frames.items():
            sheet = (sheet_name or "Sheet")[:31]
            df_out = _flatten_header_to_level0(df)
            df_out.to_excel(xw, sheet_name=sheet, index=False)

    # 2) style according to ctx
    excel_opts = getattr(getattr(ctx, "app", None), "excel", None)
    _decorate_workbook(
        out,
        auto_filter=getattr(excel_opts, "auto_filter", True) if excel_opts else True,
        header_fill_rgb=getattr(excel_opts, "header_fill_rgb", "DDDDDD") if excel_opts else "DDDDDD",
        freeze_header=getattr(excel_opts, "freeze_header", False) if excel_opts else False,
    )


# ======================================================================================
# Test-/router-facing convenience (module-level) API
# ======================================================================================

# src/spreadsheet_handling/io_backends/xlsx_backend.py
from typing import Dict, Any
import pandas as pd
from .base import BackendOptions
from .xlsx_backend import ExcelBackend  # falls die Klasse hier liegt; sonst anpassen

def _flatten_cols_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure single-level string headers for Excel output."""
    if not isinstance(df.columns, pd.MultiIndex):
        # already flat; also guard against tuple-like strings
        df.columns = [str(c) for c in df.columns]
        return df

    def first_nonempty(tup) -> str:
        for x in tup:
            s = str(x)
            if s:
                return s
        return ""

    flat = [first_nonempty(t) for t in df.columns.tolist()]
    new_df = df.copy()
    new_df.columns = flat
    return new_df

def save_xlsx(
        frames: Dict[str, pd.DataFrame],
        path: str,
        options: BackendOptions | None = None
) -> None:
    """Router-facing saver that guarantees flat string headers."""
    sanitized: Dict[str, pd.DataFrame] = {}
    for name, df in frames.items():
        sanitized[name] = _flatten_cols_for_excel(df)
    ExcelBackend().write_multi(sanitized, path, options=options)

def load_xlsx(
        path: str,
        options: BackendOptions | None = None   # <-- NEU
) -> dict[str, pd.DataFrame]:
    """
    Router-facing reader: read all sheets, assume single header row.
    (Lifts to MultiIndex of length 1 → effectively stays flat.)
    """
    return ExcelBackend().read_multi(path, header_levels=1, options=options)


__all__ = [
    "ExcelBackend",
    "save_xlsx",
    "load_xlsx",
]
