# tests/unit/io_backends/test_xlsx_writer_styling.py
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from spreadsheet_handling.pipeline.types import MetaDict, Context
from spreadsheet_handling.pipeline.config import AppConfig, IOConfig, IOEndpoint, PipelineConfig, ExcelOptions
from spreadsheet_handling.io_backends.xlsx_backend import write_xlsx

def _ctx(tmp_path: Path) -> Context:
    app = AppConfig(
        io=IOConfig(inputs={"primary": IOEndpoint(kind="json", path="unused")}, output=IOEndpoint(kind="xlsx", path=str(tmp_path/"out.xlsx"))),
        pipeline=PipelineConfig(steps=[]),
        excel=ExcelOptions(auto_filter=True, header_fill_rgb="DDDDDD", freeze_header=False, helper_fill_rgb="FFF5CC"),
        strict=True,
    )
    return Context(app=app)

def test_xlsx_header_and_autofilter(tmp_path: Path):
    frames = {"T": pd.DataFrame([{"a":"1", "b":"2"}, {"a":"3", "b":"4"}])}
    ctx = _ctx(tmp_path)
    meta = MetaDict()
    out = tmp_path/"out.xlsx"
    write_xlsx(str(out), frames, meta, ctx)

    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value == "a"
    assert ws["B1"].value == "b"
    assert ws.auto_filter is not None
